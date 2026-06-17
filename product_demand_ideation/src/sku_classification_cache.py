from __future__ import annotations

import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEET_NAME = "PowerBI Families"

CLASSIFICATION_COLUMNS = [
    "sku",
    "family",
    "category",
    "pm_responsible",
    "series",
    "source_row",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def classification_cache_path(root: Path) -> Path:
    return root / "product_demand_ideation" / "cache" / "sku_classification.sqlite"


def default_classification_workbook_path(root: Path) -> Path:
    return root / "product_demand_ideation" / "source_workbooks" / "Sku's Classification.xlsx"


def _text(value: Any) -> str:
    return str(value or "").strip()


def _normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _text(value).lower())


def _normalized_sku(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value).upper())


def _family_from_sku(value: str) -> str:
    sku = _normalized_sku(value)
    sku = re.sub(r"-SKID$", "", sku)
    sku = re.sub(r"-\d+PK$", "", sku)
    return sku


def _find_header_row(sheet: Any) -> tuple[int, dict[str, int]]:
    aliases = {
        "sku": {"sku", "mastersku", "master sku", "master_sku"},
        "family": {"family", "skufamily", "sku family"},
        "category": {"category", "categoryname", "assignedcategory", "pbi category", "powerbicategory", "high level category", "highlevelcategory"},
        "pm_responsible": {"pm", "pmresponsible", "productmanager", "owner", "categorypm"},
        "series": {"series", "seriesname", "productseries"},
    }
    normalized_aliases = {
        target: {_normalized_header(alias) for alias in values}
        for target, values in aliases.items()
    }
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True), start=1):
        header_map: dict[str, int] = {}
        for index, value in enumerate(row):
            normalized = _normalized_header(value)
            for target, candidates in normalized_aliases.items():
                if normalized in candidates and target not in header_map:
                    header_map[target] = index
        if "sku" in header_map and ("category" in header_map or "pm_responsible" in header_map or "series" in header_map):
            return row_number, header_map
    raise ValueError(f"Could not find a usable header row on sheet {SHEET_NAME!r}. Expected at least SKU plus category, PM, or series.")


def _row_value(row: tuple[Any, ...], header_map: dict[str, int], key: str) -> str:
    index = header_map.get(key)
    if index is None or index >= len(row):
        return ""
    return _text(row[index])


def _extract_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Workbook is missing required sheet {SHEET_NAME!r}.")
        sheet = workbook[SHEET_NAME]
        header_row, header_map = _find_header_row(sheet)
        rows: list[dict[str, Any]] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            sku = _normalized_sku(_row_value(row, header_map, "sku"))
            if not sku:
                continue
            family = _row_value(row, header_map, "family") or _family_from_sku(sku)
            rows.append(
                {
                    "sku": sku,
                    "family": _normalized_sku(family),
                    "category": _row_value(row, header_map, "category"),
                    "pm_responsible": _row_value(row, header_map, "pm_responsible"),
                    "series": _row_value(row, header_map, "series"),
                    "source_row": row_number,
                }
            )
        return rows
    finally:
        workbook.close()


def write_classification_cache(cache_path: Path, workbook_path: Path, rows: list[dict[str, Any]]) -> Path:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(cache_path)
    try:
        connection.execute("drop table if exists sku_classification")
        connection.execute("drop table if exists classification_metadata")
        connection.execute(
            "create table sku_classification ("
            + ", ".join(f"{column} text" for column in CLASSIFICATION_COLUMNS)
            + ")"
        )
        placeholders = ", ".join("?" for _ in CLASSIFICATION_COLUMNS)
        connection.executemany(
            "insert into sku_classification values (" + placeholders + ")",
            [tuple(row.get(column) for column in CLASSIFICATION_COLUMNS) for row in rows],
        )
        metadata = {
            "source_system": "sharepoint_powerbi_families_local_sqlite_cache",
            "source_workbook": str(workbook_path),
            "generated_at": utc_now(),
            "row_count": str(len(rows)),
        }
        connection.execute("create table classification_metadata (key text primary key, value text)")
        connection.executemany("insert into classification_metadata values (?, ?)", metadata.items())
        connection.execute("create index idx_sku_classification_sku on sku_classification(sku)")
        connection.execute("create index idx_sku_classification_family on sku_classification(family)")
        connection.execute("create index idx_sku_classification_category on sku_classification(category)")
        connection.commit()
    finally:
        connection.close()
    return cache_path


def refresh_classification_cache_from_workbook(root: Path, workbook_path: Path | None = None) -> Path:
    workbook = workbook_path or default_classification_workbook_path(root)
    if not workbook.exists():
        raise FileNotFoundError(
            f"Missing SKU classification workbook: {workbook}. "
            "Export or sync the SharePoint workbook locally, then rerun this refresh."
        )
    rows = _extract_rows(workbook)
    return write_classification_cache(classification_cache_path(root), workbook, rows)


def _metadata(cache_path: Path) -> dict[str, str]:
    connection = sqlite3.connect(cache_path)
    try:
        return dict(connection.execute("select key, value from classification_metadata").fetchall())
    finally:
        connection.close()


def classification_source(cache_path: Path) -> str:
    metadata = _metadata(cache_path)
    source = metadata.get("source_system") or "local_sku_classification_cache"
    generated_at = metadata.get("generated_at")
    row_count = metadata.get("row_count")
    details = []
    if generated_at:
        details.append(f"generated_at={generated_at}")
    if row_count:
        details.append(f"rows={row_count}")
    return source + (f" ({'; '.join(details)})" if details else "")


def load_classification_lookup(cache_path: Path) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    if not cache_path.exists():
        return {}, {}
    connection = sqlite3.connect(cache_path)
    connection.row_factory = sqlite3.Row
    try:
        rows = [dict(row) for row in connection.execute("select * from sku_classification").fetchall()]
    finally:
        connection.close()
    by_sku = {row["sku"]: row for row in rows if row.get("sku")}
    family_groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        family = row.get("family")
        if family:
            family_groups.setdefault(family, []).append(row)
    by_family = {family: group[0] for family, group in family_groups.items() if len(group) == 1}
    return by_sku, by_family


def enrich_catalog_rows_with_classification(root: Path, rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], Path | None, str | None]:
    cache_path = classification_cache_path(root)
    if not cache_path.exists():
        return rows, None, None
    by_sku, by_family = load_classification_lookup(cache_path)
    enriched: list[dict[str, Any]] = []
    for row in rows:
        candidates = [
            _normalized_sku(row.get("master_sku")),
            _normalized_sku(row.get("shopify_sku")),
        ]
        classification = next((by_sku[item] for item in candidates if item and item in by_sku), None)
        if classification is None:
            family_candidates = [_family_from_sku(item) for item in candidates if item]
            classification = next((by_family[item] for item in family_candidates if item in by_family), None)
        output = dict(row)
        if classification:
            output["classification_sku"] = classification.get("sku")
            output["classification_family"] = classification.get("family")
            output["classification_category"] = classification.get("category")
            output["classification_pm_responsible"] = classification.get("pm_responsible")
            output["classification_series"] = classification.get("series")
        enriched.append(output)
    return enriched, cache_path, classification_source(cache_path)
