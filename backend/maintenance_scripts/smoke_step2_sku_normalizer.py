from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend" / "app"))

from opportunity_engine.categories import load_categories  # noqa: E402
from opportunity_engine.ideation_template import generate_prd_ideation_workbook  # noqa: E402
from opportunity_engine.paths import ProjectPaths  # noqa: E402
from opportunity_engine.utils import timestamp  # noqa: E402


def _category_by_slug(paths: ProjectPaths, slug: str):
    for category in load_categories(paths):
        if category.slug == slug:
            return category
    raise ValueError(f"Category not found: {slug}")


def _write_step1_smoke_workbook(paths: ProjectPaths) -> Path:
    folder = paths.cache / "smoke_tests" / "step2_sku_normalizer"
    folder.mkdir(parents=True, exist_ok=True)
    output = folder / f"panels_step1_smoke_{timestamp()}.xlsx"

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Recommendations"
    headers = [
        "Recommendation",
        "Subcategory",
        "Priority",
        "Confidence",
        "Competitor / Ecommerce Example",
        "Review Link",
        "Sunco Active-Catalog Check",
        "Why This Is A True Gap",
        "PM Action",
    ]
    ws.append(headers)
    ws.append([
        "NPD: 2x2 and 2x4 selectable wattage panel with 3,200lm and 5,000lm output tiers",
        "Panels",
        "Medium",
        "Directional",
        "Step 1 ecommerce evidence: comparable selectable panel observed at $59.99",
        "",
        "0 strong Sunco active matches found.",
        (
            "Step 1 evidence explicitly names 2x2 and 2x4 panel sizes, 40W and 50W "
            "load targets, 3,200lm and 5,000lm output tiers, selectable CCT, and 0-10V dimming. "
            "Demand confidence score: 75/100."
        ),
        "Move forward as SKU-level concepts",
    ])
    ws.append([
        "Revision: add 0-10V selectable wattage merchandising to active panel family",
        "Panels",
        "High",
        "High",
        "Step 1 ecommerce evidence: active comparable calls out selectable wattage",
        "",
        "Partial Sunco active coverage found: PN_SM2X2-40W-0K: LED selectable panel family.",
        "Step 1 evidence names a missing control/merchandising callout on an active panel family.",
        "Add 0-10V/selectable-wattage callout to the active family listing and next rolling-change PRD.",
    ])

    amazon = workbook.create_sheet("Amazon Recommendations")
    amazon.append([
        "Amazon-channel recommendation",
        "Subcategory",
        "Priority",
        "Confidence",
        "Example listing",
        "Review link",
        "Sunco Amazon coverage check",
        "Stackline / Amazon evidence",
        "PM action",
        "Amazon classification",
    ])

    line_review = workbook.create_sheet("Existing SKU Line Review")
    line_review.append([
        "Family Part Number",
        "Product Title",
        "Vendor",
        "Vendor Cost",
        "Vendor Cost Source",
        "Amazon Revenue",
        "Shopify Revenue",
        "Total Revenue",
        "Pack Sizes Available",
        "Product URL",
        "Active Status",
    ])
    line_review.append([
        "PN_SM2X2-40W-0K",
        "Sunco LED selectable panel family",
        "Reference Vendor",
        18.5,
        "Step 1 line review",
        12500,
        9600,
        22100,
        "1PK",
        "https://www.sunco.com/",
        "Active",
    ])

    workbook.save(output)
    workbook.close()
    return output


def _write_residential_fans_step1_smoke_workbook(paths: ProjectPaths) -> Path:
    folder = paths.cache / "smoke_tests" / "step2_sku_normalizer"
    folder.mkdir(parents=True, exist_ok=True)
    output = folder / f"residential_fans_step1_smoke_{timestamp()}.xlsx"

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Recommendations"
    ws.append([
        "Recommendation",
        "Subcategory",
        "Priority",
        "Confidence",
        "Competitor / Ecommerce Example",
        "Review Link",
        "Sunco Active-Catalog Check",
        "Why This Is A True Gap",
        "PM Action",
    ])
    ws.append([
        "NPD: 52 inch ceiling fan with DC motor, remote control, reversible motor, matte black",
        "Residential Fans",
        "High",
        "High",
        "Verified competitor example: 52 inch indoor ceiling fan with remote and DC motor.",
        "",
        "Product case: no active Sunco ceiling fan SKU identified in this smoke fixture.",
        (
            "Step 1 evidence names a 52 inch ceiling fan, 3-blade format, DC motor, "
            "remote control, reversible motor, and matte black finish. The concept is fan-only "
            "and does not include a light kit."
        ),
        "Move forward as a fan-only NPD concept.",
    ])

    amazon = workbook.create_sheet("Amazon Recommendations")
    amazon.append([
        "Amazon-channel recommendation",
        "Subcategory",
        "Priority",
        "Confidence",
        "Example listing",
        "Review link",
        "Sunco Amazon coverage check",
        "Stackline / Amazon evidence",
        "PM action",
        "Amazon classification",
    ])

    workbook.save(output)
    workbook.close()
    return output


def _filled_ideation_rows(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    ws = workbook["Ideations"]
    rows: list[dict[str, str]] = []
    for row_index in range(4, ws.max_row + 1):
        name = ws.cell(row_index, 3).value
        if not name:
            continue
        rows.append({
            "name": str(name),
            "reference_sku": str(ws.cell(row_index, 4).value or ""),
            "action": str(ws.cell(row_index, 6).value or ""),
            "cct": str(ws.cell(row_index, 11).value or ""),
            "wattage": str(ws.cell(row_index, 8).value or ""),
            "lumens": str(ws.cell(row_index, 15).value or ""),
            "driver_type": str(ws.cell(row_index, 21).value or ""),
            "form_factor": str(ws.cell(row_index, 22).value or ""),
            "mounting": str(ws.cell(row_index, 23).value or ""),
            "efficiency": str(ws.cell(row_index, 16).value or ""),
            "notes": str(ws.cell(row_index, 54).value or ""),
        })
    workbook.close()
    return rows


def _workbook_text(workbook_path: Path) -> str:
    workbook = load_workbook(workbook_path, data_only=True)
    parts: list[str] = []
    for sheet in ["Ideations", "Source Mapping", "Run Audit"]:
        ws = workbook[sheet]
        for row in ws.iter_rows(values_only=True):
            parts.extend(str(value) for value in row if value not in (None, ""))
    workbook.close()
    return "\n".join(parts)


def main() -> int:
    paths = ProjectPaths.from_root(ROOT)
    paths.ensure()
    category = _category_by_slug(paths, "panels")
    step1_workbook = _write_step1_smoke_workbook(paths)
    step2_workbook, issues = generate_prd_ideation_workbook(paths, category, step1_workbook)
    expected_validation_notes = (
        "Excel COM open/save validation passed",
        "Excel COM validation skipped",
        "Excel COM validation failed",
    )
    if any(not issue.startswith(expected_validation_notes) for issue in issues):
        raise AssertionError("Step 2 validation issues: " + " | ".join(issues))

    rows = _filled_ideation_rows(step2_workbook)
    if len(rows) != 3:
        raise AssertionError(f"Expected 3 Step 2 rows, found {len(rows)}")
    actions = [row["action"] for row in rows]
    if actions[:2] != ["NPD", "NPD"] or actions[2] != "Revision":
        raise AssertionError(f"Unexpected action sort order: {actions}")
    if not any("2x2" in row["form_factor"] and row["lumens"] == "3,200lm" for row in rows):
        raise AssertionError("Missing 2x2 / 3,200lm SKU concept row")
    if not any("2x4" in row["form_factor"] and row["lumens"] == "5,000lm" for row in rows):
        raise AssertionError("Missing 2x4 / 5,000lm SKU concept row")
    revision_rows = [row for row in rows if row["action"] == "Revision"]
    if not revision_rows or "PN_SM2X2-40W-0K" not in revision_rows[0]["reference_sku"]:
        raise AssertionError("Revision row did not keep the active Sunco reference SKU")
    if "Revise PN_SM2X2-40W-0K" not in revision_rows[0]["notes"]:
        raise AssertionError("Revision row does not state the SKU and recommended changes")

    text = _workbook_text(step2_workbook)
    banned = [
        "schema_references",
        "templates/Competitors.md",
        "Bulb-dependent",
        "chandelier",
        "pack-count target",
        "_target_pack_count",
    ]
    hits = [token for token in banned if token.lower() in text.lower()]
    if hits:
        raise AssertionError("Banned Step 2 wording found: " + ", ".join(hits))

    fans_category = _category_by_slug(paths, "residential_fans")
    fans_step1_workbook = _write_residential_fans_step1_smoke_workbook(paths)
    fans_step2_workbook, fan_issues = generate_prd_ideation_workbook(paths, fans_category, fans_step1_workbook)
    if any(not issue.startswith(expected_validation_notes) for issue in fan_issues):
        raise AssertionError("Residential Fans Step 2 validation issues: " + " | ".join(fan_issues))
    fan_rows = _filled_ideation_rows(fans_step2_workbook)
    if len(fan_rows) != 1:
        raise AssertionError(f"Expected 1 Residential Fans Step 2 row, found {len(fan_rows)}")
    fan_row = fan_rows[0]
    if fan_row["reference_sku"].strip().lower() == "case":
        raise AssertionError("Residential Fans parser incorrectly accepted prose label 'case' as a Reference SKU")
    if "ceiling fan" not in fan_row["form_factor"].lower() or "52 in" not in fan_row["form_factor"].lower():
        raise AssertionError(f"Residential Fans form factor did not preserve fan size/category: {fan_row['form_factor']}")
    if "integrated led fixture" in fan_row["form_factor"].lower() or "lumen output target" in fan_row["lumens"].lower():
        raise AssertionError("Residential Fans row inherited integrated LED fixture defaults")
    if fan_row["lumens"]:
        raise AssertionError(f"Fan-only Residential Fans concept should not get lumens: {fan_row['lumens']}")
    if "DC Motor" not in fan_row["name"] or "Remote Control" not in fan_row["name"]:
        raise AssertionError(f"Residential Fans ideation name missed fan SKU-defining features: {fan_row['name']}")

    print(f"Step 1 smoke workbook: {step1_workbook}")
    print(f"Step 2 smoke workbook: {step2_workbook}")
    print(f"Rows: {len(rows)} ({', '.join(actions)})")
    print(f"Residential Fans Step 1 smoke workbook: {fans_step1_workbook}")
    print(f"Residential Fans Step 2 smoke workbook: {fans_step2_workbook}")
    print(f"Residential Fans row: {fan_row['name']} | Reference SKU: {fan_row['reference_sku']}")
    if issues:
        print("Validation notes:")
        for issue in issues:
            print(f"- {issue}")
    else:
        print("Validation passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
