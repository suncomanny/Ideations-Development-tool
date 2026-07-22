"""
Refresh and audit the approved competitor-source registry used by Step 3.

Usage:
  py backend\\app\\prd_research_tool\\tools\\approved_competitor_source_audit.py ^
    --source-workbook "C:\\Users\\Sunco\\Downloads\\Sunco Competitor Research (2).xlsx" ^
    --write-cache --coverage
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from approved_competitor_sources import APPROVED_COMPETITOR_SOURCE_PATH, normalize_domain


REPO_ROOT = Path(__file__).resolve().parents[4]
PRODUCT_DEMAND_SRC = REPO_ROOT / "product_demand_ideation" / "src"
OUTPUT_DIR = REPO_ROOT / "outputs" / "Research" / "Source Audits"


def sql_literal(value: Any) -> str:
    return "'" + str(value or "").replace("'", "''") + "'"


def parse_approved_sources(workbook_path: Path) -> list[dict[str, Any]]:
    """Parse the first sheet of the approved competitor research workbook."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers = [str(worksheet.cell(4, column).value or "").strip() for column in range(1, worksheet.max_column + 1)]
        index = {header: position + 1 for position, header in enumerate(headers) if header}
        sources: list[dict[str, Any]] = []
        seen: set[str] = set()
        for row in range(5, worksheet.max_row + 1):
            competitor = worksheet.cell(row, index.get("Competitor", 0)).value if index.get("Competitor") else None
            website = worksheet.cell(row, index.get("Website", 0)).value if index.get("Website") else None
            domain = normalize_domain(website)
            if not competitor or not domain or domain in seen:
                continue
            seen.add(domain)
            sources.append(
                {
                    "competitor": str(competitor).strip(),
                    "website": str(website).strip(),
                    "domain": domain,
                    "type": str(worksheet.cell(row, index.get("Type", 0)).value or "").strip() if index.get("Type") else "",
                    "tier": str(worksheet.cell(row, index.get("Tier", 0)).value or "").strip() if index.get("Tier") else "",
                    "tier_rank": str(worksheet.cell(row, index.get("Tier Rank", 0)).value or "").strip() if index.get("Tier Rank") else "",
                    "focus": str(worksheet.cell(row, index.get("Focus", 0)).value or "").strip() if index.get("Focus") else "",
                    "product_subcategories_in_scope": str(worksheet.cell(row, index.get("Product Subcategories in Scope", 0)).value or "").strip()
                    if index.get("Product Subcategories in Scope")
                    else "",
                    "strengths": str(worksheet.cell(row, index.get("Strengths", 0)).value or "").strip() if index.get("Strengths") else "",
                    "weaknesses_notes": str(worksheet.cell(row, index.get("Weaknesses / Notes", 0)).value or "").strip()
                    if index.get("Weaknesses / Notes")
                    else "",
                    "pricing_strategy": str(worksheet.cell(row, index.get("Pricing Strategy", 0)).value or "").strip()
                    if index.get("Pricing Strategy")
                    else "",
                }
            )
        return sources
    finally:
        workbook.close()


def build_coverage_sql(sources: list[dict[str, Any]]) -> str:
    """Return Redshift SQL that checks approved domains in the scrape views."""
    approved_selects = []
    for source in sources:
        approved_selects.append(
            "select "
            f"{sql_literal(source.get('competitor'))} as competitor, "
            f"{sql_literal(source.get('domain'))} as domain, "
            f"{sql_literal(source.get('website'))} as website, "
            f"{sql_literal(source.get('type'))} as type, "
            f"{sql_literal(source.get('tier'))} as tier, "
            f"{sql_literal(source.get('tier_rank'))} as tier_rank, "
            f"{sql_literal(source.get('focus'))} as focus, "
            f"{sql_literal(source.get('product_subcategories_in_scope'))} as scope"
        )
    approved_cte = "\nunion all\n".join(approved_selects)
    return f"""
with approved as (
{approved_cte}
), latest_domain as (
    select
        lower(regexp_replace(split_part(replace(replace(url, 'https://', ''), 'http://', ''), '/', 1), '^www\\.', '')) as row_domain,
        count(*) as latest_rows,
        count(distinct url) as latest_urls,
        max(scraped_at) as latest_last_scraped
    from public.v_competitors_scrapping_latest
    where url is not null and url <> ''
    group by 1
), latest_rollup as (
    select
        a.domain,
        coalesce(sum(l.latest_rows), 0) as latest_rows,
        coalesce(sum(l.latest_urls), 0) as latest_urls,
        max(l.latest_last_scraped) as latest_last_scraped
    from approved a
    left join latest_domain l
        on l.row_domain = a.domain or l.row_domain like '%.' || a.domain
    group by a.domain
), inventory_domain as (
    select
        lower(regexp_replace(split_part(replace(replace(url, 'https://', ''), 'http://', ''), '/', 1), '^www\\.', '')) as row_domain,
        count(*) as inventory_rows,
        count(distinct url) as inventory_urls,
        max(scrape_date) as inventory_last_scraped,
        sum(case when stock_qty_delta < 0 then abs(stock_qty_delta) else 0 end) as observed_stock_decrease,
        count(case when stock_qty_delta < 0 then 1 end) as decrease_events
    from public.v_competitors_inventory_daily
    where url is not null and url <> ''
    group by 1
), inventory_rollup as (
    select
        a.domain,
        coalesce(sum(i.inventory_rows), 0) as inventory_rows,
        coalesce(sum(i.inventory_urls), 0) as inventory_urls,
        max(i.inventory_last_scraped) as inventory_last_scraped,
        coalesce(sum(i.observed_stock_decrease), 0) as observed_stock_decrease,
        coalesce(sum(i.decrease_events), 0) as decrease_events
    from approved a
    left join inventory_domain i
        on i.row_domain = a.domain or i.row_domain like '%.' || a.domain
    group by a.domain
)
select
    a.competitor,
    a.domain,
    a.website,
    a.type,
    a.tier,
    a.tier_rank,
    a.focus,
    a.scope,
    lr.latest_rows,
    lr.latest_urls,
    lr.latest_last_scraped,
    ir.inventory_rows,
    ir.inventory_urls,
    ir.inventory_last_scraped,
    ir.observed_stock_decrease,
    ir.decrease_events
from approved a
left join latest_rollup lr on lr.domain = a.domain
left join inventory_rollup ir on ir.domain = a.domain
order by
    case when lower(a.tier) like '%tier 1%' then 1 when lower(a.tier) like '%tier 2%' then 2 when lower(a.tier) like '%tier 3%' then 3 else 9 end,
    nullif(a.tier_rank, '')::int nulls last,
    a.competitor;
""".strip()


def run_coverage_query(sources: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], str, str | None]:
    """Run the approved-domain coverage query through Redshift MCP, with ODBC fallback."""
    if str(PRODUCT_DEMAND_SRC) not in sys.path:
        sys.path.insert(0, str(PRODUCT_DEMAND_SRC))
    from redshift_query import execute_redshift_sql, sanitize_redshift_error

    sql = build_coverage_sql(sources)
    try:
        rows, connection_source = execute_redshift_sql(
            sql,
            timeout_seconds=240,
            client_name="sunco-approved-competitor-source-audit",
        )
        return rows, connection_source, None
    except Exception as exc:  # pragma: no cover - depends on workstation ODBC credentials.
        return [], "Redshift MCP primary; ODBC fallback", sanitize_redshift_error(exc)


def merge_coverage(sources: list[dict[str, Any]], coverage_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Attach coverage counts to the approved-source cache."""
    by_domain = {normalize_domain(row.get("domain")): row for row in coverage_rows}
    output = []
    for source in sources:
        domain = normalize_domain(source.get("domain"))
        coverage = by_domain.get(domain, {})
        merged = dict(source)
        for key in [
            "latest_rows",
            "latest_urls",
            "latest_last_scraped",
            "inventory_rows",
            "inventory_urls",
            "inventory_last_scraped",
            "observed_stock_decrease",
            "decrease_events",
        ]:
            merged[key] = coverage.get(key)
        output.append(merged)
    return output


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Refresh approved competitor source registry and audit Redshift coverage.")
    parser.add_argument("--source-workbook", type=Path, required=True, help="Workbook containing the Master List sheet.")
    parser.add_argument("--write-cache", action="store_true", help="Write backend/source_data competitor-source cache.")
    parser.add_argument("--coverage", action="store_true", help="Run Redshift coverage query for approved domains.")
    args = parser.parse_args()

    sources = parse_approved_sources(args.source_workbook)
    coverage_rows: list[dict[str, Any]] = []
    connection_source = ""
    error = None
    if args.coverage:
        coverage_rows, connection_source, error = run_coverage_query(sources)
        sources = merge_coverage(sources, coverage_rows) if coverage_rows else sources

    generated_at = datetime.now().isoformat(timespec="seconds")
    payload = {
        "generated_at": generated_at,
        "source_workbook": str(args.source_workbook),
        "source_sheet": "Master List",
        "source_row_count": len(sources),
        "coverage_connection_source": connection_source,
        "coverage_error": error,
        "sources": sources,
    }
    if args.write_cache:
        APPROVED_COMPETITOR_SOURCE_PATH.parent.mkdir(parents=True, exist_ok=True)
        APPROVED_COMPETITOR_SOURCE_PATH.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    audit_path = OUTPUT_DIR / f"approved_competitor_scrape_coverage_{stamp}.json"
    audit_payload = dict(payload)
    audit_payload["coverage_rows"] = coverage_rows
    audit_path.write_text(json.dumps(audit_payload, indent=2, default=str), encoding="utf-8")

    csv_path = OUTPUT_DIR / f"approved_competitor_scrape_coverage_{stamp}.csv"
    write_csv(csv_path, coverage_rows if coverage_rows else sources)

    domains_with_latest = sum(1 for row in sources if (row.get("latest_rows") or 0) > 0)
    domains_with_inventory = sum(1 for row in sources if (row.get("inventory_rows") or 0) > 0)
    print(f"approved_count {len(sources)}")
    print(f"domains_with_latest {domains_with_latest}")
    print(f"domains_with_inventory {domains_with_inventory}")
    print(f"cache {APPROVED_COMPETITOR_SOURCE_PATH if args.write_cache else 'not written'}")
    print(f"json {audit_path}")
    print(f"csv {csv_path}")
    if error:
        print(f"coverage_error {error}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
