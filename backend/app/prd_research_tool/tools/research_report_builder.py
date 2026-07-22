"""
Step 6A: Build Excel research reports from completed analysis artifacts.

Usage:
  python tools/research_report_builder.py "C:\\path\\to\\research_session"
  python tools/research_report_builder.py "C:\\path\\to\\research_session" --rows 3,4,5
  python tools/research_report_builder.py "C:\\path\\to\\research_session" --combined
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from approved_competitor_sources import (
    approved_competitor_source_for_domain,
    approved_sources_for_category,
    domain_matches,
    load_approved_competitor_sources,
    source_search_link,
)
from research_session_manager import artifact_path_for, packet_path_for, read_json, update_session


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
ACCENT_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
SUBHEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=15)
HYPERLINK_FONT = Font(color="0563C1", underline="single")
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
DEFAULT_COLUMN_WIDTHS = {
    "A": 24,
    "B": 34,
    "C": 18,
    "D": 18,
    "E": 18,
    "F": 18,
    "G": 18,
    "H": 24,
    "I": 18,
    "J": 42,
    "K": 14,
    "L": 24,
    "M": 24,
}
ALTERNATE_ROW_FILL = PatternFill(fill_type="solid", fgColor="F7FBFF")
SLIDE_TEAL = "009AA6"
SLIDE_NAVY = "26384F"
SLIDE_DARK_NAVY = "0B1F36"
SLIDE_LIGHT_BLUE = "EAF2F8"
SLIDE_LIGHT_GREY = "F4F7FA"
SLIDE_BORDER = "D7DEE8"
SLIDE_TEXT = "102033"
SLIDE_MUTED = "6B7A90"
SLIDE_TEAL_FILL = PatternFill(fill_type="solid", fgColor=SLIDE_TEAL)
SLIDE_NAVY_FILL = PatternFill(fill_type="solid", fgColor=SLIDE_NAVY)
SLIDE_LIGHT_FILL = PatternFill(fill_type="solid", fgColor=SLIDE_LIGHT_BLUE)
SLIDE_WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
SLIDE_SECTION_FONT = Font(color="FFFFFF", bold=True, size=9)
SLIDE_TITLE_FONT = Font(color=SLIDE_DARK_NAVY, bold=True, size=18)
SLIDE_BODY_FONT = Font(color=SLIDE_TEXT, size=9)
SLIDE_SMALL_FONT = Font(color=SLIDE_TEXT, size=8)
SLIDE_MUTED_FONT = Font(color=SLIDE_MUTED, italic=True, size=7)
SLIDE_LINK_FONT = Font(color="0563C1", underline="single", size=8)
SLIDE_THIN_SIDE = Side(style="thin", color=SLIDE_BORDER)
SLIDE_MEDIUM_SIDE = Side(style="medium", color=SLIDE_TEAL)
SLIDE_TABLE_BORDER = Border(left=SLIDE_THIN_SIDE, right=SLIDE_THIN_SIDE, top=SLIDE_THIN_SIDE, bottom=SLIDE_THIN_SIDE)
AMAZON_CHANNELS = {"amazon"}
BM_DIRECT_CHANNELS = {"home_depot", "walmart", "lowes", "brand_site"}
SKU_DECODER_CACHE = Path(__file__).resolve().parents[3] / "source_data" / "sku_decoder" / "sku_decoder_clean.csv"
ECOMMERCE_CACHE_ROOT = Path(__file__).resolve().parents[3] / "source_data" / "redshift_ecommerce_cache"
CHANNEL_DISPLAY_NAMES = {
    "amazon": "Amazon",
    "home_depot": "Home Depot",
    "walmart": "Walmart",
    "lowes": "Lowe's",
    "brand_site": "Brand Site",
    "stackline_seed": "Stackline",
}

SUMMARY_METRIC_GUIDE_ROWS = [
    [
        "Recommended Product Action",
        "PM-facing action label: NPD, Revision, Concept Review, or Hold.",
        "What should the PM do with this SKU/product recommendation next?",
    ],
    [
        "Gap Reason",
        "Supporting reason behind the action, such as missing feature, existing coverage, or new variant opportunity.",
        "Why did the tool classify this concept this way?",
    ],
    [
        "Revision Target SKU",
        "Existing Sunco/NSL SKU or active family to update when the action is Revision.",
        "Which current product would we change instead of creating a new SKU?",
    ],
    [
        "Recommended Revision Changes",
        "Specific rolling-change, feature-add, or merchandising update recommended for a Revision row.",
        "What exactly should the PM ask to change?",
    ],
    [
        "Outlook",
        "Directional recommendation for the ideation in its current configuration, such as favorable, mixed, or cautious.",
        "Given the category, competitor, pricing, and feature context, should this concept look attractive to pursue as configured?",
    ],
    [
        "Confidence",
        "Overall confidence in the launch outlook based on the quality, consistency, and completeness of the supporting research.",
        "How much trust should I place in the outlook recommendation?",
    ],
    [
        "Amazon G2",
        "Weighted Amazon Gate 2 readiness score on a 0-10 scale using the current rubric and available evidence.",
        "How ready is this concept to move forward for Amazon Gate 2 review?",
    ],
    [
        "Amazon Evidence",
        "Confidence label for the Amazon G2 score based on how direct, complete, and well-supported the underlying data is.",
        "How strong is the evidence behind the Amazon G2 score?",
    ],
]

OPPORTUNITY_TYPES = (
    "Possible feature gap",
    "Existing Sunco coverage, but missing feature",
    "Product Revision or merchandising review",
    "New variant opportunity",
    "Strategic outlier / High-output watchlist",
)
PRODUCT_ACTION_NPD = "NPD"
PRODUCT_ACTION_REVISION = "Revision"
PRODUCT_ACTION_CONCEPT_REVIEW = "Concept Review"
PRODUCT_ACTION_HOLD = "Hold"
PRODUCT_ACTIONS = (
    PRODUCT_ACTION_NPD,
    PRODUCT_ACTION_REVISION,
    PRODUCT_ACTION_CONCEPT_REVIEW,
    PRODUCT_ACTION_HOLD,
)
PRODUCT_ACTION_SORT_ORDER = {
    PRODUCT_ACTION_NPD: 0,
    PRODUCT_ACTION_REVISION: 1,
    PRODUCT_ACTION_CONCEPT_REVIEW: 2,
    PRODUCT_ACTION_HOLD: 3,
}
EVIDENCE_STRENGTH_STRONG = "Strong support"
EVIDENCE_STRENGTH_DIRECTIONAL = "Directional support"
EVIDENCE_STRENGTH_REVIEW = "Needs PM review"
MATCH_QUALITY_APPLES = "Apples-to-apples"
MATCH_QUALITY_SIMILAR = "Similar but not exact"
MATCH_QUALITY_NOT_COMPARABLE = "Not comparable"
LINK_STATUS_VERIFIED = "Verified link"
LINK_STATUS_NEEDS_CHECK = "Needs link check"
LINK_STATUS_MISSING = "No direct link"
LINK_STATUS_INVALID = "Needs product PDP"

SOURCE_LABELS = {
    "legacy_fy2025_sales_export": "Local historical sales fallback",
    "legacy_metadata_variant_price": "Local catalog price fallback",
    "legacy_fy2025_shopify_sales_csv": "Local historical Shopify sales fallback",
    "legacy_fy2025_amazon_sales_csv": "Local historical Amazon sales fallback",
    "legacy_local_metadata_and_fy2025_sales_exports": "Local catalog and historical sales fallback",
    "legacy_fy2025_amazon_export_fallback": "Local historical Amazon sales fallback",
    "mixed_legacy_fy2025_amazon_export_fallback": "Mixed live/cache plus local historical Amazon fallback",
    "target_msrp": "Target MSRP",
}

DISPLAY_TEXT_REPLACEMENTS = (
    ("max parsed wattage", "observed max wattage"),
    (" target from Step 1 evidence", " target"),
    ("No normalized competitors currently surface ", "Current competitor set does not clearly show "),
    (
        "; validate that it represents a real customer need before locking in added complexity.",
        "; treat as a hypothesis unless a PM wants this as a deliberate differentiator.",
    ),
    (
        "; validate whether that certification is a real channel requirement before adding cost.",
        "; confirm this is required for the intended channel before adding cost.",
    ),
    (
        "Autofilled locally by Codex from Stackline-backed packet seeds to reduce Claude collection cost.",
        "Autofilled from Stackline-backed market seeds to reduce manual collection work.",
    ),
    (
        "These are market-intelligence seeds, not live page fetches; use optional web enrichment for missing detailed specs/certifications.",
        "These are market-intelligence seeds, not live page fetches; use optional web enrichment only for missing specs or certifications.",
    ),
    (
        "Brand-site collection was explicitly skipped for a no-Claude/local-only clean run.",
        "Official brand-site evidence was not included in this run.",
    ),
    (
        "No brand product pages were fetched or fabricated in this artifact.",
        "No official brand product pages were used for this row.",
    ),
    (
        "Use Claude or a future web collector later if official brand-site specs are needed.",
        "Use optional web enrichment if official brand-site specs are needed.",
    ),
    ("Prioritize Claude collection next.", "Prioritize optional web collection next."),
    ("Legacy FY2025 local export fallback", "Local historical sales fallback period"),
    ("validate premium feature justification or lower price", "confirm premium feature justification or lower price"),
)


def resolved_source_channel(item: dict[str, Any]) -> str:
    """Return the effective source channel, preferring explicit retailer domains."""
    raw_channel = (optional_text(item.get("source_channel")) or "").lower()
    url = optional_text(item.get("url")) or ""
    detected_domain = urlparse(url).netloc.lower() if "://" in url else ""
    domain = detected_domain or (optional_text(item.get("source_domain")) or "")
    if domain:
        host = urlparse(domain).netloc.lower() if "://" in domain else domain.lower()
        host = host.replace("www.", "")
        if host == "amazon.com" or host.endswith(".amazon.com"):
            return "amazon"
        if host == "homedepot.com" or host.endswith(".homedepot.com"):
            return "home_depot"
        if host == "walmart.com" or host.endswith(".walmart.com"):
            return "walmart"
        if host == "lowes.com" or host.endswith(".lowes.com"):
            return "lowes"
    return raw_channel


def normalize_text(value: Any) -> str:
    """Render values as readable worksheet strings."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, list):
        return ", ".join([normalize_text(item) for item in value if normalize_text(item)])
    text = str(value)
    text = SOURCE_LABELS.get(text, text)
    for old, new in DISPLAY_TEXT_REPLACEMENTS:
        text = text.replace(old, new)
    return fix_ordinal_suffixes(text)


def ordinal_suffix(number: int) -> str:
    """Return the English ordinal suffix for small display strings."""
    if 10 <= number % 100 <= 20:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(number % 10, "th")


def fix_ordinal_suffixes(text: str) -> str:
    """Fix phrases such as 73th percentile without changing source data."""
    return re.sub(
        r"\b(\d+)(?:st|nd|rd|th)\b",
        lambda match: f"{int(match.group(1))}{ordinal_suffix(int(match.group(1)))}",
        text,
    )


def decimal_from_value(value: Any) -> Decimal | None:
    """Parse simple numeric values without keeping float representation noise."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip().replace("$", "").replace(",", "").replace("%", "")
        if not re.fullmatch(r"-?\d+(?:\.\d+)?", text):
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None
    return None


def format_decimal(value: Decimal, places: int = 2) -> str:
    """Return a readable decimal string with thousands separators."""
    quantized = value.quantize(Decimal("1") if places == 0 else Decimal("1." + ("0" * places)))
    return f"{quantized:,.{places}f}"


def display_value_for_label(label: str, value: Any) -> str:
    """Format report values using the row label/header for context."""
    text = normalize_text(value)
    label_key = normalize_text(label).lower()
    number = decimal_from_value(value if value is not None else text)
    if number is None:
        return text

    if any(term in label_key for term in ["%", "percent", "percentile", "growth", "share", "vs target", "vs market"]):
        return f"{format_decimal(number, 1)}%"
    if any(term in label_key for term in ["score", "confidence"]):
        return format_decimal(number, 2).rstrip("0").rstrip(".")
    if any(term in label_key for term in ["units", "count", "samples"]):
        return format_decimal(number, 0)
    if any(term in label_key for term in ["price", "pricing", "revenue", "sales", "msrp", "cost", "retail", "floor", "ceiling"]):
        return f"${format_decimal(number, 2)}"
    return format_decimal(number, 2).rstrip("0").rstrip(".")


def as_dict(value: Any) -> dict[str, Any]:
    """Coerce optional mappings into dicts."""
    if isinstance(value, dict):
        return value
    return {}


def as_list(value: Any) -> list[Any]:
    """Coerce optional sequences into lists."""
    if isinstance(value, list):
        return value
    return []


def slugify(value: Any) -> str:
    """Normalize category-like values into stable slugs."""
    text = normalize_text(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def category_slug_for_packet(packet: dict[str, Any]) -> str:
    """Return the category slug used by local cache folders."""
    identity = as_dict(packet.get("identity"))
    return slugify(identity.get("subcategory") or identity.get("category"))


def short_domain(value: Any) -> str:
    """Return a compact display domain from a URL or host."""
    text = normalize_text(value)
    if not text:
        return ""
    host = urlparse(text if "://" in text else f"//{text}").netloc or text
    return host.lower().replace("www.", "").strip("/")


@lru_cache(maxsize=128)
def latest_ecommerce_snapshot_payload(category_slug: str) -> dict[str, Any]:
    """Load the newest Redshift ecommerce competitor snapshot for a category."""
    if not category_slug or not ECOMMERCE_CACHE_ROOT.exists():
        return {}
    candidates = sorted(
        ECOMMERCE_CACHE_ROOT.glob(f"**/{category_slug}_ecommerce_competitor_evidence_*.json"),
        key=lambda path: path.stat().st_mtime,
    )
    if not candidates:
        return {}
    try:
        return json.loads(candidates[-1].read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def first_number_from_fields(item: dict[str, Any], keys: list[str]) -> Decimal | None:
    """Return the first numeric value from an item field list."""
    for key in keys:
        number = decimal_from_value(item.get(key))
        if number is not None:
            return number
    return None


def number_from_observation(item: dict[str, Any], observation_key: str) -> Decimal | None:
    """Parse a numeric key=value metric from raw observation strings."""
    pattern = re.compile(rf"\b{re.escape(observation_key)}\s*=\s*([0-9,]+(?:\.[0-9]+)?)", flags=re.IGNORECASE)
    for observation in as_list(item.get("raw_observations")):
        match = pattern.search(normalize_text(observation))
        if match:
            return decimal_from_value(match.group(1))
    return None


def display_number(value: Decimal, places: int = 0) -> str:
    """Format a Decimal for compact demand cells."""
    return format_decimal(value, places).rstrip("0").rstrip(".") if places else format_decimal(value, 0)


def demand_signal_parts_from_text(text: Any) -> list[str]:
    """Split legacy free-text movement notes into volume, rate, and recency/window."""
    note = normalize_text(text)
    if not note:
        return ["", "", ""]

    volume_parts: list[str] = []
    match = re.search(r"([0-9,]+(?:\.[0-9]+)?)\s+units?\s+observed stock decrease", note, flags=re.IGNORECASE)
    if match:
        volume_parts.append(f"{match.group(1)} stock decrease units")
    match = re.search(r"([0-9,]+)\s+decrease events?", note, flags=re.IGNORECASE)
    if match:
        volume_parts.append(f"{match.group(1)} decrease events")

    rate = ""
    match = re.search(r"~?\s*([0-9,]+(?:\.[0-9]+)?)\s+units/week", note, flags=re.IGNORECASE)
    if match:
        rate = f"{match.group(1)} units/week"

    recency_parts: list[str] = []
    match = re.search(r"last decrease\s+([0-9,]+)\s+day", note, flags=re.IGNORECASE)
    if match:
        recency_parts.append(f"{match.group(1)} days since last decrease")
    match = re.search(r"observed over\s+([0-9,]+)\s+day", note, flags=re.IGNORECASE)
    if match:
        recency_parts.append(f"{match.group(1)}-day observed window")
    match = re.search(r"decrease activity over\s+([0-9,]+)\s+day", note, flags=re.IGNORECASE)
    if match:
        recency_parts.append(f"{match.group(1)}-day decrease window")

    return [
        "; ".join(volume_parts) or note,
        rate,
        "; ".join(recency_parts),
    ]


@lru_cache(maxsize=1)
def sku_decoder_entries() -> list[dict[str, str]]:
    """Load the local SKU decoder cache used for concept tracking names."""
    if not SKU_DECODER_CACHE.exists():
        return []
    with SKU_DECODER_CACHE.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def decoder_rows_for(category_slug: str, normalized_code_category: str) -> list[dict[str, str]]:
    """Return SKU decoder rows that apply to a category and code family."""
    return [
        row
        for row in sku_decoder_entries()
        if row.get("mapped_category_slug") == category_slug
        and row.get("normalized_code_category") == normalized_code_category
    ]


def context_text_for_packet(packet: dict[str, Any]) -> str:
    """Collect readable packet context for token inference."""
    identity = as_dict(packet.get("identity"))
    target_profile = as_dict(packet.get("target_profile"))
    electrical = as_dict(target_profile.get("electrical"))
    physical = as_dict(target_profile.get("physical"))
    business = as_dict(target_profile.get("business_case"))
    pieces: list[str] = [
        identity.get("ideation_name"),
        identity.get("category"),
        identity.get("subcategory"),
        identity.get("sunco_reference_sku"),
        electrical.get("wattage_primary"),
        electrical.get("wattage_max"),
        electrical.get("cct_primary"),
        electrical.get("cct_max"),
        electrical.get("lumens_target"),
        physical.get("size_form_factor"),
        physical.get("finish_color"),
        physical.get("mounting_type"),
        business.get("certifications"),
    ]
    pieces.extend(as_list(target_profile.get("feature_watchlist")))
    return " ".join(normalize_text(piece) for piece in pieces if normalize_text(piece))


def preferred_product_type_code(packet: dict[str, Any]) -> str:
    """Infer the most useful product-type code for a Step 3 concept name."""
    identity = as_dict(packet.get("identity"))
    category_slug = slugify(identity.get("subcategory") or identity.get("category"))
    rows = decoder_rows_for(category_slug, "product_type")
    context = context_text_for_packet(packet).lower()
    reference_sku = normalize_text(identity.get("sunco_reference_sku")).upper()

    if category_slug == "wraparounds":
        if re.search(r"\b2\s*(?:ft|foot|feet)\b|\b2ft\b", context):
            return "WR2"
        if re.search(r"\b4\s*(?:ft|foot|feet)\b|\b4ft\b", context):
            return "WR"
    if category_slug == "panels":
        panel_map = {
            "1x2": "PN12",
            "1 x 2": "PN12",
            "1x4": "PN14",
            "1 x 4": "PN14",
            "2x2": "PN22",
            "2 x 2": "PN22",
            "2x4": "PN24",
            "2 x 4": "PN24",
        }
        for token, code in panel_map.items():
            if token in context:
                return code

    non_phasing_rows = [
        row
        for row in rows
        if "phasing" not in row.get("code_meaning", "").lower()
        and "legacy" not in row.get("code_meaning", "").lower()
    ]
    for row in non_phasing_rows:
        match_prefix = normalize_text(row.get("match_prefix")).upper()
        if match_prefix and reference_sku.startswith(match_prefix):
            return normalize_text(row.get("code")).replace("_", "")
    for row in non_phasing_rows:
        meaning = row.get("code_meaning", "").lower()
        if meaning and len(meaning) > 2 and meaning in context:
            return normalize_text(row.get("code")).replace("_", "")
    if non_phasing_rows:
        return normalize_text(non_phasing_rows[0].get("code")).replace("_", "")
    if rows:
        return normalize_text(rows[0].get("code")).replace("_", "")
    fallback = slugify(identity.get("subcategory") or identity.get("category") or "IDEA")
    return fallback[:8].upper() or "IDEA"


def first_wattage_token(text: str) -> str:
    """Return the first wattage token found in text."""
    match = re.search(r"\b(\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?)\s*W\b", text, flags=re.IGNORECASE)
    if not match:
        return ""
    number = match.group(1).replace(".0", "")
    return f"{number}W"


def first_lumen_token(text: str) -> str:
    """Return a compact lumen token when wattage is unavailable."""
    match = re.search(r"\b([0-9][0-9,]*)\s*(?:lm|lumens?)\b", text, flags=re.IGNORECASE)
    if not match:
        return ""
    return f"{match.group(1).replace(',', '')}LM"


def cct_token_from_text(text: str) -> str:
    """Infer a SKU-decoder-style CCT token from available target text."""
    lowered = text.lower()
    cct_values = [int(value) for value in re.findall(r"\b((?:27|30|35|40|50|60|65|70)00)\s*k\b", lowered)]
    expanded = sorted(set(cct_values))
    if 3000 in expanded and 5000 in expanded:
        return "3050"
    if 3000 in expanded and 6500 in expanded:
        return "3065"
    if 4000 in expanded and 6000 in expanded:
        return "4060"
    if 2700 in expanded and 5000 in expanded:
        return "2750"
    if 2700 in expanded and 4000 in expanded:
        return "2740"
    if 2700 in expanded and 6000 in expanded:
        return "2760"
    if "3cct" in lowered and not expanded:
        return "3CCT"
    single = re.search(r"\b(2700|3000|3500|4000|5000|6000|6500|7000)\s*K?\b", text, flags=re.IGNORECASE)
    return single.group(1) if single else ""


def finish_token_from_text(text: str) -> str:
    """Infer a finish token using common SKU decoder color codes."""
    lowered = text.lower()
    finish_map = [
        ("black", "BK"),
        ("bronze", "BR"),
        ("chrome", "CR"),
        ("gold", "GD"),
        ("nickel", "NK"),
        ("pure white", "PW"),
        ("silver", "SV"),
        ("white", "WH"),
    ]
    for label, code in finish_map:
        if label in lowered:
            return code
    return ""


def feature_tokens_from_text(text: str) -> list[str]:
    """Infer compact feature/control tokens for concept tracking names."""
    lowered = text.lower()
    tokens: list[str] = []
    if "ultra high output" in lowered:
        tokens.append("UO")
    elif "high output" in lowered:
        tokens.append("HO")
    if "selectable wattage" in lowered:
        tokens.append("SW")
    if "0-10" in lowered or "0 to 10" in lowered:
        tokens.append("010V")
    elif "dimmable" in lowered:
        tokens.append("DIM")
    if "dusk to dawn" in lowered or "d2d" in lowered:
        tokens.append("DD")
    if "smart" in lowered:
        tokens.append("S")
    if "sensor" in lowered:
        tokens.append("SNS")
    return list(dict.fromkeys(tokens))


def pack_token_from_text(text: str) -> str:
    """Return a compact pack suffix such as /4 when a multi-pack is explicit."""
    match = re.search(r"\b(\d+)\s*(?:pack|pk)\b", text, flags=re.IGNORECASE)
    if not match:
        return ""
    count = int(match.group(1))
    return f"/{count}" if count > 1 else ""


def concept_tracking_name(packet: dict[str, Any]) -> str:
    """Build a compact, SKU-decoder-informed concept tracking name."""
    context = context_text_for_packet(packet)
    product_code = preferred_product_type_code(packet)
    wattage = first_wattage_token(context)
    lumen = first_lumen_token(context)
    cct = cct_token_from_text(context)
    finish = finish_token_from_text(context)
    tokens = [product_code]
    if wattage:
        tokens.append(wattage)
    elif lumen:
        tokens.append(lumen)
    if cct:
        tokens.append(cct)
    tokens.extend(feature_tokens_from_text(context))
    if finish:
        tokens.append(finish)
    name = "-".join(token for token in tokens if token)
    name = re.sub(r"-+", "-", name).strip("-")
    return f"{name}{pack_token_from_text(context)}" if name else "IDEA"


def set_default_layout(ws) -> None:
    """Apply shared column widths and wrapping."""
    ws.freeze_panes = None
    ws.sheet_view.showGridLines = False
    for column, width in DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width


def sheet_title_for_row(row_number: int, packet: dict[str, Any], used: set[str] | None = None) -> str:
    """Create a compact row-first sheet title for combined reports."""
    tracking_name = re.sub(r"[^A-Za-z0-9/_-]+", " ", concept_tracking_name(packet)).strip()
    base = f"R{row_number:02d} {tracking_name}" if tracking_name else f"R{row_number:02d} Ideation"
    return safe_sheet_title(base, used or set())


def safe_sheet_title(base: str, used: set[str]) -> str:
    """Create a workbook-safe, unique worksheet title."""
    text = re.sub(r"[\[\]\*?/\\:]", " ", base or "Report").strip()
    text = re.sub(r"\s+", " ", text)
    text = text[:31] or "Report"
    candidate = text
    suffix = 2
    while candidate in used:
        trimmed = text[: max(0, 31 - len(f" ({suffix})"))]
        candidate = f"{trimmed} ({suffix})"
        suffix += 1
    used.add(candidate)
    return candidate


def section_header(ws, row: int, title: str, end_column: int = 10) -> int:
    """Write a section header row and return the next row index."""
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    return row + 1


def key_value_rows(ws, row: int, pairs: list[tuple[str, Any]], columns: int = 2) -> int:
    """Write compact key/value pairs across the sheet."""
    index = 0
    while index < len(pairs):
        for block in range(columns):
            if index >= len(pairs):
                break
            label, value = pairs[index]
            base_col = (block * 2) + 1
            ws.cell(row=row, column=base_col, value=label)
            ws.cell(row=row, column=base_col).font = SUBHEADER_FONT
            ws.cell(row=row, column=base_col).fill = SUBHEADER_FILL
            ws.cell(row=row, column=base_col + 1, value=display_value_for_label(label, value))
            ws.cell(row=row, column=base_col + 1).alignment = WRAP_ALIGNMENT
            index += 1
        row += 1
    return row


def gate_snapshot(gate_readiness: dict[str, Any], channel: str, gate: str) -> dict[str, Any]:
    """Return the matching gate snapshot when present."""
    target_channel = channel.strip().lower()
    target_gate = gate.strip().upper()
    for snapshot in as_list(gate_readiness.get("snapshots")):
        snapshot = as_dict(snapshot)
        if (
            normalize_text(snapshot.get("channel")).strip().lower() == target_channel
            and normalize_text(snapshot.get("gate")).strip().upper() == target_gate
        ):
            return snapshot
    return {}


def issue_sentence(parts: list[str]) -> str:
    """Join short issue fragments into one readable sentence."""
    parts = [part for part in parts if part]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        return f"{parts[0]} and {parts[1]}"
    return f"{', '.join(parts[:-1])}, and {parts[-1]}"


def blocker_action_summary(packet: dict[str, Any], analysis: dict[str, Any]) -> dict[str, Any]:
    """Build a compact top-of-report blocker/action summary."""
    identity = as_dict(packet.get("identity"))
    performance = as_dict(analysis.get("performance_estimation"))
    pricing = as_dict(analysis.get("pricing_analysis"))
    spec_coverage = as_dict(analysis.get("spec_coverage"))
    gate_readiness = as_dict(analysis.get("gate_readiness"))

    suggested = as_dict(pricing.get("suggested_msrp_range"))
    margin_targets = as_dict(pricing.get("margin_targets"))
    amazon_margin = as_dict(margin_targets.get("amazon"))
    amazon_snapshot = gate_snapshot(gate_readiness, "amazon", "G2")

    outlook = normalize_text(performance.get("launch_outlook")).strip().lower() or "mixed"
    confidence = normalize_text(performance.get("confidence")).strip().lower() or "medium"
    posture = normalize_text(performance.get("posture")).strip().lower() or "undetermined"
    positioning = normalize_text(suggested.get("positioning")).strip().lower() or "undetermined"
    evidence_label = normalize_text(amazon_snapshot.get("evidence_label")).strip().lower() or "unknown"

    target_msrp = pricing.get("target_msrp")
    target_vendor_cost = pricing.get("target_vendor_cost")
    margin_conflict = bool(suggested.get("margin_conflict"))
    minimum_margin_safe_price = suggested.get("minimum_margin_safe_price")
    recommended_ceiling = suggested.get("recommended_ceiling")
    notable_gaps = [normalize_text(item) for item in as_list(spec_coverage.get("notable_gaps")) if normalize_text(item)]
    gap_count = len(notable_gaps)
    feature_coverage = [as_dict(item) for item in as_list(spec_coverage.get("feature_coverage"))]
    certification_coverage = [as_dict(item) for item in as_list(spec_coverage.get("certification_coverage"))]
    zero_match_features = [entry for entry in feature_coverage if entry.get("matched_count") == 0]
    zero_match_certs = [entry for entry in certification_coverage if entry.get("matched_count") == 0]
    whitespace_labels = [
        normalize_text(entry.get("label"))
        for entry in zero_match_features
        if normalize_text(entry.get("signal")).strip().lower() == "whitespace"
    ]
    weak_signal_labels = [
        normalize_text(entry.get("label"))
        for entry in feature_coverage
        if entry.get("matched_count", 0) > 0 and normalize_text(entry.get("evidence_strength")).strip().lower() == "weak"
    ]
    zero_match_cert_labels = [normalize_text(entry.get("label")) for entry in zero_match_certs]

    actions: list[dict[str, str]] = []

    if target_msrp is None or target_vendor_cost is None:
        missing_bits = []
        if target_msrp is None:
            missing_bits.append("target MSRP")
        if target_vendor_cost is None:
            missing_bits.append("landed vendor cost")
        actions.append(
            {
                "category": "commercial",
                "issue": "Pricing inputs are incomplete",
                "why": f"Without {issue_sentence(missing_bits)}, the model has to infer key commercial inputs instead of scoring them directly.",
                "action": "Enter a provisional MSRP and landed vendor cost for this row, then rerun the report.",
                "impact": "Price Fit; Margin Viability; Outlook",
            }
        )

    if margin_conflict:
        why = "Current target pricing conflicts with the minimum margin-safe price."
        if minimum_margin_safe_price and recommended_ceiling:
            why = (
                f"The current minimum margin-safe price ({normalize_text(minimum_margin_safe_price)}) "
                f"sits above the recommended market ceiling ({normalize_text(recommended_ceiling)})."
            )
        actions.append(
            {
                "category": "commercial",
                "issue": "Margin conflict is still unresolved",
                "why": why,
                "action": "Lower vendor cost, raise MSRP, or trim feature scope until the target clears the safety floor.",
                "impact": "Margin Viability; Price Fit; Outlook",
            }
        )
    elif positioning == "premium":
        actions.append(
            {
                "category": "commercial",
                "issue": "Target pricing is above the market band",
                "why": "The current MSRP is positioned above the observed comparable range and needs stronger feature justification.",
                "action": "Either lower MSRP or strengthen the differentiator and certification story enough to defend a premium position.",
                "impact": "Price Fit; Outlook",
            }
        )

    if posture == "no_stackline_context":
        actions.append(
            {
                "category": "evidence",
                "issue": "Market growth context is missing",
                "why": "This row is missing usable Stackline market-growth context, which caps how strongly the outlook can score.",
                "action": "Confirm the row is mapped to the correct Stackline segment and refresh the run before using Outlook as a go / no-go signal.",
                "impact": "Market Support; Outlook",
            }
        )

    if gap_count >= 3:
        highlighted_labels = [label for label in whitespace_labels + zero_match_cert_labels if label][:3]
        gap_preview = issue_sentence(highlighted_labels[:2]) or issue_sentence(notable_gaps[:2])
        actions.append(
            {
                "category": "evidence",
                "issue": "Several proposed attributes still need stronger market proof",
                "why": (
                    f"The current caution is being driven by evidence blind spots, not proof the concept is wrong. "
                    f"Right now competitor records do not clearly confirm attributes such as {gap_preview}."
                ),
                "action": (
                    "Treat these as review items, not automatic negatives: keep intentional innovations, "
                    "confirm whether each item is a real customer requirement or compliance need, and only cut it if it adds cost without demand support."
                ),
                "impact": "Confidence; Market Support; Outlook",
            }
        )
    elif gap_count > 0 and outlook != "favorable":
        highlighted_labels = [label for label in whitespace_labels + zero_match_cert_labels if label][:3]
        gap_preview = issue_sentence(highlighted_labels[:2]) or issue_sentence(notable_gaps[:2])
        actions.append(
            {
                "category": "evidence",
                "issue": "A few proposed attributes still need stronger market proof",
                "why": (
                    f"The concept includes attributes that are not yet well confirmed in competitor evidence, including {gap_preview}."
                ),
                "action": (
                    "Decide whether these are intentional innovations, documentation-only blind spots, or true must-have requirements before treating them as blockers."
                ),
                "impact": "Confidence; Market Support",
            }
        )
    elif weak_signal_labels and outlook != "favorable":
        weak_preview = issue_sentence(weak_signal_labels[:2])
        actions.append(
            {
                "category": "evidence",
                "issue": "Some differentiator claims still rest on thin evidence",
                "why": f"The market set only weakly supports items such as {weak_preview}, so the report should treat them as directional rather than proven gaps or advantages.",
                "action": "Use these as demand hypotheses, not reasons to downscore the concept unless stronger competitor evidence later contradicts them.",
                "impact": "Confidence; Outlook",
            }
        )

    if evidence_label in {"low", "none"}:
        actions.append(
            {
                "category": "evidence",
                "issue": "Amazon evidence is still thin",
                "why": "The current Amazon Gate 2 score is supported by limited direct evidence, so the score is less trustworthy than the research confidence alone suggests.",
                "action": "Treat the Amazon G2 score as directional until more direct listing / competitor evidence is collected.",
                "impact": "Amazon Evidence; Confidence",
            }
        )

    if not actions:
        actions.append(
            {
                "category": "execution",
                "issue": "Protect the current favorable read",
                "why": "Current market, pricing, and competitor signals are supportive enough that the concept now reads cleanly.",
                "action": "Lock the vendor quote, confirm the key certifications, and preserve the current feature set unless cost changes materially.",
                "impact": "Outlook; Margin Viability; Execution Readiness",
            }
        )

    actions = actions[:3]
    action_categories = [entry.get("category", "") for entry in actions]
    evidence_heavy = bool(actions) and all(category in {"evidence", "market_context"} for category in action_categories)

    subcategory_label = normalize_text(identity.get("subcategory")).strip().lower()

    if outlook == "favorable":
        overall_read = (
            f"This {'concept in ' + subcategory_label if subcategory_label else 'concept'} currently looks favorable to pursue "
            f"with {confidence} research confidence."
        )
        remaining_watchouts = [entry["issue"].lower() for entry in actions if entry.get("category") != "execution"]
        if remaining_watchouts:
            why_not_stronger = (
                "The concept is already favorable, but the remaining watchouts are "
                f"{issue_sentence(remaining_watchouts)}."
            )
        else:
            why_not_stronger = "The concept is already favorable; the main task now is to protect that position as quotes and compliance details firm up."
    elif outlook == "mixed":
        if evidence_heavy:
            overall_read = (
                "This concept looks viable, and the main constraint right now is incomplete market confirmation rather than a clearly weak product position."
            )
            why_not_stronger = (
                "It is not stronger yet because the report still needs clearer evidence around "
                f"{issue_sentence([entry['issue'].lower() for entry in actions])}."
            )
        else:
            overall_read = (
                "This concept looks viable, but it still needs targeted commercial or evidence cleanup before it becomes a strong go-forward candidate."
            )
            why_not_stronger = (
                "It is not stronger yet because "
                f"{issue_sentence([entry['issue'].lower() for entry in actions])}."
            )
    elif outlook == "cautious":
        if evidence_heavy:
            overall_read = (
                "This concept has enough support to analyze, and the current caution is being driven more by evidence ambiguity than by proof the product is weak."
            )
            why_not_stronger = (
                "The main watchouts right now are "
                f"{issue_sentence([entry['issue'].lower() for entry in actions])}."
            )
        else:
            overall_read = (
                "This concept has enough support to analyze, but the current configuration is still being held back by one or more commercial or evidence blockers."
            )
            why_not_stronger = (
                "The main blockers right now are "
                f"{issue_sentence([entry['issue'].lower() for entry in actions])}."
            )
    else:
        overall_read = "This concept still needs more complete evidence before the report can support a strong directional call."
        why_not_stronger = (
            "The current read is still limited because "
            f"{issue_sentence([entry['issue'].lower() for entry in actions])}."
        )

    table_rows = [
        [f"P{index}", entry["issue"], entry["why"], entry["action"], entry["impact"]]
        for index, entry in enumerate(actions, start=1)
    ]

    return {
        "overall_read": overall_read,
        "why_not_stronger": why_not_stronger,
        "actions": table_rows,
    }


def first_sentence(text: Any) -> str:
    """Return a compact first sentence for executive display."""
    clean = normalize_text(text).strip()
    if not clean:
        return ""
    parts = re.split(r"(?<=[.!?])\s+", clean, maxsplit=1)
    return parts[0]


def extract_research_note_evidence(packet: dict[str, Any]) -> dict[str, str]:
    """Extract structured Step 1 evidence from the research notes block."""
    target_profile = as_dict(packet.get("target_profile"))
    notes = normalize_text(target_profile.get("research_notes"))
    evidence: dict[str, str] = {}

    def clean_fragment(value: str) -> str:
        return value.strip().rstrip(".")

    movement_match = re.search(
        r"Ecommerce movement:\s*([^\n]+)",
        notes,
        flags=re.IGNORECASE,
    )
    if movement_match:
        evidence["ecommerce_movement"] = clean_fragment(movement_match.group(1))

    competitor_match = re.search(
        r"Redshift ecommerce competitor evidence found\s*([^.\n]+)",
        notes,
        flags=re.IGNORECASE,
    )
    if competitor_match:
        evidence["competitor_evidence"] = clean_fragment(competitor_match.group(1))

    overlay_match = re.search(r"(?:Demand confidence score|Product Demand overlay score):\s*([0-9.]+/100)[^\n.]*", notes, flags=re.IGNORECASE)
    if overlay_match:
        evidence["overlay_score"] = clean_fragment(overlay_match.group(0))

    inventory_match = re.search(r"Inventory support:\s*([^\n]+)", notes, flags=re.IGNORECASE)
    if inventory_match:
        evidence["inventory_support"] = clean_fragment(inventory_match.group(1))

    sunco_match = re.search(r"Sunco check:\s*([^\n]+)", notes, flags=re.IGNORECASE)
    if sunco_match:
        evidence["sunco_coverage"] = clean_fragment(sunco_match.group(1))

    product_action_match = re.search(r"Recommended Product Action:\s*([^\n]+)", notes, flags=re.IGNORECASE)
    if product_action_match:
        evidence["product_action"] = clean_fragment(product_action_match.group(1))

    revision_target_match = re.search(r"Revision target SKU:\s*([^\n]+)", notes, flags=re.IGNORECASE)
    if revision_target_match:
        evidence["revision_target_sku"] = clean_fragment(revision_target_match.group(1))

    revision_changes_match = re.search(r"Recommended revision changes:\s*([^\n]+)", notes, flags=re.IGNORECASE)
    if revision_changes_match:
        evidence["revision_changes"] = clean_fragment(revision_changes_match.group(1))

    classification_match = re.search(r"Classification:\s*([^\n]+)", notes, flags=re.IGNORECASE)
    if classification_match and "product_action" not in evidence:
        evidence["product_action"] = clean_fragment(classification_match.group(1))

    recommended_match = re.search(r"Recommended action:\s*([^\n]+)", notes, flags=re.IGNORECASE)
    if recommended_match:
        evidence["recommended_action"] = clean_fragment(recommended_match.group(1))

    review_link_match = re.search(r"Review link:\s*(https?://\S+)", notes, flags=re.IGNORECASE)
    if review_link_match:
        evidence["review_link"] = review_link_match.group(1).strip()

    bias_match = re.search(r"Vendor bias risk:\s*([^.\n]+)", notes, flags=re.IGNORECASE)
    if bias_match:
        evidence["vendor_bias_risk"] = clean_fragment(bias_match.group(1))

    return evidence


def source_link_cell(url: Any, label: str = "Open source") -> dict[str, str] | str:
    """Return a short, clickable source-link payload."""
    text = optional_text(url)
    if not text or text.startswith("stackline://"):
        return ""
    return {"value": label, "hyperlink": text}


def source_link_from_item(item: dict[str, Any], label: str = "Open PDP") -> dict[str, str] | str:
    """Return a short hyperlink payload for a competitor item."""
    url = optional_text(item.get("url"))
    if not url or product_url_issue(url, resolved_source_channel(item)):
        return ""
    return source_link_cell(url, label)


def product_url_issue(url: Any, source_channel: str | None = None) -> str:
    """Return why a URL should not be shown as a verified product PDP."""
    text = optional_text(url)
    if not text or text.startswith("stackline://"):
        return ""
    parsed = urlparse(text)
    host = parsed.netloc.lower().replace("www.", "")
    path_parts = [part for part in parsed.path.strip("/").split("/") if part]
    channel = (source_channel or "").lower()

    if (channel == "amazon" or host.endswith("amazon.com")) and len(path_parts) >= 2 and path_parts[0].lower() in {"dp", "gp"}:
        candidate = path_parts[-1]
        if not re.fullmatch(r"[A-Z0-9]{10}", candidate.upper()):
            return "Amazon URL does not contain a valid ASIN."

    if channel == "home_depot" or host.endswith("homedepot.com"):
        if len(path_parts) >= 2 and path_parts[0].lower() == "p":
            last_part = path_parts[-1]
            if re.fullmatch(r"[A-Z0-9]{10}", last_part.upper()):
                return "Home Depot URL appears to use an Amazon ASIN."
            if len(path_parts) == 2 and not re.fullmatch(r"\d{6,}", last_part):
                return "Home Depot URL appears to be a synthetic non-item URL."

    if channel == "walmart" or host.endswith("walmart.com"):
        if len(path_parts) >= 2 and path_parts[0].lower() == "ip" and not re.fullmatch(r"\d{6,}", path_parts[-1]):
            return "Walmart URL does not contain a numeric item ID."

    return ""


def normalize_opportunity_type(value: Any) -> str:
    """Normalize arbitrary opportunity language into the supporting gap-reason vocabulary."""
    text = normalize_text(value).lower()
    for opportunity_type in OPPORTUNITY_TYPES:
        if opportunity_type.lower() in text:
            return opportunity_type
    if "missing feature" in text or "existing sunco coverage" in text:
        return "Existing Sunco coverage, but missing feature"
    if "feature gap" in text or "possible feature" in text:
        return "Possible feature gap"
    if "revision" in text or "merchandising" in text or "coverage exists" in text:
        return "Product Revision or merchandising review"
    if "strategic outlier" in text or "watchlist" in text:
        return "Strategic outlier / High-output watchlist"
    if "variant" in text or "true gap" in text or "new product" in text:
        return "New variant opportunity"
    return ""


def normalize_product_action(value: Any) -> str:
    """Normalize arbitrary language into the PM-facing action vocabulary."""
    text = normalize_text(value).lower()
    if not text:
        return ""
    if "hold" in text or "do not prioritize" in text or "do not launch" in text:
        return PRODUCT_ACTION_HOLD
    if "concept review" in text or "possible feature gap" in text or "strategic outlier" in text or "watchlist" in text or "partial coverage" in text:
        return PRODUCT_ACTION_CONCEPT_REVIEW
    if "revision" in text or "merchandising" in text or "existing sunco coverage" in text or "coverage exists" in text or "defend/optimize" in text:
        return PRODUCT_ACTION_REVISION
    if "npd" in text or "new product development" in text or "new product" in text or "new variant" in text or "true gap" in text or "amazon stackline opportunity" in text:
        return PRODUCT_ACTION_NPD
    return ""


def recommended_product_action(packet: dict[str, Any], analysis: dict[str, Any] | None = None) -> str:
    """Return the approved PM-facing action label for this concept."""
    identity = as_dict(packet.get("identity"))
    target_profile = as_dict(packet.get("target_profile"))
    evidence = extract_research_note_evidence(packet)
    context_values = [
        evidence.get("product_action"),
        identity.get("ideation_name"),
        target_profile.get("research_notes"),
        evidence.get("recommended_action"),
        evidence.get("sunco_coverage"),
        identity.get("strategy"),
    ]
    for value in context_values:
        action = normalize_product_action(value)
        if action:
            return action
    if analysis and recommendation_priority(analysis) == "Hold / Do Not Prioritize Yet":
        return PRODUCT_ACTION_HOLD
    return PRODUCT_ACTION_CONCEPT_REVIEW


def product_action_rank(value: Any) -> int:
    action = normalize_product_action(value)
    if action:
        return PRODUCT_ACTION_SORT_ORDER[action]
    return 99


def payload_action_sort_key(payload: tuple[int, dict[str, Any], dict[str, Any], dict[str, Any]]) -> tuple[int, int]:
    row_number, packet, analysis, _ = payload
    return (product_action_rank(recommended_product_action(packet, analysis)), row_number)


def opportunity_type_for(packet: dict[str, Any], analysis: dict[str, Any] | None = None) -> str:
    """Return the supporting gap reason for a row."""
    identity = as_dict(packet.get("identity"))
    target_profile = as_dict(packet.get("target_profile"))
    research_notes = normalize_text(target_profile.get("research_notes"))
    evidence = extract_research_note_evidence(packet)
    context_values = [
        identity.get("ideation_name"),
        identity.get("strategy"),
        research_notes,
        evidence.get("recommended_action"),
        evidence.get("sunco_coverage"),
    ]
    for value in context_values:
        normalized = normalize_opportunity_type(value)
        if normalized:
            return normalized

    sunco_coverage = normalize_text(evidence.get("sunco_coverage")).lower()
    has_anchor = bool(normalize_text(identity.get("sunco_reference_sku")))
    if "0 strong" in sunco_coverage or "assortment gap" in sunco_coverage:
        return "New variant opportunity"
    if has_anchor:
        return "Product Revision or merchandising review"
    return "New variant opportunity"


def evidence_strength_for(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Classify the row using the shared evidence-strength language."""
    performance = as_dict(analysis.get("performance_estimation"))
    confidence = normalize_text(performance.get("confidence")).lower()
    outlook = normalize_text(performance.get("launch_outlook")).lower()
    evidence = extract_research_note_evidence(packet)
    overlay_match = re.search(r"([0-9]+(?:\.[0-9]+)?)/100", normalize_text(evidence.get("overlay_score")))
    overlay_score = decimal_from_value(overlay_match.group(1)) if overlay_match else None

    if overlay_score is not None and overlay_score >= Decimal("80"):
        return EVIDENCE_STRENGTH_STRONG
    if any(token in confidence for token in ["high", "strong"]) and outlook == "favorable":
        return EVIDENCE_STRENGTH_STRONG
    if any(token in confidence for token in ["low", "limited", "weak"]) or outlook == "cautious":
        return EVIDENCE_STRENGTH_REVIEW
    return EVIDENCE_STRENGTH_DIRECTIONAL


def match_quality_for_item(item: dict[str, Any]) -> str:
    """Return the shared match-quality label for a competitor item."""
    if not item:
        return MATCH_QUALITY_NOT_COMPARABLE
    if product_url_issue(item.get("url"), resolved_source_channel(item)):
        return MATCH_QUALITY_NOT_COMPARABLE
    confidence = decimal_from_value(item.get("match_confidence"))
    if verification_status(item) == "verified_listing" and confidence is not None and confidence >= Decimal("0.80"):
        return MATCH_QUALITY_APPLES
    if verification_status(item) == "verified_listing":
        return MATCH_QUALITY_SIMILAR
    return MATCH_QUALITY_NOT_COMPARABLE


def link_status_for_item(item: dict[str, Any]) -> str:
    """Return a visible link-status label for manual PM verification."""
    if not item or not optional_text(item.get("url")):
        return LINK_STATUS_MISSING
    if product_url_issue(item.get("url"), resolved_source_channel(item)):
        return LINK_STATUS_INVALID
    if verification_status(item) == "verified_listing":
        return LINK_STATUS_VERIFIED
    return LINK_STATUS_NEEDS_CHECK


def link_status_for_url(url: Any, verified: bool = False) -> str:
    """Return the link-status label for a standalone source URL."""
    if not optional_text(url):
        return LINK_STATUS_MISSING
    return LINK_STATUS_VERIFIED if verified else LINK_STATUS_NEEDS_CHECK


def primary_competitor_item(normalized: dict[str, Any]) -> dict[str, Any]:
    """Return the best source-backed competitor record across preferred channels."""
    return best_verified_item(normalized, AMAZON_CHANNELS) or best_verified_item(normalized, BM_DIRECT_CHANNELS)


def primary_source_link(packet: dict[str, Any], normalized: dict[str, Any]) -> dict[str, str] | str:
    """Return the most useful source link for a top-level PM snapshot row."""
    item = primary_competitor_item(normalized)
    if item:
        return source_link_from_item(item)
    evidence = extract_research_note_evidence(packet)
    return source_link_cell(evidence.get("review_link"), "Open PDP")


def msrp_basis_summary(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Return a compact explanation for the recommended MSRP."""
    pricing = as_dict(analysis.get("pricing_analysis"))
    target_position = as_dict(pricing.get("target_price_position"))
    suggested = as_dict(pricing.get("suggested_msrp_range"))
    pieces: list[str] = []
    percentile = target_position.get("percentile")
    if percentile is not None:
        pieces.append(f"{display_value_for_label('Percentile', percentile)} market price percentile")
    if suggested.get("positioning"):
        pieces.append(normalize_text(suggested.get("positioning")))
    if suggested.get("margin_conflict"):
        pieces.append(f"Margin read: {normalize_text(suggested.get('margin_conflict'))}")
    if pieces:
        return "; ".join(pieces)
    evidence = extract_research_note_evidence(packet)
    if evidence.get("review_link"):
        return "Market MSRP is anchored to the verified Step 1 competitor listing until broader comps are refreshed."
    return "Recommended MSRP is based on Step 2/3 pricing analysis."


def money_or_blank(value: Any) -> str:
    """Format a money-like value for compact text."""
    if value is None:
        return ""
    return display_value_for_label("Price", value)


def number_from_text(text: Any, unit_pattern: str) -> Decimal | None:
    """Find the first numeric value before a unit pattern."""
    match = re.search(rf"\b([0-9][0-9,]*(?:\.\d+)?)\s*{unit_pattern}\b", normalize_text(text), flags=re.IGNORECASE)
    if not match:
        return None
    return decimal_from_value(match.group(1))


def cct_display_from_context(context: str) -> str:
    """Return a readable CCT summary from concept context."""
    values = sorted(set(re.findall(r"\b(?:27|30|35|40|50|60|65|70)00\s*K\b", context, flags=re.IGNORECASE)))
    if values:
        normalized_values = [value.upper().replace(" ", "") if value.upper().endswith("K") else f"{value}K" for value in values]
        return "/".join(normalized_values)
    short = re.search(r"\b([234567])K\b", context, flags=re.IGNORECASE)
    if short:
        return f"{short.group(1)}000K"
    return ""


def product_difference_summary(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Summarize how the new concept differs from the closest Sunco anchor."""
    identity = as_dict(packet.get("identity"))
    reference = as_dict(packet.get("reference_baseline"))
    target_profile = as_dict(packet.get("target_profile"))
    electrical = as_dict(target_profile.get("electrical"))
    physical = as_dict(target_profile.get("physical"))
    reference_anchor = as_dict(analysis.get("reference_anchor_context"))

    anchor_context = " ".join(
        normalize_text(value)
        for value in [
            identity.get("sunco_reference_sku"),
            reference.get("title"),
            reference.get("product_type"),
        ]
        if normalize_text(value)
    )
    target_context = context_text_for_packet(packet)

    differences: list[str] = []
    target_lumens = number_from_text(electrical.get("lumens_target"), r"(?:lm|lumens?)")
    anchor_lumens = number_from_text(anchor_context, r"(?:lm|lumens?)")
    if target_lumens and anchor_lumens and target_lumens != anchor_lumens:
        direction = "higher" if target_lumens > anchor_lumens else "lower"
        differences.append(f"{direction} light output ({format_decimal(target_lumens, 0)}lm vs {format_decimal(anchor_lumens, 0)}lm anchor)")

    target_wattage = first_wattage_token(target_context)
    anchor_wattage = first_wattage_token(anchor_context)
    if target_wattage and anchor_wattage and target_wattage != anchor_wattage:
        differences.append(f"different load class ({target_wattage} target vs {anchor_wattage} anchor)")

    target_cct = cct_display_from_context(target_context)
    anchor_cct = cct_display_from_context(anchor_context)
    if target_cct and anchor_cct and target_cct != anchor_cct:
        cct_phrase = "selectable CCT" if as_dict(target_profile.get("electrical")).get("selectable_cct") else target_cct
        differences.append(f"{cct_phrase} vs {anchor_cct} anchor")

    feature_watchlist = [normalize_text(item) for item in as_list(target_profile.get("feature_watchlist")) if normalize_text(item)]
    if feature_watchlist:
        differences.append(f"feature watchlist: {', '.join(feature_watchlist[:3])}")

    target_price = as_dict(analysis.get("pricing_analysis")).get("target_msrp")
    anchor_price = reference.get("listing_price")
    if target_price and anchor_price:
        target_number = decimal_from_value(target_price)
        anchor_number = decimal_from_value(anchor_price)
        if target_number and anchor_number and target_number != anchor_number:
            direction = "premium" if target_number > anchor_number else "value"
            differences.append(f"{direction} price position ({money_or_blank(target_price)} target vs {money_or_blank(anchor_price)} anchor)")

    if differences:
        return "; ".join(differences[:4]) + "."

    role = reference_anchor.get("primary_use")
    if role:
        return f"Closest Sunco item is mainly a category anchor; exact spec difference needs PM review. {normalize_text(role)}"
    return "No close Sunco/NSL anchor was available for this row."


def revision_target_sku(packet: dict[str, Any], analysis: dict[str, Any] | None = None) -> str:
    """Return the existing SKU/family that should be revised for Revision rows."""
    if recommended_product_action(packet, analysis) != PRODUCT_ACTION_REVISION:
        return ""
    identity = as_dict(packet.get("identity"))
    evidence = extract_research_note_evidence(packet)
    for value in [evidence.get("revision_target_sku"), identity.get("sunco_reference_sku")]:
        text = normalize_text(value)
        if text and not text.lower().startswith("tbd"):
            return text
    return "matched active Sunco SKU/family"


def revision_change_summary(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Return the concrete PM-facing change request for Revision rows."""
    if recommended_product_action(packet, analysis) != PRODUCT_ACTION_REVISION:
        return ""
    evidence = extract_research_note_evidence(packet)
    target = revision_target_sku(packet, analysis)
    for value in [evidence.get("revision_changes"), evidence.get("recommended_action")]:
        text = normalize_text(value)
        if text:
            return text
    difference = product_difference_summary(packet, analysis)
    if difference and "No close Sunco" not in difference:
        return f"Revise {target}: {difference}"
    return f"Revise {target}: compare the active SKU/family against the source links and apply the missing competitor-supported feature, spec, or merchandising change."


def product_fit_classification(packet: dict[str, Any]) -> str:
    """Return the supporting gap reason."""
    return opportunity_type_for(packet)


def related_sunco_sales_trend(reference: dict[str, Any]) -> str:
    """Return related Sunco sales trend when the packet has trend fields."""
    for key in ["sales_growth_pct", "revenue_growth_pct", "units_growth_pct", "shopify_revenue_growth_pct", "amazon_revenue_growth_pct"]:
        value = reference.get(key)
        number = decimal_from_value(value)
        if number is None:
            continue
        if number > Decimal("5"):
            return f"Growing ({display_value_for_label('Growth %', value)})"
        if number < Decimal("-5"):
            return f"Declining ({display_value_for_label('Growth %', value)})"
        return f"Flat ({display_value_for_label('Growth %', value)})"

    shopify = reference.get("shopify_revenue_12mo")
    amazon = reference.get("amazon_revenue_12mo")
    if shopify is not None or amazon is not None:
        return (
            "Monthly trend was not refreshed for this run; 12-month anchor revenue is available "
            f"(Shopify {money_or_blank(shopify) or 'n/a'}, Amazon {money_or_blank(amazon) or 'n/a'})."
        )
    return "Monthly trend was not refreshed for this run."


def related_sales_caution(reference: dict[str, Any]) -> str:
    """Explain how to treat weak or missing related-Sunco sales support."""
    trend = related_sunco_sales_trend(reference)
    lowered = trend.lower()
    if "declining" in lowered:
        return "Related anchor appears declining; use it as category context, not as proof of upside, until the decline driver is known."
    if "monthly trend was not refreshed" in lowered:
        return "12-month revenue is available, but monthly trend was not refreshed for this run."
    return "No weak/declining related-sales caution surfaced from available packet fields."


def related_sales_slide_note(reference: dict[str, Any]) -> str:
    """Return a short related-sales note that fits in the slide table."""
    trend = related_sunco_sales_trend(reference)
    lowered = trend.lower()
    if "trend not available" in lowered:
        return "Trend unavailable; 12mo context only."
    if "declining" in lowered:
        return "Declining; explain cause before using as support."
    return trend


def market_momentum_summary(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Build a concise market-momentum statement."""
    performance = as_dict(analysis.get("performance_estimation"))
    snapshot = as_dict(performance.get("market_snapshot"))
    evidence = extract_research_note_evidence(packet)
    pieces: list[str] = []

    segment = normalize_text(snapshot.get("segment_name"))
    sales = snapshot.get("segment_retail_sales")
    sales_growth = snapshot.get("segment_sales_growth_pct")
    traffic_growth = snapshot.get("segment_traffic_growth_pct")
    if traffic_growth is None:
        for channel_data in as_dict(as_dict(performance.get("channel_comparison")).get("channels")).values():
            channel_data = as_dict(channel_data)
            if channel_data.get("traffic_growth_pct") is not None:
                traffic_growth = channel_data.get("traffic_growth_pct")
                break
    if segment:
        market_text = f"{segment}"
        if sales is not None:
            market_text += f" {display_value_for_label('Sales', sales)} retail sales"
        if sales_growth is not None:
            market_text += f", {display_value_for_label('Growth %', sales_growth)} sales growth"
        pieces.append(market_text)

    movement = evidence.get("ecommerce_movement")
    if movement:
        pieces.append(f"Redshift ecommerce movement: {movement}")
    if traffic_growth is not None:
        pieces.append(f"Stackline traffic growth {display_value_for_label('Growth %', traffic_growth)}")
    return "; ".join(pieces) + "." if pieces else "Market momentum is directional; no structured movement summary was available."


def channel_strategy(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Recommend a channel posture from available evidence."""
    identity = as_dict(packet.get("identity"))
    performance = as_dict(analysis.get("performance_estimation"))
    evidence = extract_research_note_evidence(packet)
    name = normalize_text(identity.get("ideation_name")).lower()
    outlook = normalize_text(performance.get("launch_outlook")).lower()
    notes = normalize_text(as_dict(packet.get("target_profile")).get("research_notes")).lower()
    amazon_snapshot = gate_snapshot(as_dict(analysis.get("gate_readiness")), "amazon", "G2")
    amazon_score = decimal_from_value(amazon_snapshot.get("weighted_score"))

    if "shopify/front-end launch review first" in notes:
        return "Shopify-first"
    if "amazon stackline" in name:
        return "Amazon-first"
    if outlook == "favorable" and amazon_score is not None and amazon_score >= Decimal("7") and evidence.get("ecommerce_movement"):
        return "Amazon + Shopify"
    if outlook in {"mixed", "cautious"}:
        return "Staged test"
    return "Amazon + Shopify"


def certification_summary(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Return required certifications from Step 2 or observed Section F coverage."""
    business = as_dict(as_dict(packet.get("target_profile")).get("business_case"))
    explicit = normalize_text(business.get("certifications"))
    if explicit:
        return explicit
    cert_rows = as_list(as_dict(analysis.get("spec_coverage")).get("certification_coverage"))
    labels = [normalize_text(as_dict(row).get("label")) for row in cert_rows if normalize_text(as_dict(row).get("label"))]
    if labels:
        return ", ".join(labels[:4])
    return "No required certification captured in current Step 2 row."


def step1_demand_evidence_rows(packet: dict[str, Any], analysis: dict[str, Any]) -> list[list[Any]]:
    """Build source-backed rows for the Step 1 demand evidence block."""
    performance = as_dict(analysis.get("performance_estimation"))
    snapshot = as_dict(performance.get("market_snapshot"))
    evidence = extract_research_note_evidence(packet)
    rows: list[list[Any]] = []

    if evidence.get("ecommerce_movement"):
        rows.append(["Ecommerce inventory movement", evidence["ecommerce_movement"], "Redshift ecommerce competitor snapshot", ""])
    if evidence.get("inventory_support"):
        rows.append(["Matched inventory support", evidence["inventory_support"], "Redshift ecommerce competitor snapshot", ""])
    if evidence.get("competitor_evidence"):
        rows.append(["Competitor PDP coverage", evidence["competitor_evidence"], "Redshift ecommerce competitor snapshot", ""])
    if evidence.get("review_link"):
        rows.append(["Primary competitor PDP", evidence["review_link"], "Step 1 verified review link", {"value": "Open listing", "hyperlink": evidence["review_link"]}])
    if evidence.get("sunco_coverage"):
        rows.append(["Sunco active-catalog coverage", evidence["sunco_coverage"], "Postgres product/catalog snapshot", ""])
    if snapshot:
        rows.append(
            [
                "Stackline market segment",
                market_momentum_summary(packet, analysis),
                normalize_text(snapshot.get("segment_name")) or "Stackline segment",
                "",
            ]
        )
    if evidence.get("vendor_bias_risk"):
        rows.append(["Evidence caveat", f"Vendor bias risk: {evidence['vendor_bias_risk']}", "Step 1 demand evidence", ""])
    return rows


def leadership_brief_rows(
    row_number: int,
    packet: dict[str, Any],
    analysis: dict[str, Any],
    normalized: dict[str, Any],
    action_summary: dict[str, Any],
) -> list[tuple[str, Any]]:
    """Build the screenshot-friendly top-of-sheet leadership brief."""
    identity = as_dict(packet.get("identity"))
    reference = as_dict(packet.get("reference_baseline"))
    performance = as_dict(analysis.get("performance_estimation"))
    pricing = as_dict(analysis.get("pricing_analysis"))
    normalized_summary = as_dict(normalized.get("summary"))
    amazon_snapshot = gate_snapshot(as_dict(analysis.get("gate_readiness")), "amazon", "G2")
    action = recommended_product_action(packet, analysis)

    return [
        ("Recommended Product Action", action),
        ("Concept Tracking Name", concept_tracking_name(packet)),
        ("Naming Basis", "SKU Decoder product type + Step 2 target specs; tracking name only, not a final SKU."),
        ("Gap Reason", product_fit_classification(packet)),
        ("Channel Strategy", channel_strategy(packet, analysis)),
        ("Closest Sunco / NSL Anchor", f"{normalize_text(identity.get('sunco_reference_sku'))} - {normalize_text(reference.get('title'))}"),
        ("Difference vs Current Sunco", product_difference_summary(packet, analysis)),
        ("Related Sunco Sales Trend", related_sunco_sales_trend(reference)),
        ("Related Sales Caution", related_sales_caution(reference)),
        ("Why Worth Reviewing Now", market_momentum_summary(packet, analysis)),
        ("Target MSRP", pricing.get("target_msrp")),
        ("Target Vendor Cost", pricing.get("target_vendor_cost")),
        ("Amazon G2 / Evidence", f"{display_value_for_label('Score', amazon_snapshot.get('weighted_score')) or 'n/a'} / {normalize_text(as_dict(amazon_snapshot.get('evidence_confidence')).get('label')) or 'n/a'}"),
        ("Competitor Records", f"{normalized_summary.get('verified_listing_count', 0)} verified listings; {normalized_summary.get('inferred_competitor_count', 0)} inferred competitors"),
        ("Required Certifications", certification_summary(packet, analysis)),
        ("Leadership Watchout", action_summary.get("why_not_stronger")),
    ]


def compact_competitor_read(item: dict[str, Any]) -> str:
    """Return a short competitor proof statement for PM review."""
    if not item:
        return "No direct competitor PDP was selected in this report."
    pieces = [
        normalize_text(item.get("brand")),
        listing_identifier(item),
        clean_slide_text(item.get("product_title"), 80),
    ]
    price = display_value_for_label("Price", item.get("price"))
    if price:
        pieces.append(price)
    return " | ".join(piece for piece in pieces if piece)


def product_identifier(item: dict[str, Any]) -> str:
    """Return the SKU/model/ASIN value without the channel prefix."""
    if not item:
        return ""
    source_channel = resolved_source_channel(item)
    sku = optional_text(item.get("sku"))
    model = optional_text(item.get("model_number"))
    if source_channel == "amazon" and sku:
        return sku
    return model or sku or listing_identifier(item)


def directional_identifier(item: dict[str, Any]) -> str:
    """Return an identifier that does not imply an unverified retailer PDP exists."""
    issue = product_url_issue(item.get("url"), resolved_source_channel(item))
    if verification_status(item) == "verified_listing" and not issue:
        return product_identifier(item)
    value = (
        optional_text(item.get("model_number"))
        or optional_text(item.get("retailer_sku"))
        or optional_text(item.get("sku"))
        or listing_identifier(item)
    )
    return f"Unverified ID {value}" if value else ""


def directional_channel_label(item: dict[str, Any]) -> str:
    """Return a channel label that separates discovery seeds from verified PDPs."""
    issue = product_url_issue(item.get("url"), resolved_source_channel(item))
    if verification_status(item) == "verified_listing" and not issue:
        return source_channel_label(item)
    if issue:
        return "Unverified retailer seed"
    source = discovery_source_label(item)
    return f"{source} discovery seed" if source else "Discovery seed"


def directional_supporting_source_label(item: dict[str, Any]) -> str:
    """Return source context for unverified rows without claiming a PDP exists."""
    issue = product_url_issue(item.get("url"), resolved_source_channel(item))
    if issue:
        return "PDP not verified"
    return discovery_source_label(item)


def product_pack_size(item: dict[str, Any]) -> str:
    """Return pack/size context for competitor tables."""
    if not item:
        return ""
    parts = []
    pack_quantity = item.get("pack_quantity")
    if pack_quantity not in (None, "", []):
        parts.append(f"Pack {display_value_for_label('Count', pack_quantity)}")
    for key in ["size", "size_form_factor", "length", "dimensions"]:
        value = item.get(key)
        if value not in (None, "", []):
            parts.append(normalize_text(value))
            break
    return "; ".join(parts)


def product_key_specs(item: dict[str, Any]) -> str:
    """Return short spec context for product-level competitor rows."""
    if not item:
        return ""
    parts: list[str] = []
    for label, key in [
        ("W", "wattage"),
        ("lm", "lumens"),
        ("CCT", "cct"),
        ("V", "voltage"),
        ("CRI", "cri"),
        ("Features", "features"),
        ("Certs", "certifications"),
    ]:
        value = item.get(key)
        if value not in (None, "", []):
            parts.append(f"{label}: {normalize_text(value)}")
    return "; ".join(parts[:5])


def competitor_name(item: dict[str, Any]) -> str:
    """Return the competitor/brand label for tables."""
    return (
        normalize_text(as_dict(item).get("brand"))
        or normalize_text(as_dict(item).get("approved_competitor"))
        or source_channel_label(item)
    )


def recommendation_priority(analysis: dict[str, Any]) -> str:
    """Return the move-forward read for PM-facing action tables."""
    outlook = normalize_text(as_dict(analysis.get("performance_estimation")).get("launch_outlook")).lower()
    if outlook == "favorable":
        return "Strong Move-Forward Candidate"
    if outlook == "cautious":
        return "Hold / Do Not Prioritize Yet"
    return "Review Further Before Moving Forward"


def decision_read_label(analysis: dict[str, Any]) -> str:
    """Return a short go-forward decision label."""
    priority = recommendation_priority(analysis)
    if priority == "Strong Move-Forward Candidate":
        return "Pursue"
    if priority == "Hold / Do Not Prioritize Yet":
        return "Hold for now"
    return "Review further"


def opportunity_type_reason(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Explain the opportunity type without exposing backend vocabulary logic."""
    evidence = extract_research_note_evidence(packet)
    coverage = normalize_text(evidence.get("sunco_coverage"))
    if coverage:
        return coverage
    return product_difference_summary(packet, analysis)


def source_label_or_blank(source: Any) -> str:
    """Return a compact non-link source label."""
    text = normalize_text(source)
    if not text:
        return ""
    return text


def observed_market_price_range(pricing: dict[str, Any]) -> str:
    """Return a compact market-price range from available benchmarks."""
    suggested = as_dict(pricing.get("suggested_msrp_range"))
    floor = suggested.get("observed_unit_price_floor") or suggested.get("recommended_floor")
    ceiling = suggested.get("observed_unit_price_ceiling") or suggested.get("recommended_ceiling")
    median = as_dict(pricing.get("unit_price_benchmarks")).get("median") or as_dict(pricing.get("price_benchmarks")).get("median")
    parts = []
    if floor is not None:
        parts.append(f"Low/P25 {display_value_for_label('Price', floor)}")
    if median is not None:
        parts.append(f"Median {display_value_for_label('Price', median)}")
    if ceiling is not None:
        parts.append(f"High/P75 {display_value_for_label('Price', ceiling)}")
    return "; ".join(parts)


def margin_read(pricing: dict[str, Any]) -> str:
    """Return a PM-facing margin read."""
    suggested = as_dict(pricing.get("suggested_msrp_range"))
    if suggested.get("margin_conflict"):
        return "Margin risk"
    minimum_margin_safe = decimal_from_value(suggested.get("minimum_margin_safe_price"))
    target_msrp = decimal_from_value(pricing.get("target_msrp"))
    if minimum_margin_safe is not None and target_msrp is not None:
        if target_msrp < minimum_margin_safe:
            return "Below margin floor"
        if target_msrp <= minimum_margin_safe * Decimal("1.05"):
            return "Near margin floor"
        return "Meets landed margin target"
    return "Needs quote before price can be trusted"


def price_position_read(pricing: dict[str, Any]) -> str:
    """Return a compact price-position read."""
    target_position = as_dict(pricing.get("target_price_position"))
    percentile = target_position.get("percentile")
    bucket = normalize_text(target_position.get("bucket")).replace("_", " ")
    if percentile is not None and bucket:
        return f"{display_value_for_label('Percentile', percentile)} percentile; {bucket}"
    return normalize_text(as_dict(pricing.get("suggested_msrp_range")).get("positioning"))


def channel_strategy_label(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Return controlled PM-facing channel labels."""
    text = channel_strategy(packet, analysis).lower()
    if "shopify" in text and "amazon" in text and ("both" in text or "+" in text):
        return "Amazon + Shopify"
    if "shopify" in text and ("first" in text or "front-end" in text):
        return "Shopify-first"
    if "amazon" in text and "first" in text:
        return "Amazon-first"
    if "staged" in text:
        return "Staged test"
    if "do not" in text or "hold" in text:
        return "Do not launch yet"
    if "both" in text or ("amazon" in text and "shopify" in text):
        return "Amazon + Shopify"
    return "Staged test"


def channel_reason(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Explain the channel label in PM-facing language."""
    label = channel_strategy_label(packet, analysis)
    evidence = extract_research_note_evidence(packet)
    if label == "Shopify-first":
        return "Competitor ecommerce movement is stronger than current Amazon proof."
    if label == "Amazon-first":
        return "Amazon demand proof is the strongest available channel signal."
    if label == "Amazon + Shopify":
        return "Both Amazon and ecommerce/front-end evidence support a broader launch read."
    if label == "Do not launch yet":
        return "Demand or margin support is not strong enough yet."
    if evidence.get("ecommerce_movement"):
        return "Evidence is mixed; use a limited test before broad launch."
    return "Channel evidence is still directional."


def pm_decision_snapshot_rows(
    row_number: int,
    packet: dict[str, Any],
    analysis: dict[str, Any],
    normalized: dict[str, Any],
    action_summary: dict[str, Any],
) -> list[list[Any]]:
    """Build the compact PM-facing decision table at the top of each row sheet."""
    identity = as_dict(packet.get("identity"))
    reference = as_dict(packet.get("reference_baseline"))
    pricing = as_dict(analysis.get("pricing_analysis"))
    product_action = recommended_product_action(packet, analysis)
    gap_reason = opportunity_type_for(packet, analysis)
    evidence_strength = evidence_strength_for(packet, analysis)
    competitor = primary_competitor_item(normalized)
    competitor_match = match_quality_for_item(competitor)
    competitor_link = source_link_from_item(competitor) if competitor else primary_source_link(packet, normalized)
    evidence = extract_research_note_evidence(packet)
    pricing_link = source_link_cell(evidence.get("review_link"), "Open price source")

    rows = [
        [
            "Concept Tracking Name",
            concept_tracking_name(packet),
            evidence_strength,
            "Working concept name based on target product type, wattage, CCT, and feature set. Not a final SKU.",
            "",
        ],
        [
            "Recommended Product Action",
            product_action,
            evidence_strength,
            first_sentence(action_summary.get("overall_read")) or "Use the controlled action label to decide whether this moves to NPD, Revision, Concept Review, or Hold.",
            "",
        ],
        [
            "Gap Reason",
            gap_reason,
            evidence_strength,
            opportunity_type_reason(packet, analysis),
            "",
        ],
        [
            "Recommended MSRP",
            pricing.get("target_msrp"),
            evidence_strength,
            msrp_basis_summary(packet, analysis),
            pricing_link,
        ],
        [
            "Target Vendor Cost",
            pricing.get("target_vendor_cost"),
            evidence_strength,
            "Landed cost ceiling from the 50-55% pre-ads gross-margin policy midpoint.",
            "",
        ],
        [
            "Sunco Gap",
            gap_reason,
            evidence_strength,
            evidence.get("sunco_coverage") or product_difference_summary(packet, analysis),
            "",
        ],
        [
            "Closest Sunco / NSL Anchor",
            f"{normalize_text(identity.get('sunco_reference_sku'))} - {normalize_text(reference.get('title'))}".strip(" -"),
            evidence_strength,
            related_sunco_sales_trend(reference),
            "",
        ],
        [
            "Competitor Proof",
            compact_competitor_read(competitor),
            evidence_strength,
            f"{competitor_match}; {item_demand_signal(competitor) if competitor else normalize_text(evidence.get('ecommerce_movement'))}",
            competitor_link,
        ],
        [
            "Link Status",
            link_status_for_item(competitor) if competitor else link_status_for_url(evidence.get("review_link"), verified=True),
            evidence_strength,
            "PMs should use this visible link to manually confirm product, price, and attributes before presenting.",
            competitor_link,
        ],
        [
            "Channel Strategy",
            channel_strategy_label(packet, analysis),
            evidence_strength,
            channel_reason(packet, analysis),
            "",
        ],
        [
            "Main Watchout",
            first_sentence(action_summary.get("why_not_stronger")),
            EVIDENCE_STRENGTH_REVIEW if action_summary.get("why_not_stronger") else evidence_strength,
            "Use this as the first follow-up before leadership review, PRD build, or vendor RFQ.",
            "",
        ],
    ]
    if product_action == PRODUCT_ACTION_REVISION:
        rows[2:2] = [
            [
                "Revision Target SKU",
                revision_target_sku(packet, analysis),
                evidence_strength,
                "Existing Sunco/NSL SKU or active family the PM should revise instead of creating a new part number by default.",
                "",
            ],
            [
                "Recommended Revision Changes",
                revision_change_summary(packet, analysis),
                evidence_strength,
                "Specific rolling-change, feature-add, or merchandising change to evaluate before RFQ.",
                "",
            ],
        ]
    return rows


def write_pm_decision_snapshot(
    ws,
    row: int,
    row_number: int,
    packet: dict[str, Any],
    analysis: dict[str, Any],
    normalized: dict[str, Any],
    action_summary: dict[str, Any],
) -> int:
    """Write the compact PM decision snapshot at the top of the row sheet."""
    return write_table(
        ws,
        row,
        "PM Decision Snapshot",
        ["Field", "Recommendation", "Evidence Strength", "PM Read / Evidence", "Source Link"],
        pm_decision_snapshot_rows(row_number, packet, analysis, normalized, action_summary),
    )


DEMAND_PROOF_HEADERS = [
    "Competitor",
    "Product Name",
    "SKU / Model / ASIN",
    "Price",
    "Demand Volume",
    "Demand Rate / Share",
    "Demand Recency / Window",
    "Product Link",
    "Match Quality",
    "PM Read",
]
SUNCO_COVERAGE_HEADERS = [
    "Sunco / NSL SKU",
    "Product Name",
    "Shopify Revenue",
    "Shopify Units Sold",
    "Amazon Revenue",
    "Amazon Units Sold",
    "Total Revenue",
    "Coverage Read",
    "Gap vs Ideation",
    "Product Link / Source",
]
COMPETITOR_COMPARABLE_HEADERS = [
    "Channel",
    "Competitor",
    "Product Name",
    "SKU / Model / ASIN",
    "Price",
    "Pack / Size",
    "Key Specs",
    "Demand Volume",
    "Demand Rate / Share",
    "Demand Recency / Window",
    "Match Quality",
    "Product Link",
    "PM Read",
]
SOURCE_DETAIL_HEADERS = [
    "Brand",
    "Product Name",
    "SKU / Model / ASIN",
    "Channel",
    "Price",
    "Key Specs",
    "Demand Volume",
    "Demand Rate / Share",
    "Demand Recency / Window",
    "Match Quality",
    "Product Link",
    "Link Status",
    "Source Basis",
]
APPROVED_SOURCE_COVERAGE_HEADERS = [
    "Approved Competitor",
    "Domain",
    "Tier",
    "Scope",
    "Current Snapshot Status",
    "Warehouse Coverage",
    "Discovery Link",
    "Next Step",
]
PRICING_CHANNEL_HEADERS = [
    "Recommended MSRP",
    "Target Vendor Cost",
    "Observed Market Price Range",
    "Price Position",
    "Margin Read",
    "Recommended Channel",
    "Channel Reason",
    "Source Link",
]
WATCHOUT_HEADERS = [
    "Watchout",
    "Risk Level",
    "Why It Matters",
    "Next Checkpoint",
    "Recommended Follow-Up",
    "Source / Evidence",
]


def top_verified_items(normalized: dict[str, Any], channels: set[str], limit: int) -> list[dict[str, Any]]:
    """Return a small, PM-readable set of verified competitor items."""
    rows: list[dict[str, Any]] = []
    for item in sort_candidates(as_list(normalized.get("items"))):
        if resolved_source_channel(item) not in channels:
            continue
        if verification_status(item) != "verified_listing":
            continue
        if product_url_issue(item.get("url"), resolved_source_channel(item)):
            continue
        rows.append(item)
        if len(rows) >= limit:
            break
    return rows


def demand_proof_rows(packet: dict[str, Any], analysis: dict[str, Any], normalized: dict[str, Any]) -> list[list[Any]]:
    """Build product-level demand proof rows."""
    evidence = extract_research_note_evidence(packet)
    rows: list[list[Any]] = []
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in verified_scraped_competitor_items(packet, limit=3) + top_verified_items(normalized, AMAZON_CHANNELS | BM_DIRECT_CHANNELS, 3):
        key = optional_text(item.get("url")) or product_identifier(item) or clean_slide_text(item.get("product_title"), 120)
        if key in seen:
            continue
        seen.add(key)
        items.append(item)
        if len(items) >= 3:
            break
    for item in items:
        rows.append(
            [
                competitor_name(item),
                clean_slide_text(item.get("product_title"), 120),
                product_identifier(item),
                item.get("price"),
                *item_demand_signal_parts(item),
                source_link_from_item(item),
                match_quality_for_item(item),
                "Product-level demand support; cite when the PDP visibly matches the target spec.",
            ]
        )
    if not rows:
        rows.append(
            [
                "",
                "",
                "",
                "",
                *demand_signal_parts_from_text(evidence.get("ecommerce_movement") or "No structured demand proof captured."),
                "",
                MATCH_QUALITY_NOT_COMPARABLE,
                "No product-level demand proof yet; use only as directional context.",
            ]
        )
    return rows


def sunco_coverage_decision_rows(packet: dict[str, Any], analysis: dict[str, Any], normalized: dict[str, Any]) -> list[list[Any]]:
    """Summarize Sunco coverage using channel revenue and units."""
    identity = as_dict(packet.get("identity"))
    reference = as_dict(packet.get("reference_baseline"))
    evidence = extract_research_note_evidence(packet)
    shopify_revenue = decimal_from_value(reference.get("shopify_revenue_12mo")) or Decimal("0")
    amazon_revenue = decimal_from_value(reference.get("amazon_revenue_12mo")) or Decimal("0")
    total_revenue = shopify_revenue + amazon_revenue
    trend_note = related_sunco_sales_trend(reference)
    if "trend not available" in trend_note.lower():
        trend_note = "12-month revenue is available, but monthly trend was not refreshed for this run."
    return [
        [
            identity.get("sunco_reference_sku"),
            reference.get("title"),
            reference.get("shopify_revenue_12mo"),
            reference.get("shopify_units_12mo"),
            reference.get("amazon_revenue_12mo"),
            reference.get("amazon_units_12mo"),
            float(total_revenue),
            trend_note,
            evidence.get("sunco_coverage") or product_difference_summary(packet, analysis),
            source_label_or_blank(reference.get("reference_data_source")),
        ]
    ]


def competitor_decision_rows(packet: dict[str, Any], analysis: dict[str, Any], normalized: dict[str, Any]) -> list[list[Any]]:
    """Summarize top competitor comparables with exact product links."""
    rows: list[list[Any]] = []
    for item in top_verified_items(normalized, AMAZON_CHANNELS | BM_DIRECT_CHANNELS, 5):
        match_quality = match_quality_for_item(item)
        pm_read = "Use as primary comp" if match_quality == MATCH_QUALITY_APPLES else "Use as directional price/spec comp"
        rows.append(
            [
                source_channel_label(item),
                competitor_name(item),
                clean_slide_text(item.get("product_title"), 120),
                product_identifier(item),
                item.get("price"),
                product_pack_size(item),
                product_key_specs(item),
                *item_demand_signal_parts(item),
                match_quality,
                source_link_from_item(item),
                pm_read,
            ]
        )
    if not rows:
        rows.append(["", "", "No verified competitor comparables were captured.", "", "", "", "", "", "", MATCH_QUALITY_NOT_COMPARABLE, "", "Do not cite until a product link is verified."])
    return rows


def directional_competitor_rows(packet: dict[str, Any], analysis: dict[str, Any], normalized: dict[str, Any]) -> list[list[Any]]:
    """Return weaker or inferred comparables for discovery only."""
    rows: list[list[Any]] = []
    for item in sort_candidates(as_list(normalized.get("items"))):
        if len(rows) >= 5:
            break
        if verification_status(item) == "verified_listing" and match_quality_for_item(item) == MATCH_QUALITY_APPLES:
            continue
        issue = product_url_issue(item.get("url"), resolved_source_channel(item))
        verified = verification_status(item) == "verified_listing" and not issue
        link = source_link_from_item(item) if verified else ""
        pm_read = (
            "Directional only; source product PDP is verified, but specs are not apples-to-apples."
            if verified
            else "Discovery only; needs a verified product PDP before citation."
        )
        if issue:
            pm_read = f"Discovery only; {issue} Needs a verified product PDP before citation."
        rows.append(
            [
                directional_channel_label(item),
                competitor_name(item),
                clean_slide_text(item.get("product_title"), 120),
                directional_identifier(item),
                item.get("price"),
                product_pack_size(item),
                product_key_specs(item),
                *item_demand_signal_parts(item),
                match_quality_for_item(item),
                link,
                pm_read,
            ]
        )
    if not rows:
        rows.append(["", "", "No directional comparables captured.", "", "", "", "", "", "", MATCH_QUALITY_NOT_COMPARABLE, "", "No secondary comps available in this run."])
    return rows


def pricing_channel_decision_rows(packet: dict[str, Any], analysis: dict[str, Any], normalized: dict[str, Any]) -> list[list[Any]]:
    """Summarize MSRP, cost, and channel positioning."""
    pricing = as_dict(analysis.get("pricing_analysis"))
    evidence = extract_research_note_evidence(packet)
    price_link = source_link_cell(evidence.get("review_link"), "Open price source")
    return [
        [
            pricing.get("target_msrp"),
            pricing.get("target_vendor_cost"),
            observed_market_price_range(pricing),
            price_position_read(pricing),
            margin_read(pricing),
            channel_strategy_label(packet, analysis),
            channel_reason(packet, analysis),
            price_link,
        ]
    ]


def risk_level_from_action(issue: str, analysis: dict[str, Any]) -> str:
    """Map action/watchout language into PM-facing risk levels."""
    lowered = normalize_text(issue).lower()
    if recommendation_priority(analysis) == "Hold / Do Not Prioritize Yet":
        return "High"
    if any(term in lowered for term in ["margin", "cost", "quote", "certification", "competitor match"]):
        return "High"
    if any(term in lowered for term in ["evidence", "pricing", "sunco", "overlap"]):
        return "Medium"
    return "Low"


def checkpoint_from_action(issue: str) -> str:
    """Return the Step 3-scope checkpoint for a watchout."""
    lowered = normalize_text(issue).lower()
    if any(term in lowered for term in ["data", "evidence", "trend"]):
        return "Data Refresh Needed"
    if any(term in lowered for term in ["quote", "cost", "margin", "vendor"]):
        return "Before Vendor RFQ"
    if any(term in lowered for term in ["spec", "certification", "prd"]):
        return "Before PRD Build"
    return "Before Leadership Review"


def risk_watchout_rows(packet: dict[str, Any], analysis: dict[str, Any], action_summary: dict[str, Any]) -> list[list[Any]]:
    """Summarize the highest-value PM watchouts."""
    rows: list[list[Any]] = []
    for action in as_list(action_summary.get("actions"))[:4]:
        action_row = as_list(action)
        issue = action_row[1] if len(action_row) > 1 else ""
        why = action_row[2] if len(action_row) > 2 else ""
        recommendation = action_row[3] if len(action_row) > 3 else ""
        rows.append([issue, risk_level_from_action(issue, analysis), why, checkpoint_from_action(issue), recommendation, ""])
    if not rows:
        rows.append(["No major watchout generated", "Low", "The concept does not currently show a major blocker in the available analysis.", "Before Leadership Review", "Protect the assumptions through PRD and RFQ work.", ""])
    return rows


def pm_action_checklist_rows(packet: dict[str, Any], analysis: dict[str, Any], action_summary: dict[str, Any]) -> list[list[Any]]:
    """Return PM-facing next actions without internal P1/P2/P3 codes."""
    product_action = recommended_product_action(packet, analysis)
    rows: list[list[Any]] = []
    for action in as_list(action_summary.get("actions")):
        action_row = as_list(action)
        rows.append(
            [
                product_action,
                action_row[1] if len(action_row) > 1 else "",
                action_row[2] if len(action_row) > 2 else "",
                action_row[3] if len(action_row) > 3 else "",
                action_row[4] if len(action_row) > 4 else "",
            ]
        )
    if not rows:
        rows.append([product_action, "No immediate issue surfaced", "Current evidence supports a normal PM review.", "Review the summary and source links before presenting.", "Leadership readiness"])
    return rows


def write_leadership_brief(
    ws,
    row: int,
    row_number: int,
    packet: dict[str, Any],
    analysis: dict[str, Any],
    normalized: dict[str, Any],
    action_summary: dict[str, Any],
) -> int:
    """Write a compact leadership brief designed for screenshotting."""
    row = section_header(ws, row, "Leadership Brief", end_column=10)
    row = key_value_rows(
        ws,
        row,
        leadership_brief_rows(row_number, packet, analysis, normalized, action_summary),
        columns=2,
    )
    return row + 1


def merged_text_row(ws, row: int, label: str, value: Any) -> int:
    """Write one labeled merged text row."""
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=row, column=2, value=normalize_text(value))
    ws.cell(row=row, column=2).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    return row + 1


def write_list_section(ws, row: int, title: str, values: list[str]) -> int:
    """Write a simple bulleted list section."""
    row = section_header(ws, row, title)
    if not values:
        ws.cell(row=row, column=1, value="No applicable items for this section.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        return row + 2
    for value in values:
        ws.cell(row=row, column=1, value=f"- {normalize_text(value)}")
        ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
    return row + 1


def write_table(ws, row: int, title: str, headers: list[str], rows: list[list[Any]]) -> int:
    """Write a basic table and return the next row index."""
    row = section_header(ws, row, title, end_column=max(10, len(headers)))
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_index, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = WRAP_ALIGNMENT
    row += 1

    if not rows:
        ws.cell(row=row, column=1, value="No applicable rows for this section.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        return row + 2

    for row_offset, values in enumerate(rows):
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_index)
            header = headers[col_index - 1] if col_index <= len(headers) else ""
            if len(headers) >= 2 and headers[0] == "Metric" and headers[1] == "Value" and col_index == 2 and values:
                header = normalize_text(values[0])
            elif len(headers) >= 2 and headers[0] == "Field" and headers[1] == "Recommendation" and col_index == 2 and values:
                header = normalize_text(values[0])
            elif header:
                header = f"{title} {header}"
            apply_table_cell(cell, value, header)
            cell.alignment = WRAP_ALIGNMENT
            if row_offset % 2 == 1:
                cell.fill = ALTERNATE_ROW_FILL
        row += 1
    return row + 1


def apply_table_cell(cell, value: Any, label: str = "") -> None:
    """Write a table cell, optionally attaching a hyperlink."""
    hyperlink = None
    text_value = value
    if isinstance(value, dict):
        text_value = value.get("value")
        hyperlink = value.get("hyperlink")

    cell.value = display_value_for_label(label, text_value)
    if hyperlink:
        cell.hyperlink = hyperlink
        cell.font = HYPERLINK_FONT


def optional_text(value: Any) -> str | None:
    """Return a stripped string or None."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def looks_like_amazon_asin(value: str | None) -> bool:
    """Return whether a value looks like an Amazon ASIN."""
    return bool(value and re.fullmatch(r"[A-Z0-9]{10}", value.upper()))


def source_channel_label(item: dict[str, Any]) -> str:
    """Render a human-friendly source channel label."""
    channel = resolved_source_channel(item)
    return CHANNEL_DISPLAY_NAMES.get(channel, channel.replace("_", " ").title() or "Unknown")


def source_detail_channel_label(item: dict[str, Any]) -> str:
    """Render the channel label used in source-detail tables."""
    if optional_text(item.get("collection_method")) == "redshift_ecommerce_snapshot":
        return short_domain(item.get("source_domain") or item.get("url")) or "Verified direct competitor"
    return source_channel_label(item)


def discovery_source_label(item: dict[str, Any]) -> str:
    """Render the original discovery source for inferred competitors."""
    domain = optional_text(item.get("discovery_source_domain"))
    if domain:
        host = urlparse(domain if "://" in domain else f"//{domain}").netloc or domain
        return host.replace("www.", "")
    channel = optional_text(item.get("discovery_source_channel"))
    if channel:
        lowered = channel.lower()
        return CHANNEL_DISPLAY_NAMES.get(lowered, lowered.replace("_", " ").title())
    domain = optional_text(item.get("source_domain"))
    if domain:
        host = urlparse(domain if "://" in domain else f"//{domain}").netloc or domain
        return host.replace("www.", "")
    return source_channel_label(item)


def listing_identifier_from_url(url: str | None, source_channel: str | None) -> str | None:
    """Best-effort identifier fallback derived from the listing URL."""
    if not url:
        return None

    channel = (source_channel or "").lower()
    patterns = []
    if channel == "amazon":
        patterns = [(r"/dp/([A-Z0-9]{10})(?:[/?]|$)", "ASIN {match}")]
    elif channel == "home_depot":
        patterns = [(r"/p/(?:[^/]+/)?(\d+)(?:[/?]|$)", "Item {match}")]
    elif channel == "walmart":
        patterns = [(r"/ip/(?:[^/]+/)?(\d+)(?:[/?]|$)", "Item {match}")]
    elif channel == "lowes":
        patterns = [(r"/pd/[^/]+/(\d+)(?:[/?]|$)", "Item {match}")]

    for pattern, template in patterns:
        match = re.search(pattern, url, flags=re.IGNORECASE)
        if match:
            return template.format(match=match.group(1))
    return None


def listing_identifier(item: dict[str, Any]) -> str:
    """Return the best channel-aware identifier label for one competitor listing."""
    source_channel = resolved_source_channel(item)
    sku = optional_text(item.get("sku"))
    model_number = optional_text(item.get("model_number"))
    url = optional_text(item.get("url"))

    if source_channel == "amazon":
        asin = sku if looks_like_amazon_asin(sku) else None
        if not asin and looks_like_amazon_asin(model_number):
            asin = model_number
        if not asin and url:
            derived = listing_identifier_from_url(url, source_channel)
            if derived and derived.startswith("ASIN "):
                asin = derived.replace("ASIN ", "", 1)
        if asin:
            return f"ASIN {asin}"
        if model_number:
            return f"Model {model_number}"

    if model_number:
        return f"Model {model_number}"

    if sku:
        if source_channel == "brand_site":
            return f"Part {sku}"
        if source_channel in {"home_depot", "walmart", "lowes"}:
            return f"Item {sku}"
        return f"SKU {sku}"

    derived = listing_identifier_from_url(url, source_channel)
    return derived or ""


def listing_link_cell(item: dict[str, Any]) -> dict[str, str] | str:
    """Return a hyperlink cell payload for a competitor source listing."""
    url = optional_text(item.get("url"))
    if not url or url.startswith("stackline://") or product_url_issue(url, resolved_source_channel(item)):
        return ""
    return {
        "value": source_detail_channel_label(item),
        "hyperlink": url,
    }


def verification_status(item: dict[str, Any]) -> str:
    """Return the normalized verification status for one competitor record."""
    status = (optional_text(item.get("verification_status")) or "").lower()
    if status == "verified_listing" and product_url_issue(item.get("url"), resolved_source_channel(item)):
        return "inferred_competitor"
    if status in {"verified_listing", "inferred_competitor"}:
        return status
    return "verified_listing"


def verification_label(item: dict[str, Any]) -> str:
    """Render a user-facing verification label."""
    return "Verified Listing" if verification_status(item) == "verified_listing" else "Inferred Competitor"


def sort_candidates(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sort candidate records by confidence, then by price."""
    def sort_key(item: dict[str, Any]) -> tuple[float, float, str]:
        confidence = item.get("match_confidence")
        if not isinstance(confidence, (int, float)):
            confidence = 0
        price = item.get("price")
        if not isinstance(price, (int, float)):
            price = 0
        title = normalize_text(item.get("product_title"))
        return (-confidence, price, title)

    return sorted(items, key=sort_key)


def candidate_rows(
    normalized_items: list[dict[str, Any]],
    channels: set[str],
    limit: int = 10,
    verification_filter: str | None = "verified_listing",
) -> list[list[Any]]:
    """Build a compact competitor table filtered by source channel."""
    rows = []
    for item in sort_candidates(normalized_items):
        if resolved_source_channel(item) not in channels:
            continue
        if verification_filter and verification_status(item) != verification_filter:
            continue
        rows.append(
            [
                item.get("brand"),
                item.get("product_title"),
                listing_identifier(item),
                source_channel_label(item),
                item.get("price"),
                item.get("wattage"),
                item.get("lumens"),
                item.get("cct"),
                item.get("cri"),
                listing_link_cell(item),
                item.get("match_confidence"),
            ]
        )
        if len(rows) >= limit:
            break
    return rows


def item_demand_signal_parts(item: dict[str, Any]) -> list[str]:
    """Return demand evidence as volume, rate/share, and recency/window columns."""
    volume = ""
    rate_or_share = ""
    recency_window = ""

    decrease = first_number_from_fields(item, ["observed_stock_decrease", "stock_decrease_units", "inventory_movement"])
    decrease_events = first_number_from_fields(item, ["decrease_events", "stock_decrease_events"])
    if decrease is not None and decrease > 0:
        volume = f"{display_number(decrease)} stock decrease units"
        if decrease_events is not None and decrease_events > 0:
            volume += f"; {display_number(decrease_events)} decrease events"
    else:
        stackline_units = first_number_from_fields(item, ["units_sold"]) or number_from_observation(item, "Stackline units_sold")
        if stackline_units is not None and stackline_units > 0:
            volume = f"{display_number(stackline_units)} Stackline units sold"

    velocity = first_number_from_fields(
        item,
        ["velocity_units_per_week", "avg_units_per_week_decrease_window", "avg_units_per_week_observed_window"],
    )
    if velocity is not None and velocity > 0:
        rate_or_share = f"{display_number(velocity, 1)} units/week"
    else:
        sales_share = first_number_from_fields(item, ["sales_share_pct"]) or number_from_observation(item, "Stackline sales_share_pct")
        if sales_share is not None and sales_share > 0:
            rate_or_share = f"{display_number(sales_share, 1)}% Stackline sales share"
        elif item.get("sales_rank") not in (None, ""):
            rate_or_share = f"Sales rank {normalize_text(item.get('sales_rank'))}"

    recency_parts: list[str] = []
    days_since = first_number_from_fields(item, ["days_since_last_decrease"])
    if days_since is not None:
        recency_parts.append(f"{display_number(days_since)} days since last decrease")
    latest_date = normalize_text(item.get("latest_inventory_scrape_date") or item.get("last_observed_date"))
    if latest_date:
        recency_parts.append(f"latest scrape {latest_date[:10]}")
    observation_window = first_number_from_fields(item, ["observation_window_days"])
    decrease_window = first_number_from_fields(item, ["decrease_window_days"])
    if observation_window is not None and observation_window > 0:
        recency_parts.append(f"{display_number(observation_window)}-day observed window")
    if decrease_window is not None and decrease_window > 0:
        recency_parts.append(f"{display_number(decrease_window)}-day decrease window")
    if not recency_parts and (volume or rate_or_share) and any("Stackline" in normalize_text(value) for value in as_list(item.get("raw_observations"))):
        recency_parts.append("Stackline segment period")
    recency_window = "; ".join(recency_parts)

    if not any([volume, rate_or_share, recency_window]):
        observations = [normalize_text(value) for value in as_list(item.get("raw_observations")) if normalize_text(value)]
        if observations:
            legacy = demand_signal_parts_from_text("; ".join(observations[:3]))
            return legacy

    return [volume, rate_or_share, recency_window]


def item_demand_signal(item: dict[str, Any]) -> str:
    """Return compact demand evidence attached to one competitor item."""
    parts = [part for part in item_demand_signal_parts(item) if normalize_text(part)]
    if parts:
        return "; ".join(parts)
    signal_keys = [
        "inventory_movement",
        "stock_decrease_units",
        "stock_decrease_events",
        "velocity_units_per_week",
        "sales_rank",
        "units_sold",
        "sales_share_pct",
        "review_count",
        "rating",
    ]
    pieces = []
    for key in signal_keys:
        value = item.get(key)
        if value is not None and value != "":
            label = key.replace("_", " ").title()
            pieces.append(f"{label}: {normalize_text(value)}")
    return "; ".join(pieces[:3])


def item_attribute_summary(item: dict[str, Any]) -> str:
    """Return compact attribute evidence for a competitor item."""
    parts: list[str] = []
    for label, key in [
        ("Voltage", "voltage"),
        ("Pack", "pack_quantity"),
        ("Certs", "certifications"),
        ("Features", "features"),
        ("Dimming", "dimming"),
        ("Mount", "mounting_type"),
        ("Finish", "finish_color"),
    ]:
        value = item.get(key)
        if value is not None and value != "" and value != []:
            parts.append(f"{label}: {normalize_text(value)}")
    return "; ".join(parts[:4])


def detailed_candidate_rows(
    normalized_items: list[dict[str, Any]],
    channels: set[str],
    limit: int = 10,
    verification_filter: str | None = "verified_listing",
) -> list[list[Any]]:
    """Build a competitor table that includes movement and attribute evidence."""
    rows = []
    for item in sort_candidates(normalized_items):
        if resolved_source_channel(item) not in channels:
            continue
        if verification_filter and verification_status(item) != verification_filter:
            continue
        rows.append(
            source_detail_row(item)
        )
        if len(rows) >= limit:
            break
    return rows


def source_detail_row(item: dict[str, Any], source_basis: str | None = None) -> list[Any]:
    """Build one Section B/C source-detail row."""
    return [
        item.get("brand"),
        item.get("product_title"),
        listing_identifier(item),
        source_detail_channel_label(item),
        item.get("price"),
        product_key_specs(item) or item_attribute_summary(item),
        *item_demand_signal_parts(item),
        match_quality_for_item(item),
        listing_link_cell(item),
        link_status_for_item(item),
        source_basis or discovery_source_label(item),
    ]


def ecommerce_cache_item_from_row(row: dict[str, Any]) -> dict[str, Any]:
    """Convert one Redshift ecommerce snapshot row to the report item shape."""
    domain = short_domain(row.get("domain") or row.get("url"))
    approved_source = approved_competitor_source_for_domain(domain) or {}
    features: list[str] = []
    if normalize_text(row.get("dimmable")) in {"1", "Yes", "true", "True"}:
        features.append("dimmable")
    if normalize_text(row.get("ip_rating")):
        features.append(normalize_text(row.get("ip_rating")))
    return {
        "source_channel": "brand_site",
        "source_domain": domain,
        "discovery_source_channel": "redshift_ecommerce_snapshot",
        "discovery_source_domain": domain,
        "collection_method": "redshift_ecommerce_snapshot",
        "approved_competitor": approved_source.get("competitor"),
        "approved_source_tier": approved_source.get("tier"),
        "brand": row.get("brand") or approved_source.get("competitor"),
        "product_title": row.get("name"),
        "model_number": row.get("sku"),
        "sku": row.get("sku"),
        "url": row.get("url"),
        "price": decimal_from_value(row.get("price")) or row.get("price"),
        "currency": row.get("currency") or "USD",
        "wattage": row.get("wattage"),
        "lumens": row.get("lumens"),
        "cct": row.get("cct"),
        "cri": row.get("cri"),
        "voltage": row.get("voltage"),
        "certifications": [value for value in [row.get("ip_rating")] if normalize_text(value)],
        "features": features,
        "image_url": row.get("image"),
        "availability": row.get("availability") or row.get("stock_status"),
        "observed_stock_decrease": row.get("observed_stock_decrease"),
        "decrease_events": row.get("decrease_events"),
        "observation_count": row.get("observation_count"),
        "observation_window_days": row.get("observation_window_days"),
        "decrease_window_days": row.get("decrease_window_days"),
        "avg_units_per_week_observed_window": row.get("avg_units_per_week_observed_window"),
        "avg_units_per_week_decrease_window": row.get("avg_units_per_week_decrease_window"),
        "days_since_last_decrease": row.get("days_since_last_decrease"),
        "latest_inventory_scrape_date": row.get("latest_inventory_scrape_date") or row.get("scraped_at"),
        "verification_status": "verified_listing",
        "verification_reason": "Verified direct competitor row from Redshift ecommerce scraper snapshot.",
        "match_confidence": 0.9,
    }


def ecommerce_cache_sort_key(row: dict[str, Any]) -> tuple[Decimal, Decimal, Decimal, str]:
    """Sort Redshift ecommerce snapshot rows by observed movement and freshness."""
    velocity = first_number_from_fields(
        row,
        ["avg_units_per_week_observed_window", "avg_units_per_week_decrease_window"],
    ) or Decimal("0")
    decrease = decimal_from_value(row.get("observed_stock_decrease")) or Decimal("0")
    events = decimal_from_value(row.get("decrease_events")) or Decimal("0")
    scraped_at = normalize_text(row.get("latest_inventory_scrape_date") or row.get("scraped_at"))
    return (velocity, decrease, events, scraped_at)


def verified_scraped_competitor_items(packet: dict[str, Any], limit: int = 6) -> list[dict[str, Any]]:
    """Return verified direct-competitor items from the Redshift ecommerce cache."""
    payload = latest_ecommerce_snapshot_payload(category_slug_for_packet(packet))
    rows = as_list(payload.get("rows"))
    if not rows:
        return []
    require_approved_source = bool(load_approved_competitor_sources())
    output: list[dict[str, Any]] = []
    for raw_row in sorted(rows, key=ecommerce_cache_sort_key, reverse=True):
        row = as_dict(raw_row)
        domain = short_domain(row.get("domain") or row.get("url"))
        if not normalize_text(row.get("url")) or "sunco" in short_domain(row.get("url")):
            continue
        if require_approved_source and not approved_competitor_source_for_domain(domain):
            continue
        output.append(ecommerce_cache_item_from_row(row))
        if len(output) >= limit:
            break
    return output


def verified_scraped_competitor_rows(packet: dict[str, Any], limit: int = 6) -> list[list[Any]]:
    """Return verified direct-competitor rows from the Redshift ecommerce cache."""
    return [
        source_detail_row(item, "Redshift verified ecommerce scraper snapshot")
        for item in verified_scraped_competitor_items(packet, limit=limit)
    ]


def ecommerce_snapshot_domains(packet: dict[str, Any]) -> set[str]:
    """Return domains represented in the active category ecommerce snapshot."""
    payload = latest_ecommerce_snapshot_payload(category_slug_for_packet(packet))
    domains: set[str] = set()
    for raw_row in as_list(payload.get("rows")):
        row = as_dict(raw_row)
        domain = short_domain(row.get("domain") or row.get("url"))
        if domain:
            domains.add(domain)
    return domains


def approved_source_snapshot_status(source: dict[str, Any], snapshot_domains: set[str]) -> str:
    """Explain whether an approved source has current category product evidence."""
    domain = source.get("domain")
    if any(domain_matches(snapshot_domain, domain) for snapshot_domain in snapshot_domains):
        return "Verified PDP rows in current category snapshot"
    if decimal_from_value(source.get("latest_rows")) and decimal_from_value(source.get("latest_rows")) > 0:
        return "Source exists in Redshift, but no matching PDP rows in current category snapshot"
    return "Not currently covered by Redshift scrape views"


def approved_source_warehouse_coverage(source: dict[str, Any]) -> str:
    """Return compact warehouse coverage counts for the approved source."""
    latest_rows = decimal_from_value(source.get("latest_rows")) or Decimal("0")
    latest_urls = decimal_from_value(source.get("latest_urls")) or Decimal("0")
    inventory_urls = decimal_from_value(source.get("inventory_urls")) or Decimal("0")
    parts = [f"{display_number(latest_urls)} latest PDP URLs"] if latest_urls > 0 else ["0 latest PDP URLs"]
    if inventory_urls > 0:
        parts.append(f"{display_number(inventory_urls)} inventory URLs")
    latest_scrape = normalize_text(source.get("latest_last_scraped"))
    if latest_scrape:
        parts.append(f"latest scrape {latest_scrape[:10]}")
    if latest_rows <= 0 and normalize_text(source.get("tier")):
        parts.append("approved source list only")
    return "; ".join(parts)


def approved_source_next_step(status: str) -> str:
    """Return the safe next action for one approved source."""
    lowered = status.lower()
    if "current category snapshot" in lowered and "verified pdp" in lowered:
        return "Use verified PDP rows above when product/spec match is visible."
    if "source exists in redshift" in lowered:
        return "Refresh/query ecommerce snapshot with broader category terms or site-specific terms."
    return "Add source to scraper coverage or run manual site discovery before citing products."


def approved_source_coverage_rows(packet: dict[str, Any], limit: int = 12) -> list[list[Any]]:
    """Show approved competitor-source coverage for the active category."""
    category_slug = category_slug_for_packet(packet)
    identity = as_dict(packet.get("identity"))
    snapshot_domains = ecommerce_snapshot_domains(packet)
    sources = approved_sources_for_category(category_slug, limit=limit)
    if not sources:
        return [
            [
                "No approved source registry loaded",
                "",
                "",
                "",
                "Run approved_competitor_source_audit.py to create the backend source cache.",
                "",
                "",
                "Do not treat missing Section C rows as proof that no competitors exist.",
            ]
        ]

    rows: list[list[Any]] = []
    for source in sources:
        status = approved_source_snapshot_status(source, snapshot_domains)
        search_link = source_link_cell(source_search_link(source, category_slug, identity.get("ideation_name")), "Search source")
        rows.append(
            [
                source.get("competitor"),
                source.get("domain"),
                source.get("tier"),
                source.get("product_subcategories_in_scope") or source.get("focus"),
                status,
                approved_source_warehouse_coverage(source),
                search_link,
                approved_source_next_step(status),
            ]
        )
    return rows


def inferred_competitor_rows(
    normalized_items: list[dict[str, Any]],
    limit: int = 12,
) -> list[list[Any]]:
    """Build a table of inferred competitors that need listing verification."""
    rows = []
    for item in sort_candidates(normalized_items):
        if verification_status(item) != "inferred_competitor":
            continue
        why_included = (
            product_url_issue(item.get("url"), resolved_source_channel(item))
            or item.get("verification_reason")
            or item.get("extraction_notes")
            or item.get("match_notes")
        )
        rows.append(
            [
                item.get("brand"),
                item.get("product_title"),
                directional_channel_label(item),
                directional_supporting_source_label(item),
                listing_link_cell(item),
                link_status_for_item(item),
                why_included,
                item.get("match_confidence"),
            ]
        )
        if len(rows) >= limit:
            break
    return rows


def top_brand_rows(brands: list[dict[str, Any]]) -> list[list[Any]]:
    """Format summarized brand rows."""
    rows = []
    for brand in brands:
        rows.append(
            [
                brand.get("brand"),
                brand.get("candidate_count"),
                normalize_text(brand.get("source_channels")),
                brand.get("median_unit_price"),
            ]
        )
    return rows


def coverage_rows(entries: list[dict[str, Any]]) -> list[list[Any]]:
    """Format feature/certification coverage rows."""
    rows = []
    for entry in entries:
        rows.append([entry.get("label"), entry.get("matched_count"), entry.get("coverage_pct")])
    return rows


def benchmark_rows(pricing: dict[str, Any]) -> list[list[Any]]:
    """Format multi-metric pricing benchmark rows."""
    metrics = [
        ("Raw Price", as_dict(pricing.get("price_benchmarks"))),
        ("Unit Price", as_dict(pricing.get("unit_price_benchmarks"))),
        ("Unit Price / Watt", as_dict(pricing.get("unit_price_per_watt_benchmarks"))),
        ("Unit Price / Lumen", as_dict(pricing.get("unit_price_per_lumen_benchmarks"))),
    ]
    rows = []
    for label, metric in metrics:
        rows.append(
            [
                label,
                metric.get("sample_size"),
                metric.get("min"),
                metric.get("p25"),
                metric.get("median"),
                metric.get("mean"),
                metric.get("p75"),
                metric.get("max"),
            ]
        )
    return rows


def pricing_position_rows(pricing: dict[str, Any]) -> list[list[Any]]:
    """Format pricing positioning rows."""
    suggested = as_dict(pricing.get("suggested_msrp_range"))
    target_position = as_dict(pricing.get("target_price_position"))
    return [
        ["Target MSRP", pricing.get("target_msrp")],
        ["Evaluated Price", target_position.get("evaluated_value")],
        ["Evaluated Price Source", target_position.get("evaluated_value_source")],
        ["Target Price Percentile", target_position.get("percentile")],
        ["Target Price Bucket", target_position.get("bucket")],
        ["Target vs Market Median %", target_position.get("vs_median_pct")],
        ["Observed Unit Price Floor (P25)", suggested.get("observed_unit_price_floor")],
        ["Observed Unit Price Ceiling (P75)", suggested.get("observed_unit_price_ceiling")],
        ["Recommended Floor", suggested.get("recommended_floor")],
        ["Recommended Ceiling", suggested.get("recommended_ceiling")],
        ["Minimum Margin-Safe MSRP", suggested.get("minimum_margin_safe_price")],
        ["Suggested Positioning", suggested.get("positioning")],
        ["Margin Conflict", suggested.get("margin_conflict")],
    ]


def margin_rows(pricing: dict[str, Any]) -> list[list[Any]]:
    """Format channel-specific margin guidance rows."""
    rows = []
    for channel in ["shopify", "amazon"]:
        entry = as_dict(as_dict(pricing.get("margin_targets")).get(channel))
        if not entry:
            continue
        rows.append(
            [
                channel.title(),
                entry.get("target_margin_pct"),
                entry.get("minimum_viable_msrp"),
                entry.get("vs_target_msrp_pct"),
                entry.get("vs_market_median_pct"),
            ]
        )
    return rows


def value_position_rows(pricing: dict[str, Any]) -> list[list[Any]]:
    """Format value-ranking rows for unit price, price per watt, and price per lumen."""
    rows = []
    metrics = [
        ("Unit Price", as_dict(pricing.get("target_price_position"))),
        ("Unit Price / Watt", as_dict(pricing.get("target_price_per_watt_position"))),
        ("Unit Price / Lumen", as_dict(pricing.get("target_price_per_lumen_position"))),
    ]
    for label, metric in metrics:
        if not metric:
            continue
        rows.append(
            [
                label,
                metric.get("evaluated_value"),
                metric.get("percentile"),
                metric.get("bucket"),
                metric.get("vs_median_pct"),
            ]
        )
    return rows


def channel_comparison_rows(performance: dict[str, Any]) -> list[list[Any]]:
    """Format Stackline channel comparison rows for Section A."""
    comparison = as_dict(performance.get("channel_comparison"))
    channels = as_dict(comparison.get("channels"))
    rows = []
    for channel_name, channel in channels.items():
        channel = as_dict(channel)
        rows.append(
            [
                channel_name,
                channel.get("retail_sales"),
                channel.get("units_sold"),
                channel.get("avg_retail_price"),
                channel.get("retail_sales_growth_pct"),
                channel.get("sunco_sales_share_pct"),
            ]
        )
    return rows


def spec_action_rows(spec_coverage: dict[str, Any]) -> list[list[Any]]:
    """Format feature/certification coverage in PM-facing language."""
    def label_marker(value: Any) -> str:
        return re.sub(r"[^a-z0-9]+", " ", normalize_text(value).lower()).strip()

    def signal_read(entry: dict[str, Any]) -> str:
        signal = normalize_text(entry.get("signal")).lower()
        if signal == "whitespace":
            return "Competitor-supported gap"
        if signal == "table_stakes":
            return "Common market requirement"
        if signal in {"competitive", "leading"}:
            return "Potential differentiator"
        if signal in {"low_signal", "below_market"}:
            return "Low-confidence signal"
        return normalize_text(entry.get("signal")) or "Market signal"

    def sunco_coverage_read(entry: dict[str, Any]) -> str:
        coverage = entry.get("coverage_pct")
        matched = entry.get("matched_count")
        if coverage is None and matched is None:
            return "Coverage not quantified"
        return f"{display_value_for_label('Coverage %', coverage)} coverage; {display_value_for_label('Count', matched)} matched"

    rows = []
    covered_labels = set()
    for entry in as_list(spec_coverage.get("attribute_decisions")):
        entry = as_dict(entry)
        for label in as_list(entry.get("covered_labels")):
            marker = label_marker(label)
            if marker:
                covered_labels.add(marker)
        rows.append(
            [
                entry.get("attribute"),
                entry.get("type"),
                entry.get("market_evidence"),
                entry.get("sunco_coverage"),
                entry.get("recommendation"),
                entry.get("source_notes"),
            ]
        )

    for entry in as_list(spec_coverage.get("feature_coverage")):
        if label_marker(entry.get("label")) in covered_labels:
            continue
        rows.append(
            [
                entry.get("label"),
                "Feature",
                signal_read(entry),
                sunco_coverage_read(entry),
                entry.get("recommended_action"),
                entry.get("evidence_strength"),
            ]
        )
    for entry in as_list(spec_coverage.get("certification_coverage")):
        rows.append(
            [
                entry.get("label"),
                "Certification",
                signal_read(entry),
                sunco_coverage_read(entry),
                entry.get("recommended_action"),
                entry.get("evidence_strength"),
            ]
        )
    return rows


def numeric_guidance_rows(spec_coverage: dict[str, Any]) -> list[list[Any]]:
    """Format numeric target-positioning rows."""
    rows = []
    for entry in as_list(spec_coverage.get("numeric_guidance")):
        rows.append(
            [
                entry.get("label"),
                entry.get("target_value"),
                entry.get("median"),
                entry.get("p75"),
                entry.get("target_percentile"),
                entry.get("recommended_action"),
            ]
        )
    return rows


def gate_readiness_snapshot_rows(gate_readiness: dict[str, Any]) -> list[list[Any]]:
    """Format gate/channel readiness rows for the report."""
    rows = []
    for snapshot in as_list(gate_readiness.get("snapshots")):
        evidence = as_dict(snapshot.get("evidence_confidence"))
        rows.append(
            [
                snapshot.get("channel"),
                snapshot.get("gate"),
                snapshot.get("family_state"),
                snapshot.get("weighted_score"),
                evidence.get("score"),
                evidence.get("label"),
                f"{evidence.get('implemented_questions')}/{evidence.get('methodology_active_questions')}",
            ]
        )
    return rows


def gate_readiness_pillar_rows(gate_readiness: dict[str, Any]) -> list[list[Any]]:
    """Format pillar-level rollups from the primary G2 channel snapshot."""
    primary_channel = normalize_text(gate_readiness.get("primary_channel"))
    for snapshot in as_list(gate_readiness.get("snapshots")):
        if snapshot.get("channel") == primary_channel and snapshot.get("gate") == "G2":
            rows = []
            for pillar in as_list(snapshot.get("pillar_scores")):
                rows.append(
                    [
                        pillar.get("label"),
                        pillar.get("base_weight"),
                        pillar.get("effective_weight"),
                        pillar.get("average_score"),
                        f"{pillar.get('scored_question_count')}/{pillar.get('question_count')}",
                        pillar.get("status"),
                    ]
                )
            return rows
    return []


def vendor_request_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    """Format vendor optimization requests."""
    rows = []
    for item in items:
        rows.append(
            [
                item.get("priority"),
                item.get("linked_metric"),
                item.get("request"),
                item.get("reason"),
            ]
        )
    return rows


def optimization_driver_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    """Format category-aware decision drivers for the report."""
    rows = []
    for item in items:
        rows.append(
            [
                item.get("tier"),
                item.get("label"),
                item.get("driver_type"),
                item.get("signal"),
                item.get("reason"),
            ]
        )
    return rows


def low_signal_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    """Format lower-signal attributes that should be validated before over-weighting."""
    rows = []
    for item in items:
        rows.append(
            [
                item.get("label"),
                item.get("driver_type"),
                item.get("signal"),
                item.get("reason"),
            ]
        )
    return rows


def optimization_modifier_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    """Format active optimization modifiers for the report."""
    rows = []
    for item in items:
        rows.append(
            [
                item.get("label"),
                ", ".join(as_list(item.get("matched_keywords"))),
                item.get("match_score"),
            ]
        )
    return rows


def optimization_scorecard_rows(scorecard: dict[str, Any]) -> list[list[Any]]:
    """Format optimization score components for the report."""
    rows = []
    for item in as_list(scorecard.get("components")):
        rows.append(
            [
                item.get("component"),
                item.get("score"),
                item.get("weight"),
                item.get("reason"),
            ]
        )
    return rows


def report_filename(row_number: int) -> str:
    """Return the stable report filename for a row."""
    return f"row_{row_number:03d}_research_report.xlsx"


def prd_prefill_pairs(packet: dict[str, Any], analysis: dict[str, Any]) -> list[tuple[str, Any]]:
    """Build the PRD-oriented prefill block from packet targets."""
    identity = as_dict(packet.get("identity"))
    target_profile = as_dict(packet.get("target_profile"))
    electrical = as_dict(target_profile.get("electrical"))
    physical = as_dict(target_profile.get("physical"))
    business = as_dict(target_profile.get("business_case"))
    reference = as_dict(packet.get("reference_baseline"))

    return [
        ("Ideation Name", normalize_text(identity.get("ideation_name"))),
        ("Reference Image URL", reference.get("image_url")),
        ("Voltage", electrical.get("voltage")),
        ("Wattage Primary", electrical.get("wattage_primary")),
        ("Wattage Max", electrical.get("wattage_max")),
        ("Selectable Wattage", electrical.get("selectable_wattage")),
        ("CCT Primary", electrical.get("cct_primary")),
        ("CCT Max", electrical.get("cct_max")),
        ("Selectable CCT", electrical.get("selectable_cct")),
        ("CRI", electrical.get("cri")),
        ("Lumens Target", electrical.get("lumens_target")),
        ("Dimmable", electrical.get("dimmable")),
        ("Dimming Type", electrical.get("dimming_type")),
        ("Size / Form Factor", physical.get("size_form_factor")),
        ("Mounting Type", physical.get("mounting_type")),
        ("Material", physical.get("material")),
        ("Finish / Color", physical.get("finish_color")),
        ("IP Rating", physical.get("ip_rating")),
        ("Moisture Rating", physical.get("moisture_rating")),
        ("Target MSRP", as_dict(analysis.get("pricing_analysis")).get("target_msrp")),
        ("Target Vendor Cost", as_dict(analysis.get("pricing_analysis")).get("target_vendor_cost")),
        ("Certifications", business.get("certifications")),
        ("Lifetime Hours", business.get("lifetime_hours")),
        ("Warranty", business.get("warranty")),
    ]


def render_row_sheet(
    ws,
    row_number: int,
    packet: dict[str, Any],
    analysis: dict[str, Any],
    normalized: dict[str, Any],
) -> None:
    """Render one ideation sheet into an existing worksheet."""
    set_default_layout(ws)

    identity = as_dict(packet.get("identity"))
    reference = as_dict(packet.get("reference_baseline"))
    performance = as_dict(analysis.get("performance_estimation"))
    pricing = as_dict(analysis.get("pricing_analysis"))
    summary = as_dict(analysis.get("summary"))
    spec_coverage = as_dict(analysis.get("spec_coverage"))
    reference_anchor = as_dict(analysis.get("reference_anchor_context"))
    gate_readiness = as_dict(analysis.get("gate_readiness"))
    vendor_requests = as_list(analysis.get("highest_impact_vendor_requests"))
    ideation_optimization = as_dict(analysis.get("ideation_optimization"))
    optimization_scorecard = as_dict(ideation_optimization.get("optimization_scorecard"))
    optimization_modifiers = as_list(ideation_optimization.get("active_modifiers"))
    normalized_summary = as_dict(normalized.get("summary"))
    action_summary = blocker_action_summary(packet, analysis)
    tracking_name = concept_tracking_name(packet)

    ws.cell(row=1, column=1, value=f"{tracking_name} | {normalize_text(identity.get('ideation_name'))}")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = WRAP_ALIGNMENT
    ws.row_dimensions[1].height = 48
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=2, column=1, value=f"Row {row_number} Research Report | Concept Tracking Name generated from SKU Decoder cache")
    ws.cell(row=2, column=1).fill = ACCENT_FILL
    ws.cell(row=2, column=1).alignment = WRAP_ALIGNMENT
    ws.row_dimensions[2].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    row = 4
    row = write_pm_decision_snapshot(ws, row, row_number, packet, analysis, normalized, action_summary)
    row = write_table(
        ws,
        row,
        "PM Action Checklist",
        ["Recommended Product Action", "Issue / Reason", "Why It Matters", "Next Action", "Expected Impact"],
        pm_action_checklist_rows(packet, analysis, action_summary),
    )
    row = write_table(
        ws,
        row,
        "Demand Proof",
        DEMAND_PROOF_HEADERS,
        demand_proof_rows(packet, analysis, normalized),
    )
    row = write_table(
        ws,
        row,
        "Sunco Coverage",
        SUNCO_COVERAGE_HEADERS,
        sunco_coverage_decision_rows(packet, analysis, normalized),
    )
    row = write_table(
        ws,
        row,
        "Primary Competitor Comparables",
        COMPETITOR_COMPARABLE_HEADERS,
        competitor_decision_rows(packet, analysis, normalized),
    )
    row = write_table(
        ws,
        row,
        "Directional / Secondary Comparables",
        COMPETITOR_COMPARABLE_HEADERS,
        directional_competitor_rows(packet, analysis, normalized),
    )
    row = write_table(
        ws,
        row,
        "Pricing / Channel Read",
        PRICING_CHANNEL_HEADERS,
        pricing_channel_decision_rows(packet, analysis, normalized),
    )
    row = write_table(
        ws,
        row,
        "Go-Forward Watchouts",
        WATCHOUT_HEADERS,
        risk_watchout_rows(packet, analysis, action_summary),
    )
    row = section_header(ws, row, "Section A - Ideation + Current Sunco Context")
    section_a_rows = [
        ("Concept Tracking Name", tracking_name),
        ("Recommended Product Action", recommended_product_action(packet, analysis)),
    ]
    if recommended_product_action(packet, analysis) == PRODUCT_ACTION_REVISION:
        section_a_rows.extend([
            ("Revision Target SKU", revision_target_sku(packet, analysis)),
            ("Recommended Revision Changes", revision_change_summary(packet, analysis)),
        ])
    section_a_rows.extend([
        ("Gap Reason", product_fit_classification(packet)),
        ("Recommended Channel", channel_strategy_label(packet, analysis)),
        ("Category Owner", identity.get("category_owner")),
        ("Category", identity.get("category")),
        ("Subcategory", identity.get("subcategory")),
        ("Reference Anchor SKU", identity.get("sunco_reference_sku")),
        ("Reference Anchor Product", reference.get("title")),
        ("Anchor Listing Price", reference.get("listing_price")),
        ("Anchor Shopify Revenue 12mo", reference.get("shopify_revenue_12mo")),
        ("Anchor Shopify Units 12mo", reference.get("shopify_units_12mo")),
        ("Anchor Amazon Revenue 12mo", reference.get("amazon_revenue_12mo")),
        ("Anchor Amazon Units 12mo", reference.get("amazon_units_12mo")),
        ("Anchor Role", reference_anchor.get("primary_use")),
        ("Gap vs Ideation", product_difference_summary(packet, analysis)),
    ])
    row = key_value_rows(ws, row, section_a_rows)
    row = merged_text_row(
        ws,
        row,
        "Performance Rationale",
        " | ".join(as_list(performance.get("rationale"))),
    )
    row = key_value_rows(
        ws,
        row,
        [
            ("Verified Listings", normalized_summary.get("verified_listing_count")),
            ("Inferred Competitors", normalized_summary.get("inferred_competitor_count")),
        ],
        columns=2,
    )
    row = write_table(
        ws,
        row,
        "Step 1 Demand + Coverage Evidence",
        ["Signal", "Evidence", "Source", "Link"],
        step1_demand_evidence_rows(packet, analysis),
    )
    row = write_table(
        ws,
        row,
        "Stackline Channel Comparison",
        ["Channel", "Retail Sales", "Units Sold", "Avg Price", "Sales Growth %", "Sunco Share %"],
        channel_comparison_rows(performance),
    )
    row = merged_text_row(ws, row, "Gate Readiness Summary", gate_readiness.get("summary"))
    row = write_table(
        ws,
        row,
        "Gate Readiness by Channel / Gate",
        ["Channel", "Gate", "Family State", "Score", "Evidence Score", "Evidence Label", "Implemented Questions"],
        gate_readiness_snapshot_rows(gate_readiness),
    )
    row = write_table(
        ws,
        row,
        "Primary G2 Pillar Rollup",
        ["Pillar", "Base Weight", "Effective Weight", "Avg Score", "Scored Questions", "Status"],
        gate_readiness_pillar_rows(gate_readiness),
    )

    row = write_table(
        ws,
        row,
        "Section B - Amazon Source Detail",
        SOURCE_DETAIL_HEADERS,
        detailed_candidate_rows(as_list(normalized.get("items")), AMAZON_CHANNELS, limit=6, verification_filter="verified_listing"),
    )

    verified_direct_rows = verified_scraped_competitor_rows(packet, limit=6)
    row = write_table(
        ws,
        row,
        "Section C - Direct / Retail Source Detail",
        SOURCE_DETAIL_HEADERS,
        verified_direct_rows
        or detailed_candidate_rows(as_list(normalized.get("items")), BM_DIRECT_CHANNELS, limit=6, verification_filter="verified_listing"),
    )
    row = write_table(
        ws,
        row,
        "Section C Source Coverage Audit",
        APPROVED_SOURCE_COVERAGE_HEADERS,
        approved_source_coverage_rows(packet, limit=12),
    )

    row = write_table(
        ws,
        row,
        "Section D - Directional Competitors - Not Verified",
        ["Brand", "Product Name", "Discovery Channel", "Supporting Source", "Product Link", "Link Status", "Why Included", "Confidence"],
        inferred_competitor_rows(as_list(normalized.get("items")), limit=6),
    )

    row = write_table(
        ws,
        row,
        "Section E - Pricing Detail / Audit",
        ["Metric", "Value"],
        pricing_position_rows(pricing),
    )
    row = write_table(
        ws,
        row,
        "Pricing Benchmarks",
        ["Metric", "Samples", "Min", "P25", "Median", "Mean", "P75", "Max"],
        benchmark_rows(pricing),
    )
    row = write_table(
        ws,
        row,
        "Margin Targets",
        ["Channel", "Target Margin %", "Min MSRP", "Vs Target MSRP %", "Vs Market Median %"],
        margin_rows(pricing),
    )
    row = write_table(
        ws,
        row,
        "Value Ranking",
        ["Metric", "Target", "Percentile", "Bucket", "Vs Median %"],
        value_position_rows(pricing),
    )
    row = write_table(
        ws,
        row,
        "Top Brands",
        ["Brand", "Candidate Count", "Channels", "Median Unit Price"],
        top_brand_rows(as_list(summary.get("top_brands"))),
    )

    row = write_table(
        ws,
        row,
        "Section F - Feature + Certification Evidence",
        ["Attribute", "Type", "Market Evidence", "Sunco Coverage", "Recommendation", "Source / Notes"],
        spec_action_rows(spec_coverage),
    )
    row = write_table(
        ws,
        row,
        "Numeric Target Positioning",
        ["Metric", "Target", "Median", "P75", "Percentile", "Recommendation"],
        numeric_guidance_rows(spec_coverage),
    )
    row = merged_text_row(ws, row, "Category Optimization Summary", ideation_optimization.get("summary"))
    row = key_value_rows(
        ws,
        row,
        [
            ("Optimization Profile", ideation_optimization.get("profile_label")),
            ("Profile Match Basis", ", ".join(as_list(ideation_optimization.get("matched_taxonomy"))) or ", ".join(as_list(ideation_optimization.get("matched_keywords")))),
            ("Active Modifiers", ", ".join(item.get("label") for item in optimization_modifiers if item.get("label"))),
            ("Optimization Score", optimization_scorecard.get("score")),
            ("Optimization Label", optimization_scorecard.get("label")),
            ("Optimization Confidence", optimization_scorecard.get("confidence")),
        ],
    )
    row = write_table(
        ws,
        row,
        "Active Variant Modifiers",
        ["Modifier", "Matched Keywords", "Match Score"],
        optimization_modifier_rows(optimization_modifiers),
    )
    row = write_table(
        ws,
        row,
        "Optimization Scorecard",
        ["Component", "Score", "Weight", "Reason"],
        optimization_scorecard_rows(optimization_scorecard),
    )
    row = write_table(
        ws,
        row,
        "Primary Decision Drivers",
        ["Tier", "Driver", "Type", "Signal", "Why It Matters"],
        optimization_driver_rows(as_list(ideation_optimization.get("primary_decision_drivers"))),
    )
    row = write_table(
        ws,
        row,
        "Secondary Decision Drivers",
        ["Tier", "Driver", "Type", "Signal", "Why It Matters"],
        optimization_driver_rows(as_list(ideation_optimization.get("secondary_decision_drivers"))),
    )
    row = write_table(
        ws,
        row,
        "Do Not Over-Weight Yet",
        ["Driver", "Type", "Signal", "Reason"],
        low_signal_rows(as_list(ideation_optimization.get("low_signal_attributes"))),
    )
    row = write_table(
        ws,
        row,
        "Highest-Impact Vendor Requests",
        ["Priority", "Area", "Request", "Reason"],
        vendor_request_rows(vendor_requests),
    )
    row = write_list_section(ws, row, "Recommendations", as_list(analysis.get("recommendations")))
    row = write_list_section(ws, row, "Audit Notes", as_list(analysis.get("notes")))

    row = section_header(ws, row, "Section G - PRD / RFQ Draft Inputs")
    row = key_value_rows(ws, row, prd_prefill_pairs(packet, analysis))


def clean_slide_text(value: Any, limit: int | None = None) -> str:
    """Return compact text suitable for one-page slide summary cells."""
    text = normalize_text(value).strip()
    text = re.sub(r"\s+", " ", text)
    if limit and len(text) > limit:
        return text[: max(0, limit - 1)].rstrip(" ,.;") + "..."
    return text


def slide_title(packet: dict[str, Any]) -> str:
    """Create a human-readable slide title from the ideation packet."""
    identity = as_dict(packet.get("identity"))
    target_profile = as_dict(packet.get("target_profile"))
    physical = as_dict(target_profile.get("physical"))
    context = context_text_for_packet(packet)
    category = normalize_text(identity.get("subcategory") or identity.get("category") or "Product")
    lumen_match = re.search(r"\b([0-9][0-9,]*)\s*(?:lm|lumens?)\b", context, flags=re.IGNORECASE)
    lumens = f"{lumen_match.group(1).replace(',', ',')}lm" if lumen_match else ""
    size = normalize_text(physical.get("size_form_factor"))
    size_match = re.search(r"\b(\d+\s*(?:ft|foot|feet)|\d+x\d+|1\s*x\s*4|2\s*x\s*2|2\s*x\s*4)\b", size, flags=re.IGNORECASE)
    size_label = size_match.group(1).replace(" ", "") if size_match else ""
    pieces = [category]
    if lumens:
        pieces.append(lumens)
    if size_label:
        pieces.append(size_label)
    title = ", ".join(pieces)
    if len(title) < 8:
        title = normalize_text(identity.get("ideation_name")) or concept_tracking_name(packet)
    return clean_slide_text(title, 95)


def slide_sheet_title_for_row(row_number: int, packet: dict[str, Any], used: set[str]) -> str:
    """Return a workbook-safe slide summary sheet name."""
    return safe_sheet_title(f"S{row_number:02d} {concept_tracking_name(packet)}", used)


def set_slide_layout(ws) -> None:
    """Apply one-page slide-style layout defaults."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.sheet_view.zoomScale = 85

    widths = {
        "A": 13,
        "B": 13,
        "C": 13,
        "D": 13,
        "E": 12,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row_index in range(1, 39):
        ws.row_dimensions[row_index].height = 18
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 48
    ws.row_dimensions[6].height = 46
    ws.row_dimensions[7].height = 46
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[19].height = 24
    ws.row_dimensions[24].height = 24
    ws.row_dimensions[29].height = 24
    ws.row_dimensions[34].height = 24
    ws.row_dimensions[35].height = 16
    ws.print_area = "A1:M37"


def style_range(
    ws,
    cell_range: str,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    border: Border | None = None,
    alignment: Alignment | None = None,
) -> None:
    """Apply style fields to every cell in a range."""
    for row in ws[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if border is not None:
                cell.border = border
            if alignment is not None:
                cell.alignment = alignment


def merge_write(
    ws,
    cell_range: str,
    value: Any,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    border: Border | None = SLIDE_TABLE_BORDER,
    alignment: Alignment | None = None,
) -> None:
    """Merge a range and write a styled value into the top-left cell."""
    ws.merge_cells(cell_range)
    start = cell_range.split(":", 1)[0]
    cell = ws[start]
    cell.value = normalize_text(value)
    style_range(ws, cell_range, fill=fill, font=font, border=border, alignment=alignment or WRAP_ALIGNMENT)


def section_bar(ws, row: int, start_col: int, end_col: int, title: str, *, fill: PatternFill = SLIDE_TEAL_FILL) -> None:
    """Write a colored slide section bar across columns."""
    start = f"{get_column_letter(start_col)}{row}"
    end = f"{get_column_letter(end_col)}{row}"
    merge_write(
        ws,
        f"{start}:{end}",
        title,
        fill=fill,
        font=SLIDE_SECTION_FONT,
        border=SLIDE_TABLE_BORDER,
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )


def write_slide_table(
    ws,
    row: int,
    start_col: int,
    headers: list[str],
    rows: list[list[Any]],
    *,
    header_fill: PatternFill = SLIDE_TEAL_FILL,
    max_rows: int | None = None,
) -> int:
    """Write a compact slide-ready table."""
    if max_rows is not None:
        rows = rows[:max_rows]
    for offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + offset, value=header)
        cell.fill = header_fill
        cell.font = SLIDE_SECTION_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = SLIDE_TABLE_BORDER
    row += 1
    body_rows = rows or [["No available source-backed row."] + [""] * (len(headers) - 1)]
    for row_values in body_rows:
        for offset in range(len(headers)):
            value = row_values[offset] if offset < len(row_values) else ""
            cell = ws.cell(row=row, column=start_col + offset)
            apply_table_cell(cell, value, headers[offset])
            cell.font = SLIDE_SMALL_FONT if not cell.hyperlink else SLIDE_LINK_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = SLIDE_TABLE_BORDER
            cell.fill = SLIDE_WHITE_FILL
        row += 1
    return row


def write_span_table(
    ws,
    row: int,
    sections: list[tuple[str, int, int]],
    rows: list[list[Any]],
    *,
    header_fill: PatternFill = SLIDE_TEAL_FILL,
    max_rows: int | None = None,
) -> int:
    """Write a compact table where logical columns can span multiple Excel columns."""
    if max_rows is not None:
        rows = rows[:max_rows]
    for header, start_col, end_col in sections:
        section_bar(ws, row, start_col, end_col, header, fill=header_fill)
    row += 1
    body_rows = rows or [["No available source-backed row."] + [""] * (len(sections) - 1)]
    for row_values in body_rows:
        for index, (_, start_col, end_col) in enumerate(sections):
            value = row_values[index] if index < len(row_values) else ""
            display_value = value.get("value") if isinstance(value, dict) else value
            display_text = display_value_for_label(sections[index][0], display_value)
            start = f"{get_column_letter(start_col)}{row}"
            end = f"{get_column_letter(end_col)}{row}"
            merge_write(
                ws,
                f"{start}:{end}",
                display_text,
                fill=SLIDE_WHITE_FILL,
                font=SLIDE_SMALL_FONT,
                border=SLIDE_TABLE_BORDER,
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            )
            if isinstance(value, dict) and value.get("hyperlink"):
                ws[start].hyperlink = value.get("hyperlink")
                ws[start].font = SLIDE_LINK_FONT
        row += 1
    return row


def target_spec_lines(packet: dict[str, Any], analysis: dict[str, Any]) -> list[str]:
    """Build concise target-spec bullets for a slide card."""
    target_profile = as_dict(packet.get("target_profile"))
    electrical = as_dict(target_profile.get("electrical"))
    physical = as_dict(target_profile.get("physical"))
    pricing = as_dict(analysis.get("pricing_analysis"))
    specs = [
        ("Wattage", electrical.get("wattage_primary") or electrical.get("wattage_max")),
        ("Lumens", electrical.get("lumens_target")),
        ("CCTs", electrical.get("cct_primary") or electrical.get("cct_max")),
        ("Voltage", electrical.get("voltage")),
        ("Dimming", electrical.get("dimming_type") or ("Dimmable" if electrical.get("dimmable") else "")),
        ("CRI", electrical.get("cri")),
        ("Moisture", physical.get("moisture_rating") or physical.get("ip_rating")),
        ("Certs", certification_summary(packet, analysis)),
        ("Target MSRP", pricing.get("target_msrp")),
    ]
    lines = []
    for label, value in specs:
        text = display_value_for_label(label, value)
        if text:
            lines.append(f"- {label}: {text}")
    return lines[:9]


def primary_feature_summary(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Return a short key-differentiator statement."""
    target_profile = as_dict(packet.get("target_profile"))
    feature_watchlist = [normalize_text(item) for item in as_list(target_profile.get("feature_watchlist")) if normalize_text(item)]
    spec_coverage = as_dict(analysis.get("spec_coverage"))
    feature_rows = as_list(spec_coverage.get("feature_coverage"))
    supported = [
        normalize_text(as_dict(row).get("label"))
        for row in feature_rows
        if normalize_text(as_dict(row).get("signal")).lower() in {"competitive", "differentiator", "table_stakes"}
    ]
    merged = list(dict.fromkeys(feature_watchlist + supported))
    return clean_slide_text(", ".join(merged[:4]) or product_difference_summary(packet, analysis), 130)


def slide_strategy_text(packet: dict[str, Any], analysis: dict[str, Any]) -> str:
    """Build the slide strategy narrative from Step 3 evidence."""
    identity = as_dict(packet.get("identity"))
    evidence = extract_research_note_evidence(packet)
    performance = as_dict(analysis.get("performance_estimation"))
    snapshot = as_dict(performance.get("market_snapshot"))
    movement = evidence.get("ecommerce_movement") or ""
    movement_short = "; ".join(movement.split(";")[:3]).strip()
    channel = channel_strategy(packet, analysis).replace(
        "Shopify/front-end first; Amazon follow-up if Stackline, pack economics, and margin support.",
        "Shopify first; Amazon follow-up if economics support.",
    )
    parts = [
        f"{concept_tracking_name(packet)} is recommended as {recommended_product_action(packet, analysis)} in {normalize_text(identity.get('subcategory') or identity.get('category'))}.",
        f"Gap reason: {product_fit_classification(packet)}.",
        f"Gap vs current line: {clean_slide_text(product_difference_summary(packet, analysis), 140)}",
    ]
    market = []
    if snapshot.get("segment_retail_sales") is not None:
        market.append(f"{display_value_for_label('Sales', snapshot.get('segment_retail_sales'))} Stackline sales")
    if snapshot.get("segment_sales_growth_pct") is not None:
        market.append(f"{display_value_for_label('Growth %', snapshot.get('segment_sales_growth_pct'))} growth")
    if movement_short:
        market.append(f"Redshift movement: {movement_short}")
    if market:
        parts.append("; ".join(market) + ".")
    if evidence.get("sunco_coverage"):
        coverage = evidence["sunco_coverage"].split(".", 1)[0]
        if re.search(r"\b0\s+strong\b", coverage, flags=re.IGNORECASE):
            coverage = "0 exact Sunco active-catalog matches"
        parts.append(f"Coverage check: {coverage}.")
    parts.append(f"Channel: {channel}")
    return clean_slide_text(" ".join(part for part in parts if part), 500)


def best_verified_item(normalized: dict[str, Any], channels: set[str]) -> dict[str, Any]:
    """Return the best verified competitor item for a channel set."""
    for item in sort_candidates(as_list(normalized.get("items"))):
        if resolved_source_channel(item) in channels and verification_status(item) == "verified_listing":
            return item
    return {}


def competitor_table_row(item: dict[str, Any], *, direct: bool = False) -> list[Any]:
    """Format the key competitor row for the slide summary."""
    if not item:
        return []
    if direct:
        identifier = listing_identifier(item)
        url = optional_text(item.get("url"))
        identifier_cell: Any = {"value": identifier, "hyperlink": url} if url else identifier
        return [
            identifier_cell,
            item.get("brand"),
            item.get("price"),
            clean_slide_text(item.get("product_title"), 100),
        ]
    return [
        listing_identifier(item),
        item.get("brand"),
        item.get("price"),
        *item_demand_signal_parts(item),
        item.get("rating") or "",
        item.get("review_count") or "",
        listing_link_cell(item),
    ]


def source_footer(packet: dict[str, Any], analysis: dict[str, Any], normalized: dict[str, Any]) -> str:
    """Build a compact source footer for the slide summary."""
    evidence = extract_research_note_evidence(packet)
    performance = as_dict(analysis.get("performance_estimation"))
    snapshot = as_dict(performance.get("market_snapshot"))
    sources = ["Sources: Step 3 packet/analysis"]
    if snapshot.get("segment_name"):
        sources.append(f"Stackline segment: {snapshot.get('segment_name')}")
    if evidence.get("ecommerce_movement") or evidence.get("inventory_support"):
        sources.append("Redshift ecommerce competitor snapshot")
    if evidence.get("sunco_coverage"):
        sources.append("Postgres/Sunco catalog coverage")
    if evidence.get("review_link"):
        sources.append(evidence["review_link"])
    direct_item = best_verified_item(normalized, BM_DIRECT_CHANNELS)
    if direct_item and optional_text(direct_item.get("url")):
        sources.append(optional_text(direct_item.get("url")) or "")
    return clean_slide_text(" | ".join(sources), 420)


def best_slide_image_url(packet: dict[str, Any], normalized: dict[str, Any]) -> str:
    """Return the best available product image URL for a slide summary."""
    reference = as_dict(packet.get("reference_baseline"))
    if optional_text(reference.get("image_url")):
        return optional_text(reference.get("image_url")) or ""
    for item in sort_candidates(as_list(normalized.get("items"))):
        image_url = optional_text(item.get("image_url") or item.get("image"))
        if image_url:
            return image_url
    return ""


def cached_slide_image(session_dir: Path, url: str) -> Path | None:
    """Download and cache an image URL for slide workbook embedding."""
    if not url or not url.startswith(("http://", "https://")):
        return None
    parsed = urlparse(url)
    suffix = Path(parsed.path).suffix.lower()
    if suffix not in {".png", ".jpg", ".jpeg"}:
        suffix = ".jpg"
    cache_dir = session_dir / "reports" / "_slide_image_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    image_path = cache_dir / f"{hashlib.sha1(url.encode('utf-8')).hexdigest()[:16]}{suffix}"
    if image_path.exists() and image_path.stat().st_size:
        return image_path
    try:
        request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(request, timeout=8) as response:
            data = response.read(5_000_000)
        if not data:
            return None
        image_path.write_bytes(data)
        return image_path
    except Exception:
        return None


def add_slide_image(ws, session_dir: Path, image_url: str) -> None:
    """Add a product/reference image to the slide summary when possible."""
    merge_write(
        ws,
        "A4:D15",
        "",
        fill=SLIDE_LIGHT_FILL,
        border=Border(left=SLIDE_THIN_SIDE, right=SLIDE_THIN_SIDE, top=SLIDE_THIN_SIDE, bottom=SLIDE_THIN_SIDE),
    )
    image_path = cached_slide_image(session_dir, image_url)
    if image_path:
        try:
            from openpyxl.drawing.image import Image as XLImage

            image = XLImage(str(image_path))
            image.width = 260
            image.height = 185
            ws.add_image(image, "B5")
            return
        except Exception:
            pass
    if image_url:
        ws["A4"].value = f"Image URL:\n{image_url}"
        ws["A4"].hyperlink = image_url
        ws["A4"].font = SLIDE_LINK_FONT
    else:
        ws["A4"].value = "No image available in Step 3 packet."
        ws["A4"].font = SLIDE_SMALL_FONT


def render_slide_summary_sheet(
    ws,
    session_dir: Path,
    row_number: int,
    packet: dict[str, Any],
    analysis: dict[str, Any],
    normalized: dict[str, Any],
) -> None:
    """Render one slide-ready SKU summary sheet."""
    set_slide_layout(ws)
    identity = as_dict(packet.get("identity"))
    reference = as_dict(packet.get("reference_baseline"))
    pricing = as_dict(analysis.get("pricing_analysis"))
    amazon_snapshot = gate_snapshot(as_dict(analysis.get("gate_readiness")), "amazon", "G2")
    image_url = best_slide_image_url(packet, normalized)

    merge_write(
        ws,
        "A1:J2",
        slide_title(packet),
        fill=SLIDE_WHITE_FILL,
        font=SLIDE_TITLE_FONT,
        border=Border(bottom=Side(style="dotted", color="7F8DA0")),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    merge_write(
        ws,
        "K1:M2",
        "Total SKU Count: 1 variation",
        fill=SLIDE_WHITE_FILL,
        font=Font(color=SLIDE_MUTED, size=9),
        border=Border(bottom=Side(style="dotted", color="7F8DA0")),
        alignment=Alignment(horizontal="right", vertical="center", wrap_text=True),
    )

    add_slide_image(ws, session_dir, image_url)

    ws["E4"].value = "Strategy:"
    ws["E4"].font = Font(color=SLIDE_TEAL, bold=True, underline="single", size=10)
    ws["E4"].alignment = WRAP_ALIGNMENT
    merge_write(
        ws,
        "E5:M8",
        slide_strategy_text(packet, analysis),
        fill=SLIDE_WHITE_FILL,
        font=SLIDE_BODY_FONT,
        border=None,
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    ws["E10"].value = "Target Specs:"
    ws["E10"].font = Font(color=SLIDE_TEAL, bold=True, underline="single", size=10)
    merge_write(
        ws,
        "E11:H16",
        "\n".join(target_spec_lines(packet, analysis)),
        fill=SLIDE_WHITE_FILL,
        font=SLIDE_SMALL_FONT,
        border=None,
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    section_bar(ws, 10, 10, 13, "Sunco Anchor Rev (12mo)")
    write_slide_table(
        ws,
        11,
        10,
        ["Comparable SKU", "Shopify Rev", "Amazon Rev", "Price"],
        [
            [
                identity.get("sunco_reference_sku"),
                reference.get("shopify_revenue_12mo"),
                reference.get("amazon_revenue_12mo"),
                reference.get("listing_price"),
            ]
        ],
        max_rows=1,
    )

    ws["A17"].value = "Product Variants:"
    ws["A17"].font = Font(color=SLIDE_TEAL, bold=True, underline="single", size=10)
    write_span_table(
        ws,
        18,
        [
            ("Concept Tracking Name", 1, 4),
            ("Description", 5, 9),
            ("Key Differentiators", 10, 13),
        ],
        [
            [
                concept_tracking_name(packet),
                clean_slide_text(identity.get("ideation_name"), 120),
                primary_feature_summary(packet, analysis),
            ]
        ],
        max_rows=1,
    )

    ws["A22"].value = "Comparable SKU Performance (12mo):"
    ws["A22"].font = Font(color=SLIDE_TEAL, bold=True, underline="single", size=10)
    write_slide_table(
        ws,
        23,
        1,
        ["Comparable SKU", "Shopify Rev", "Amazon Rev", "Shopify Units", "Amazon Units", "Sale Price", "Trend / Caveat"],
        [
            [
                identity.get("sunco_reference_sku"),
                reference.get("shopify_revenue_12mo"),
                reference.get("amazon_revenue_12mo"),
                reference.get("shopify_units_12mo"),
                reference.get("amazon_units_12mo"),
                reference.get("listing_price"),
                related_sales_slide_note(reference),
            ]
        ],
        max_rows=1,
    )

    amazon_item = best_verified_item(normalized, AMAZON_CHANNELS)
    ws["A27"].value = "Amazon Segment Competitor"
    ws["A27"].font = Font(color=SLIDE_DARK_NAVY, bold=True, size=8)
    ws["C27"].value = f"Amazon G2 {display_value_for_label('Score', amazon_snapshot.get('weighted_score')) or 'n/a'} | evidence {normalize_text(as_dict(amazon_snapshot.get('evidence_confidence')).get('label')) or 'n/a'}"
    ws["C27"].font = SLIDE_MUTED_FONT
    write_slide_table(
        ws,
        28,
        1,
        ["ASIN / ID", "Brand", "Price", "Volume", "Rate / Share", "Window", "Rating", "Reviews", "Source"],
        [competitor_table_row(amazon_item)],
        header_fill=SLIDE_NAVY_FILL,
        max_rows=1,
    )

    direct_item = best_verified_item(normalized, BM_DIRECT_CHANNELS)
    ws["A32"].value = "Direct Segment Competitor"
    ws["A32"].font = Font(color=SLIDE_DARK_NAVY, bold=True, size=8)
    write_span_table(
        ws,
        33,
        [
            ("Product Code", 1, 3),
            ("Brand", 4, 5),
            ("Price", 6, 7),
            ("Product Name", 8, 13),
        ],
        [competitor_table_row(direct_item, direct=True)],
        header_fill=SLIDE_NAVY_FILL,
        max_rows=1,
    )

    merge_write(
        ws,
        "A36:M36",
        source_footer(packet, analysis, normalized),
        fill=SLIDE_WHITE_FILL,
        font=SLIDE_MUTED_FONT,
        border=None,
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    ws["M37"].value = "Sunco"
    ws["M37"].font = Font(color=SLIDE_TEAL, bold=True, size=10)
    ws["M37"].alignment = Alignment(horizontal="right", vertical="bottom")


def build_slide_index_sheet(ws, payloads: list[tuple[int, dict[str, Any], dict[str, Any], dict[str, Any]]]) -> None:
    """Build a clean index sheet for the slide summary workbook."""
    set_default_layout(ws)
    ws.title = "Index"
    ws.cell(row=1, column=1, value="Step 3 Slide Summary Index")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    rows = []
    used_titles = {ws.title}
    for row_number, packet, analysis, normalized in payloads:
        rows.append(
            [
                row_number,
                concept_tracking_name(packet),
                slide_title(packet),
                recommended_product_action(packet, analysis),
                revision_target_sku(packet, analysis),
                revision_change_summary(packet, analysis),
                product_fit_classification(packet),
                channel_strategy(packet, analysis),
                as_dict(analysis.get("performance_estimation")).get("launch_outlook"),
                as_dict(analysis.get("performance_estimation")).get("confidence"),
                slide_sheet_title_for_row(row_number, packet, used_titles),
            ]
        )
    write_table(
        ws,
        3,
        "Slide Summary Sheets",
        ["Row", "Concept Tracking Name", "Slide Title", "Recommended Product Action", "Revision Target SKU", "Recommended Revision Changes", "Gap Reason", "Channel Strategy", "Outlook", "Confidence", "Worksheet"],
        rows,
    )


def slide_summary_path_for(destination: Path, session_dir: Path) -> Path:
    """Resolve the companion slide-summary workbook path for a detailed workbook."""
    if destination.name.endswith("_completed_rows.xlsx"):
        return destination.with_name(destination.name.replace("_completed_rows.xlsx", "_slide_summaries.xlsx"))
    return destination.with_name(f"{destination.stem}_slide_summaries{destination.suffix}")


def build_slide_summary_workbook(
    session_dir: Path,
    payloads: list[tuple[int, dict[str, Any], dict[str, Any], dict[str, Any]]],
    output_path: Path,
) -> Path | None:
    """Build the slide-ready companion workbook with one sheet per concept."""
    if not payloads:
        return None

    wb = Workbook()
    index_ws = wb.active
    build_slide_index_sheet(index_ws, payloads)

    used_titles = {index_ws.title}
    for row_number, packet, analysis, normalized in payloads:
        title = slide_sheet_title_for_row(row_number, packet, used_titles)
        ws = wb.create_sheet(title=title)
        render_slide_summary_sheet(ws, session_dir, row_number, packet, analysis, normalized)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def completed_row_payloads(
    session_dir: Path,
    rows: list[int] | None = None,
) -> list[tuple[int, dict[str, Any], dict[str, Any], dict[str, Any]]]:
    """Load completed packet/analysis/normalized payloads for selected rows."""
    manifest = read_json(session_dir / "manifest.json")
    target_rows = set(rows or [row["row_number"] for row in manifest.get("rows", [])])
    payloads = []

    for row in manifest.get("rows", []):
        row_number = row["row_number"]
        if row_number not in target_rows:
            continue
        if row.get("stages", {}).get("analyzed") != "complete":
            continue
        packet = read_json(packet_path_for(session_dir, row_number))
        analysis = read_json(artifact_path_for(session_dir, row_number, "analyzed"))
        normalized = read_json(artifact_path_for(session_dir, row_number, "normalized"))
        payloads.append((row_number, packet, analysis, normalized))

    return payloads


def build_report(session_dir: Path, row_number: int) -> Path | None:
    """Build one Excel report for a completed analyzed row."""
    payloads = completed_row_payloads(session_dir, rows=[row_number])
    if not payloads:
        return None

    _, packet, analysis, normalized = payloads[0]
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title_for_row(row_number, packet, set())
    render_row_sheet(ws, row_number, packet, analysis, normalized)

    report_path = artifact_path_for(session_dir, row_number, "reported")
    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(report_path)
    return report_path


def build_summary_sheet(ws, payloads: list[tuple[int, dict[str, Any], dict[str, Any], dict[str, Any]]]) -> None:
    """Build a summary sheet for the combined workbook."""
    set_default_layout(ws)
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="Step 3 Research Summary")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    row = key_value_rows(
        ws,
        3,
        [
            ("Completed Row Count", len(payloads)),
            ("Workbook Scope", "One summary sheet plus one sheet per completed ideation row"),
        ],
        columns=1,
    )
    row = write_table(
        ws,
        row + 1,
        "Summary Metric Guide",
        ["Metric", "What It Means", "Question It Answers"],
        SUMMARY_METRIC_GUIDE_ROWS,
    )
    used_titles = {ws.title}
    completed_rows = []
    for row_number, packet, analysis, _ in payloads:
        identity = as_dict(packet.get("identity"))
        amazon_g2 = next(
            (
                snapshot.get("weighted_score")
                for snapshot in as_list(as_dict(analysis.get("gate_readiness")).get("snapshots"))
                if snapshot.get("channel") == "amazon" and snapshot.get("gate") == "G2"
            ),
            None,
        )
        amazon_evidence = next(
            (
                as_dict(snapshot.get("evidence_confidence")).get("label")
                for snapshot in as_list(as_dict(analysis.get("gate_readiness")).get("snapshots"))
                if snapshot.get("channel") == "amazon" and snapshot.get("gate") == "G2"
            ),
            None,
        )
        completed_rows.append(
            [
                row_number,
                concept_tracking_name(packet),
                normalize_text(identity.get("ideation_name")),
                recommended_product_action(packet, analysis),
                revision_target_sku(packet, analysis),
                revision_change_summary(packet, analysis),
                product_fit_classification(packet),
                channel_strategy(packet, analysis),
                as_dict(analysis.get("performance_estimation")).get("launch_outlook"),
                as_dict(analysis.get("performance_estimation")).get("confidence"),
                amazon_g2,
                amazon_evidence,
                sheet_title_for_row(row_number, packet, used_titles),
            ]
        )
    write_table(
        ws,
        row,
        "Completed Ideation Rows",
        ["Row", "Concept Tracking Name", "Ideation", "Recommended Product Action", "Revision Target SKU", "Recommended Revision Changes", "Gap Reason", "Channel Strategy", "Outlook", "Confidence", "Amazon G2", "Amazon Evidence", "Worksheet"],
        completed_rows,
    )


def build_combined_workbook(
    session_dir: Path,
    rows: list[int] | None = None,
    output_path: str | None = None,
) -> Path | None:
    """Build one workbook with a summary sheet plus one sheet per completed row."""
    payloads = completed_row_payloads(session_dir, rows=rows)
    if not payloads:
        return None
    payloads.sort(key=payload_action_sort_key)

    wb = Workbook()
    summary_ws = wb.active
    build_summary_sheet(summary_ws, payloads)

    used_titles = {summary_ws.title}
    for row_number, packet, analysis, normalized in payloads:
        title = sheet_title_for_row(row_number, packet, used_titles)
        ws = wb.create_sheet(title=title)
        render_row_sheet(ws, row_number, packet, analysis, normalized)

    if output_path:
        destination = Path(output_path).resolve()
    else:
        destination = session_dir / "reports" / f"{session_dir.name}_completed_rows.xlsx"
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)
    return destination


def parse_rows_argument(value: str | None) -> list[int] | None:
    """Parse an optional comma-separated row list."""
    if not value:
        return None
    rows = []
    for part in value.split(","):
        part = part.strip()
        if not part:
            continue
        rows.append(int(part))
    return rows or None


def build_reports(
    session_root: str,
    rows: list[int] | None = None,
    combined: bool = False,
    output_path: str | None = None,
) -> dict[str, Any]:
    """Build report artifacts for selected session rows."""
    session_dir = Path(session_root).resolve()

    if combined:
        combined_path = build_combined_workbook(session_dir, rows=rows, output_path=output_path)
        update_result = update_session(str(session_dir))
        return {
            "session_root": str(session_dir),
            "rows_requested": sorted(rows or []),
            "combined_workbook": str(combined_path) if combined_path else None,
            "manifest_summary": update_result["summary"],
        }

    manifest = read_json(session_dir / "manifest.json")
    target_rows = set(rows or [row["row_number"] for row in manifest.get("rows", [])])

    written_rows = []
    skipped_rows = []
    report_files = []

    for row in manifest.get("rows", []):
        row_number = row["row_number"]
        if row_number not in target_rows:
            continue
        report_path = build_report(session_dir, row_number)
        if report_path is None:
            skipped_rows.append(row_number)
            continue
        written_rows.append(row_number)
        report_files.append(str(report_path))

    update_result = update_session(str(session_dir))
    return {
        "session_root": str(session_dir),
        "rows_requested": sorted(target_rows),
        "rows_written": written_rows,
        "rows_skipped": skipped_rows,
        "report_files": report_files,
        "manifest_summary": update_result["summary"],
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build Excel report artifacts for completed analyzed rows."
    )
    parser.add_argument("session_root", help="Path to an initialized research session.")
    parser.add_argument(
        "--rows",
        default=None,
        help="Optional comma-separated row numbers to report.",
    )
    parser.add_argument(
        "--combined",
        action="store_true",
        help="Build one combined workbook with one sheet per completed row.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Optional explicit output path for the combined workbook.",
    )
    args = parser.parse_args()

    result = build_reports(
        args.session_root,
        rows=parse_rows_argument(args.rows),
        combined=args.combined,
        output_path=args.output,
    )
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
