from __future__ import annotations

import ctypes
import platform
import subprocess
import time
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


def _retry_com_cleanup(action) -> Exception | None:
    last_error: Exception | None = None
    for _ in range(3):
        try:
            action()
            return None
        except Exception as exc:
            last_error = exc
            time.sleep(0.5)
    return last_error


def _retry_com_result(action):
    last_error: Exception | None = None
    for _ in range(5):
        try:
            return action(), None
        except Exception as exc:
            last_error = exc
            time.sleep(0.75)
    return None, last_error


def _excel_process_id(excel) -> int | None:
    try:
        hwnd = int(excel.Hwnd)
        pid = ctypes.c_ulong()
        ctypes.windll.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
        return int(pid.value) or None
    except Exception:
        return None


def _terminate_excel_process(pid: int) -> None:
    subprocess.run(
        ["taskkill", "/PID", str(pid), "/T", "/F"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=False,
    )


def validate_workbook(path: Path, required_sheets: list[str]) -> list[str]:
    issues: list[str] = []
    if not path.exists():
        return [f"Workbook does not exist: {path}"]

    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:
        return [f"Workbook failed to open with openpyxl: {exc}"]

    for sheet in required_sheets:
        if sheet not in workbook.sheetnames:
            issues.append(f"Missing required sheet: {sheet}")

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(token in value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")):
                    issues.append(f"Formula/error text found at {sheet.title}!{cell.coordinate}: {value}")
    workbook.close()
    issues.extend(validate_sheet_views(path))
    return issues


def validate_sheet_views(path: Path) -> list[str]:
    issues: list[str] = []
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    try:
        with ZipFile(path) as archive:
            sheet_files = [
                name for name in archive.namelist()
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            ]
            for name in sheet_files:
                root = ET.fromstring(archive.read(name))
                for view in root.findall(".//x:sheetView", ns):
                    pane = view.find("x:pane", ns)
                    if pane is None:
                        continue
                    x_split = pane.get("xSplit")
                    y_split = pane.get("ySplit")
                    valid_panes = {"topLeft"}
                    if x_split:
                        valid_panes.add("topRight")
                    if y_split:
                        valid_panes.add("bottomLeft")
                    if x_split and y_split:
                        valid_panes.add("bottomRight")
                    seen: set[tuple[str, str, str]] = set()
                    for selection in view.findall("x:selection", ns):
                        selection_pane = selection.get("pane") or "topLeft"
                        if selection_pane not in valid_panes:
                            issues.append(f"Invalid worksheet view in {name}: selection references missing pane {selection_pane}.")
                        key = (
                            selection_pane,
                            selection.get("activeCell") or "",
                            selection.get("sqref") or "",
                        )
                        if key in seen:
                            issues.append(f"Duplicate worksheet view selection in {name}: {selection_pane} {key[1]} {key[2]}.")
                        seen.add(key)
    except Exception as exc:
        issues.append(f"Workbook XML view validation failed: {exc}")
    return issues


def try_excel_com_open_save(path: Path) -> tuple[bool, str]:
    if platform.system().lower() != "windows":
        return False, "Excel COM validation skipped: this is not Windows."

    try:
        import win32com.client  # type: ignore
    except Exception as exc:
        return False, f"Excel COM validation skipped: pywin32 is not available ({exc})."

    excel = None
    workbook = None
    excel_pid = None
    ok = False
    message = ""
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel_pid = _excel_process_id(excel)
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook, exc = _retry_com_result(lambda: excel.Workbooks.Open(str(path.resolve())))
        if exc:
            raise exc
        exc = _retry_com_cleanup(lambda: workbook.Save())
        if exc:
            raise exc
        ok = True
        message = "Excel COM open/save validation passed."
    except Exception as exc:
        message = f"Excel COM validation failed: {exc}"
    finally:
        cleanup_errors: list[str] = []
        if workbook is not None:
            exc = _retry_com_cleanup(lambda: workbook.Close(SaveChanges=False))
            if exc:
                cleanup_errors.append(f"workbook close failed: {exc}")
        if excel is not None:
            exc = _retry_com_cleanup(lambda: excel.Quit())
            if exc:
                cleanup_errors.append(f"Excel quit failed: {exc}")
        if excel_pid is not None:
            time.sleep(0.5)
            _terminate_excel_process(excel_pid)
        if cleanup_errors:
            ok = False
            message = f"{message} Cleanup warning: {' | '.join(cleanup_errors)}"
    return ok, message
