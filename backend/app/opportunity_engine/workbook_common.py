from __future__ import annotations

import json
import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .paths import ProjectPaths


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def parse_generated_date(payload: dict[str, Any]) -> date | None:
    raw = payload.get("generated") or payload.get("generated_at")
    if not raw:
        return None
    text = str(raw)[:10]
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def age_days(generated: date | None) -> int | None:
    if generated is None:
        return None
    return (date.today() - generated).days


def ensure_template_copy(template: Path, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)


def clear_sheet_rows(ws, start_row: int) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)
    if hasattr(ws, "_images"):
        ws._images = []


def clear_row_values(ws, row_index: int, max_col: int | None = None) -> None:
    limit = max_col or ws.max_column
    for col in range(1, limit + 1):
        ws.cell(row_index, col).value = None


def copy_row_style(ws, template_row: int, target_row: int, max_col: int | None = None) -> None:
    limit = max_col or ws.max_column
    for col in range(1, limit + 1):
        src = ws.cell(template_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def add_or_replace_audit_sheet(workbook, rows: list[tuple[str, str]], sql_text: str) -> None:
    if "Run Audit" in workbook.sheetnames:
        del workbook["Run Audit"]
    ws = workbook.create_sheet("Run Audit")
    ws.append(["Audit Item", "Value"])
    for label, value in rows:
        ws.append([label, value])
    ws.append([])
    ws.append(["SQL Used", sql_text])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120
    ws.row_dimensions[ws.max_row].height = 120
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def find_cached_image(paths: ProjectPaths, original_path: str | None) -> Path | None:
    if not original_path:
        return None
    original = Path(original_path)
    candidates = []
    if original.exists():
        candidates.append(original)
    image_root = paths.cache / "images"
    if image_root.exists():
        candidates.extend(image_root.rglob(original.name))
    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def add_image(ws, image_path: Path | None, anchor: str, width: int = 96, height: int = 96) -> str:
    if image_path is None:
        return "missing"
    try:
        from openpyxl.drawing.image import Image as XLImage

        image = XLImage(str(image_path))
        image.width = width
        image.height = height
        ws.add_image(image, anchor)
        return "embedded"
    except Exception as exc:
        ws[anchor].value = f"Image unavailable: {image_path.name}"
        return f"failed: {exc}"


def write_headers(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = max(14, min(36, len(header) + 4))


def create_minimal_workbook(path: Path, sheets: dict[str, list[str]]) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, headers in sheets.items():
        ws = workbook.create_sheet(sheet_name)
        if headers:
            write_headers(ws, headers)
    workbook.save(path)
