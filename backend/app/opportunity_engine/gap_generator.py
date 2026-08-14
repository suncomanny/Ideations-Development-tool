from __future__ import annotations

from contextlib import contextmanager
from copy import deepcopy
from datetime import date, datetime
import os
from pathlib import Path
import re
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .categories import Category
from .category_intelligence import format_intelligence_audit, load_category_intelligence
from .line_review import (
    prepare_line_review_context,
    run_audit_rows as line_review_run_audit_rows,
    source_audit_rows as line_review_source_audit_rows,
    write_line_review_sheet,
)
from .paths import ProjectPaths
from .refresh_redshift_stackline_cache import cache_created_date, latest_category_cache
from .sql_audit import collect_sql_text
from .utils import slugify, timestamp
from .validation import try_excel_com_open_save, validate_workbook
from .workbook_common import (
    add_image,
    add_or_replace_audit_sheet,
    age_days,
    create_minimal_workbook,
    ensure_template_copy,
    find_cached_image,
    load_json,
    parse_generated_date,
)


TRUE_GAP_HEADERS = {
    "Summary": [],
    "Recommendations": [
        "Image",
        "Subcategory",
        "PM Recommendation Name",
        "Recommendation",
        "Priority",
        "Confidence",
        "Competitor / Ecommerce Example",
        "Review Link",
        "Sunco Active-Catalog Check",
        "Why This Is A True Gap",
        "PM Action",
    ],
    "Sources and Audit": ["Type", "Subcategory", "Item", "Evidence / Note", "URL"],
    "Amazon Recommendations": [
        "Image",
        "Subcategory",
        "PM Recommendation Name",
        "Amazon-channel recommendation",
        "Priority",
        "Confidence",
        "Amazon classification",
        "Stackline / Amazon evidence",
        "Sunco Amazon coverage check",
        "Example listing",
        "Review link",
        "PM action",
    ],
    "Amazon Source Audit": ["Source", "Evidence captured", "Used for", "Caveat", "Reference"],
}


ALIASES = {
    "panels": {"panels", "panel lights", "ceiling panels", "ceiling panel lights", "ceiling panels / fixtures", "flat panels"},
    "ceiling_fixtures": {"ceiling fixtures", "residential decor fixtures"},
    "wall_sconces": {"wall sconces", "wall sconces"},
    "under_cabinet": {"under cabinet", "under cabinet / tape"},
    "tape_rope_light": {"tape/rope light", "under cabinet / tape"},
}

SUPPLEMENTAL_CANDIDATES_PER_TAB = 1

RELATED_CATEGORY_BACKFILL = {
    "ceiling_fixtures": ["chandeliers", "pendants", "lamps", "wall_sconces"],
    "lamps": ["ceiling_fixtures", "chandeliers", "pendants", "wall_sconces"],
    "chandeliers": ["pendants", "ceiling_fixtures", "lamps", "wall_sconces", "vanity"],
    "pendants": ["chandeliers", "ceiling_fixtures", "lamps", "wall_sconces"],
    "wall_sconces": ["vanity", "ceiling_fixtures", "pendants", "chandeliers"],
    "vanity": ["wall_sconces", "ceiling_fixtures", "chandeliers", "pendants"],
    "under_cabinet": ["tape_rope_light", "ceiling_fixtures", "lamps"],
    "tape_rope_light": ["under_cabinet", "ceiling_fixtures", "lamps"],
}

SUPPLEMENTAL_EXCLUDE_TERMS = {
    "chandeliers": ["fan", "puck", "under-cabinet", "under cabinet", "tape", "rope", "sconce", "vanity", "floor lamp", "table lamp"],
}

DECISION_TREE_STEPS = [
    "1. Category fit: only exact selected category rows are promoted first; parent categories are ignored.",
    "2. Demand/assortment signal: require a competitor/ecommerce example, Amazon/Stackline/BSR evidence when available, or a reviewed public listing/category page.",
    "3. Sunco gap gate: require evidence that Sunco does not already have active comparable coverage; Sunco.com/Shopify coverage has priority over marketplace-only checks.",
    "4. Actionability gate: require a PM action that can plausibly become a sourced SKU/family, variant, bundle, or channel listing.",
    "5. Ranking: exact-category rows sort ahead of adjacent candidates; High priority and High confidence sort ahead of Medium/Low.",
    "6. Natural count +1: if exact results are sparse, include up to one adjacent candidate per recommendation tab, clearly labeling it as exploratory.",
    "7. Research handoff: Step 2/Step 3 refine attributes, live links, market pricing, competitive listings, and final launch fit before this becomes a PRD-ready product.",
]

SUCCESS_PROXY_TEXT = (
    "The workbook does not claim guaranteed launch success. It chooses higher-probability ideations by triangulating "
    "market pull, competitor catalog coverage, Amazon/Stackline or BSR evidence where available, a documented Sunco "
    "assortment gap, and a concrete PM action path. Ideas with weaker category fit are labeled as supplemental and "
    "remain exploratory until refreshed evidence supports them."
)

INCLUDE_LEGACY_GAP_EVIDENCE_ENV = "OPPORTUNITY_ENGINE_INCLUDE_LEGACY_GAP_EVIDENCE"


@contextmanager
def _category_run_lock(paths: ProjectPaths, category: Category):
    lock_dir = paths.cache / "run_locks"
    lock_dir.mkdir(parents=True, exist_ok=True)
    lock_path = lock_dir / f"step1_{category.slug}.lock"
    try:
        handle = lock_path.open("x", encoding="utf-8")
    except FileExistsError as exc:
        raise RuntimeError(
            f"Step 1 is already running for {category.run_name}. "
            "Wait for the existing run to finish before starting another one."
        ) from exc
    try:
        handle.write(timestamp())
        handle.close()
        yield
    finally:
        try:
            lock_path.unlink()
        except FileNotFoundError:
            pass


def _match_category(category: Category, raw: str | None) -> bool:
    if not raw:
        return False
    category_slug = category.slug
    raw_text = raw.strip().lower()
    raw_slug = slugify(raw_text)
    if raw_slug == category_slug:
        return True
    aliases = ALIASES.get(category_slug, set())
    if raw_text in aliases:
        return True
    return category_slug in {slugify(part.strip()) for part in raw_text.split("/")}


def _category_slug(value: str | None) -> str:
    if not value:
        return ""
    raw = value.strip().lower()
    for slug, aliases in ALIASES.items():
        if raw in aliases:
            return slug
    return slugify(value)


def _row_identity(row: dict[str, Any]) -> str:
    return str(row.get("recommendation") or "").strip().lower()


def _unique_warnings(warnings: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for warning in warnings:
        if warning in seen:
            continue
        seen.add(warning)
        unique.append(warning)
    return unique


def _priority_score(row: dict[str, Any]) -> tuple[int, int]:
    priority = str(row.get("priority") or "").lower()
    confidence = str(row.get("confidence") or "").lower()
    priority_score = {"high": 0, "medium": 1, "low": 2}.get(priority, 3)
    confidence_score = {"high": 0, "medium": 1, "low": 2}.get(confidence, 3)
    return priority_score, confidence_score


def _is_high(value: str | None) -> bool:
    return str(value or "").strip().lower() == "high"


def _has_gap_language(value: str | None) -> bool:
    text = str(value or "").lower()
    return any(
        phrase in text
        for phrase in [
            "active title match: 0",
            "found no",
            "no exact",
            "no sunco",
            "gap",
            "not show",
            "does not show",
            "0 for",
        ]
    )


def _has_market_signal(row: dict[str, Any], source_kind: str) -> bool:
    fields = ["example", "evidence", "why_gap", "source_url", "review_url", "source_systems"]
    haystack = " ".join(str(row.get(field) or "") for field in fields).lower()
    if "amazon" in source_kind.lower():
        return any(term in haystack for term in ["stackline", "bsr", "sales", "asin", "review", "amazon"])
    return bool(row.get("example") or row.get("source_url") or row.get("why_gap"))


def _decision_tree_text() -> str:
    return "\n".join(DECISION_TREE_STEPS)


def _decision_outcome(row: dict[str, Any], source_kind: str) -> str:
    if row.get("_supplemental"):
        return "Exploratory adjacent seed - included only as the +1 candidate and not yet proven."
    if _is_high(row.get("priority")) and _is_high(row.get("confidence")) and _has_market_signal(row, source_kind) and _has_gap_language(row.get("sunco_check")):
        return "High-confidence exact gap - strongest Step 1 selection tier."
    if _is_high(row.get("priority")) or _is_high(row.get("confidence")):
        return "Qualified exact-category candidate - selected for Step 2 review."
    return "Lower-confidence exact-category candidate - review before moving into Step 2."


def _row_decision_rationale(row: dict[str, Any], source_kind: str) -> str:
    signals = [
        "Exact category match" if not row.get("_supplemental") else f"Adjacent candidate from {row.get('_source_category') or row.get('subcategory') or 'unknown category'}",
        f"Priority: {row.get('priority') or 'Unknown'}",
        f"Confidence: {row.get('confidence') or 'Unknown'}",
    ]
    if _has_market_signal(row, source_kind):
        signals.append("Market signal present")
    if row.get("source_url") or row.get("review_url"):
        signals.append("Review URL present")
    if _has_gap_language(row.get("sunco_check")):
        signals.append("Sunco coverage gap documented")
    if row.get("pm_action") or row.get("action"):
        signals.append("PM action path present")

    cautions: list[str] = []
    if row.get("_supplemental"):
        cautions.append("Supplemental row; review before treating as true gap")
    if not (_is_high(row.get("priority")) and _is_high(row.get("confidence"))):
        cautions.append("Priority/confidence is not High/High")
    if not _has_gap_language(row.get("sunco_check")):
        cautions.append("Sunco gap language should be manually reviewed")

    text = f"Decision outcome: {_decision_outcome(row, source_kind)}\nSelection signals: " + "; ".join(signals)
    if cautions:
        text += "\nCautions: " + "; ".join(cautions)
    return text


CONTROLLED_ACTIONS = ("NPD", "Revision", "Concept Review")

SOURCE_NAME_MARKERS = (
    "amazon stackline opportunity",
    "amazon-channel recommendation",
    "stackline/amazon leader",
    "stackline amazon leader",
    "amazon recommendation",
    "shopify ecommerce candidate",
    "ecommerce competitor demand candidate",
    "competitor evidence",
    "source audit",
    "step 1 evidence",
    "home depot marketplace",
)


def _controlled_action(row: dict[str, Any]) -> str:
    text = " ".join(
        str(row.get(key) or "")
        for key in ["classification", "recommendation", "pm_action", "action", "gap_reason"]
    ).lower()
    if "revision" in text or "existing sunco coverage" in text or "coverage exists" in text or "sunco amazon incumbent" in text:
        return "Revision"
    if "concept review" in text or "possible feature gap" in text or "strategic outlier" in text or "watchlist" in text or "partial coverage" in text:
        return "Concept Review"
    return "NPD"


def _step1_name_text(row: dict[str, Any]) -> str:
    return "\n".join(
        str(row.get(key) or "")
        for key in [
            "recommendation",
            "classification",
            "example",
            "why_gap",
            "evidence",
            "sunco_check",
            "pm_action",
            "action",
        ]
    )


def _title_name_component(value: str) -> str:
    acronyms = {
        "led": "LED",
        "cct": "CCT",
        "cri": "CRI",
        "ip": "IP",
        "ul": "UL",
        "etl": "ETL",
        "fcc": "FCC",
        "0-10v": "0-10V",
        "triac": "TRIAC",
        "ft": "ft",
        "in": "in",
        "lm": "lm",
    }
    pieces = re.split(r"(\s+|/|-)", value.strip())
    output: list[str] = []
    for piece in pieces:
        key = piece.lower()
        if not piece or piece.isspace() or piece in {"/", "-"}:
            output.append(piece)
        elif key in acronyms:
            output.append(acronyms[key])
        elif re.fullmatch(r"\d+(?:\.\d+)?(?:x\d+(?:\.\d+)?|ft|in|w|k|lm)?", key):
            output.append(piece)
        elif re.fullmatch(r"\d+cct", key):
            output.append(piece.upper())
        else:
            output.append(piece[:1].upper() + piece[1:].lower())
    text = "".join(output)
    text = re.sub(r"\bLed\b", "LED", text)
    text = re.sub(r"\bCct\b", "CCT", text)
    text = re.sub(r"\bIp(\d+)\b", r"IP\1", text)
    text = re.sub(r"\b0-10v\b", "0-10V", text, flags=re.IGNORECASE)
    return text


def _dedupe_name_parts(parts: list[str]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for part in parts:
        clean = re.sub(r"\s+", " ", str(part or "")).strip(" ,;:-")
        normalized = re.sub(r"[^a-z0-9]+", "", clean.lower())
        if not normalized or normalized in seen:
            continue
        if any(normalized in re.sub(r"[^a-z0-9]+", "", existing.lower()) for existing in output):
            continue
        seen.add(normalized)
        output.append(clean)
    return output


def _format_number(value: str) -> str:
    number = float(value.replace(",", ""))
    if number.is_integer():
        return f"{int(number):,}"
    return f"{number:,.1f}".rstrip("0").rstrip(".")


def _extract_size_token(text: str) -> str:
    for pattern, formatter in [
        (r"\b(\d+(?:\.\d+)?)\s*(?:ft|foot|feet|')\b", lambda m: f"{m.group(1).rstrip('0').rstrip('.') if '.' in m.group(1) else m.group(1)} ft"),
        (r"\b(\d+)\s*[xX]\s*(\d+)\b", lambda m: f"{m.group(1)}x{m.group(2)}"),
        (r"\b(\d+(?:\.\d+)?)\s*(?:in|inch|inches|\")\b", lambda m: f"{m.group(1).rstrip('0').rstrip('.') if '.' in m.group(1) else m.group(1)} in"),
    ]:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return formatter(match)
    return ""


def _extract_wattage_text(text: str) -> str:
    for pattern in [
        r"\bmax parsed wattage\s+([0-9]+(?:\.[0-9]+)?)\s*W\b",
        r"\bload target\s+([0-9]+(?:\.[0-9]+)?)\s*W\b",
        r"\b([0-9]+(?:\.[0-9]+)?)\s*W(?:att)?\b",
    ]:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return f"{_format_number(match.group(1))}W"
    return ""


def _extract_lumens_text(text: str) -> str:
    patterns = [
        r"\bbrightness target\s+([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:lm|lumens)\b",
        r"\b([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:lm|lumens)\b",
        r"\b([0-9][0-9,]*(?:\.[0-9]+)?)\s+lumen(?:s)?\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            trailing = text[match.end(): match.end() + 8]
            if re.match(r"\s*/\s*w", trailing, flags=re.IGNORECASE):
                continue
            numeric = float(match.group(1).replace(",", ""))
            if numeric < 300 or numeric > 200000:
                continue
            return f"{_format_number(match.group(1))}lm"
    return ""


def _extract_cct_text(text: str) -> str:
    lowered = text.lower()
    values: list[str] = []
    for match in re.finditer(r"\b((?:27|30|35|40|50|65)00|\d{4})\s*(?:K|Kelvin)\b", text, flags=re.IGNORECASE):
        value = f"{match.group(1)}K".upper()
        if value not in values:
            values.append(value)
    if values:
        return "/".join(values[:4])
    if "5cct" in lowered:
        return "5CCT"
    if "3cct" in lowered:
        return "3CCT"
    if "selectable cct" in lowered or "switchable cct" in lowered:
        return "Selectable CCT"
    return ""


def _infer_form_factor(category: Category, text: str) -> str:
    lowered = text.lower()
    category_text = category.run_name.replace("_", " ").lower()
    combined = f"{lowered} {category_text}"
    labels = [
        (("vapor tight", "vaportight", "vapor proof"), "LED Vapor Tight Fixture"),
        (("wraparound", "wrap around"), "LED Wraparound Fixture"),
        (("troffer", "center basket"), "LED Troffer"),
        (("flat panel", "panel"), "LED Panel"),
        (("linear high bay",), "Linear High Bay"),
        (("high bay",), "LED High Bay"),
        (("shop light",), "LED Shop Light"),
        (("strip light", "striplight"), "LED Strip Light"),
        (("canopy",), "LED Canopy Light"),
        (("wall pack",), "LED Wall Pack"),
        (("flood",), "LED Flood Light"),
        (("emergency", "exit sign"), "Emergency/Exit Fixture"),
    ]
    for needles, label in labels:
        if any(needle in combined for needle in needles):
            return label
    category_name = _title_name_component(category.run_name.replace("_", " "))
    if "fixture" in category_name.lower() or "light" in category_name.lower():
        return category_name
    return f"LED {category_name} Fixture"


def _extract_name_features(text: str) -> list[str]:
    lowered = text.lower()
    features: list[str] = []
    for needle, label in [
        ("selectable wattage", "Selectable Wattage"),
        ("wattage selectable", "Selectable Wattage"),
        ("multi-wattage", "Selectable Wattage"),
        ("emergency backup", "Emergency Backup"),
        ("emergency battery", "Emergency Backup"),
        ("motion sensor", "Motion Sensor"),
        ("daylight sensor", "Daylight Sensor"),
        ("smart", "Smart Connected"),
        ("linkable", "Linkable"),
        ("0-10v", "0-10V Dimming"),
        ("0-10 v", "0-10V Dimming"),
        ("triac", "TRIAC Dimming"),
        ("frosted lens", "Frosted Lens"),
        ("prismatic lens", "Prismatic Lens"),
        ("polycarbonate lens", "Polycarbonate Lens"),
        ("daisy chain", "Daisy-Chain Ready"),
        ("daisy-chain", "Daisy-Chain Ready"),
    ]:
        if needle in lowered and label not in features:
            features.append(label)
    ip_match = re.search(r"\bIP\s*([0-9]{2})\b", text, flags=re.IGNORECASE)
    if ip_match:
        features.append(f"IP{ip_match.group(1)}")
    return features


def _pm_recommendation_name(category: Category, row: dict[str, Any]) -> str:
    action = _controlled_action(row)
    text = _step1_name_text(row)
    form_factor = _infer_form_factor(category, text)
    size = _extract_size_token(text)
    name_base = f"{size} {form_factor}".strip() if size and size.lower() not in form_factor.lower() else form_factor
    parts = _dedupe_name_parts(
        [
            _title_name_component(name_base),
            _extract_wattage_text(text),
            _extract_lumens_text(text),
            _extract_cct_text(text),
            *_extract_name_features(text),
        ]
    )
    body = ", ".join(parts) if parts else _title_name_component(form_factor)
    if any(marker in body.lower() for marker in SOURCE_NAME_MARKERS):
        body = _title_name_component(form_factor)
    return f"{action}: {body}"


def _related_rank(category: Category, row: dict[str, Any]) -> int:
    related = RELATED_CATEGORY_BACKFILL.get(category.slug, [])
    row_slug = _category_slug(row.get("subcategory"))
    if row_slug in related:
        return related.index(row_slug)
    return 99


def _backfill_allowed(category: Category, row: dict[str, Any]) -> bool:
    haystack = " ".join(
        str(row.get(key) or "")
        for key in ["subcategory", "recommendation", "example", "classification"]
    ).lower()
    excluded = SUPPLEMENTAL_EXCLUDE_TERMS.get(category.slug, [])
    return not any(term in haystack for term in excluded)


def _source_supplement(row: dict[str, Any], category: Category) -> dict[str, Any]:
    supplement = deepcopy(row)
    source_category = row.get("subcategory") or "Unknown"
    original_confidence = row.get("confidence") or "Unknown"
    original_priority = row.get("priority") or "Unknown"
    supplement["_supplemental"] = True
    supplement["_source_category"] = source_category
    supplement["subcategory"] = category.run_name
    supplement["recommendation"] = f"[Supplemental candidate] {row.get('recommendation')}"
    supplement["priority"] = "Exploratory"
    supplement["confidence"] = "Needs category fit review"
    supplement["example"] = (
        f"SUPPLEMENTAL WARNING: original source category was {source_category}; "
        f"original priority/confidence was {original_priority}/{original_confidence}. "
        f"{row.get('example') or ''}"
    ).strip()
    supplement["sunco_check"] = (
        f"SUPPLEMENTAL WARNING: this is not yet proven as an exact {category.run_name} gap. "
        f"It was included as the +1 adjacent candidate and needs category fit review before Step 2. "
        f"{row.get('sunco_check') or ''}"
    ).strip()
    supplement["why_gap"] = (
        f"Adjacent-category seed for {category.run_name}; review category fit, search demand, and Sunco coverage before treating as a true gap. "
        f"Original rationale: {row.get('why_gap') or ''}"
    ).strip()
    supplement["pm_action"] = (
        f"Determine whether this adjacent idea belongs in {category.run_name}; only promote if fresh source evidence supports it. "
        f"{row.get('pm_action') or ''}"
    ).strip()
    return supplement


def _amazon_supplement(row: dict[str, Any], category: Category) -> dict[str, Any]:
    supplement = deepcopy(row)
    source_category = row.get("subcategory") or "Unknown"
    original_confidence = row.get("confidence") or "Unknown"
    original_priority = row.get("priority") or "Unknown"
    supplement["_supplemental"] = True
    supplement["_source_category"] = source_category
    supplement["subcategory"] = category.run_name
    supplement["recommendation"] = f"[Supplemental candidate] {row.get('recommendation')}"
    supplement["priority"] = "Exploratory"
    supplement["confidence"] = "Needs category fit review"
    supplement["classification"] = "Supplemental adjacent candidate - not exact category proof"
    supplement["evidence"] = (
        f"SUPPLEMENTAL WARNING: original source category was {source_category}; "
        f"original priority/confidence was {original_priority}/{original_confidence}. "
        f"{row.get('evidence') or ''}"
    ).strip()
    supplement["sunco_check"] = (
        f"SUPPLEMENTAL WARNING: this is not yet proven as an exact {category.run_name} Amazon gap. "
        f"It was included as the +1 adjacent candidate and needs category fit review before Step 2. "
        f"{row.get('sunco_check') or ''}"
    ).strip()
    supplement["action"] = (
        f"Determine whether this adjacent Amazon idea belongs in {category.run_name}; only promote if fresh source evidence supports it. "
        f"{row.get('action') or ''}"
    ).strip()
    return supplement


def _amazon_to_main_tab_row(row: dict[str, Any], category: Category) -> dict[str, Any]:
    display = deepcopy(row)
    display["_amazon_main_tab_display"] = True
    display["subcategory"] = category.run_name
    display["recommendation"] = row.get("recommendation")
    display["example"] = row.get("example")
    display["source_url"] = row.get("review_url")
    display["why_gap"] = (
        "Amazon/Stackline-derived display row shown because no separate Sunco.com/ecommerce "
        f"recommendation rows were available for {category.run_name}. "
        f"Evidence: {row.get('evidence') or ''}"
    ).strip()
    display["pm_action"] = row.get("action")
    display["source_systems"] = row.get("source_systems") or ["Amazon/Stackline evidence"]
    return display


def _backfill_rows(
    *,
    category: Category,
    exact_rows: list[dict[str, Any]],
    all_rows: list[dict[str, Any]],
    row_factory,
    supplemental_limit: int = SUPPLEMENTAL_CANDIDATES_PER_TAB,
) -> tuple[list[dict[str, Any]], list[str]]:
    if supplemental_limit <= 0:
        return exact_rows, []

    exact_keys = {_row_identity(row) for row in exact_rows}
    allowed_related = set(RELATED_CATEGORY_BACKFILL.get(category.slug, []))
    if not allowed_related:
        return list(exact_rows), []

    candidates = [
        row for row in all_rows
        if _row_identity(row)
        and _row_identity(row) not in exact_keys
        and not _match_category(category, row.get("subcategory"))
        and _category_slug(row.get("subcategory")) in allowed_related
        and _backfill_allowed(category, row)
    ]
    candidates.sort(key=lambda row: (_related_rank(category, row), *_priority_score(row), _row_identity(row)))

    output = list(exact_rows)
    warnings: list[str] = []
    for row in candidates:
        if len(warnings) >= supplemental_limit:
            break
        output.append(row_factory(row, category))
        warnings.append(
            f"Added supplemental candidate from {row.get('subcategory') or 'Unknown'}: {row.get('recommendation')}"
        )
    return output, warnings


def _reference_manifest_paths(paths: ProjectPaths) -> tuple[Path, Path, Path]:
    reference = paths.source_data / "schema_references"
    return (
        reference / "source_manifest_indoor_residential_2026-05-13.json",
        reference / "amazon_rerun_evidence_indoor_residential_2026-05-13.json",
        reference / "amazon_recommendations_manifest_indoor_residential_2026-05-13.json",
    )


def _date_from_text(value: Any) -> date | None:
    if not value:
        return None
    text = str(value)[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _date_from_json_or_mtime(path: Path, keys: tuple[str, ...] = ("generated_at", "generated", "created_at")) -> date | None:
    if not path or not path.exists():
        return None
    try:
        payload = load_json(path)
        for key in keys:
            generated = _date_from_text(payload.get(key))
            if generated:
                return generated
    except Exception:
        pass
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).date()
    except OSError:
        return None


def _latest_category_file(paths: ProjectPaths, category: Category, folder: Path, patterns: list[str]) -> Path | None:
    if not folder.exists():
        return None
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(folder.glob(pattern))
    matches = sorted(
        {path.resolve(): path for path in candidates if path.is_file()}.values(),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return matches[0] if matches else None


def _freshness_record(
    source: str,
    path: Path | None,
    generated: date | None,
    source_system: str,
    role: str,
    active: bool = True,
    note: str = "",
) -> dict[str, Any]:
    return {
        "source": source,
        "source_system": source_system,
        "role": role,
        "path": str(path) if path else "",
        "generated": generated.isoformat() if generated else "Unknown",
        "age_days": age_days(generated),
        "active": active,
        "note": note,
    }


def _latest_line_review_snapshot(paths: ProjectPaths, category: Category) -> Path | None:
    return _latest_category_file(
        paths,
        category,
        paths.source_data / "postgres_exports" / "line_reviews",
        [
            f"{category.slug}_line_review*.json",
            f"{category.slug}*line_review*.json",
            f"*{category.slug}*line_review*.json",
        ],
    ) or _latest_category_file(
        paths,
        category,
        paths.source_data / "redshift_exports" / "line_reviews",
        [
            f"{category.slug}_line_review*.json",
            f"{category.slug}*line_review*.json",
            f"*{category.slug}*line_review*.json",
        ],
    )


def build_step1_freshness_rows(
    paths: ProjectPaths,
    category: Category,
    generated: date | None,
    source_path: Path | None = None,
    amazon_rerun_path: Path | None = None,
    amazon_path: Path | None = None,
    include_legacy_active: bool = True,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if source_path:
        rows.append(
            _freshness_record(
                "Legacy ecommerce recommendation seed",
                source_path,
                generated or _date_from_json_or_mtime(source_path),
                "schema_reference_json",
                "Step 1 seed/reference",
                active=include_legacy_active,
                note="Older seed manifest. Step 1 should prefer refreshed Redshift/Postgres-backed sources when available.",
            )
        )
    if amazon_rerun_path:
        rows.append(
            _freshness_record(
                "Legacy Amazon rerun seed",
                amazon_rerun_path,
                _date_from_json_or_mtime(amazon_rerun_path) or generated,
                "schema_reference_json",
                "Amazon seed/reference",
                active=include_legacy_active,
                note="Older seed manifest. Kept for backward compatibility and audit context.",
            )
        )
    if amazon_path:
        rows.append(
            _freshness_record(
                "Legacy Amazon recommendation manifest",
                amazon_path,
                _date_from_json_or_mtime(amazon_path) or generated,
                "schema_reference_json",
                "Amazon fallback seed/reference",
                active=False,
                note="Fallback reference only unless no fresher Amazon evidence exists.",
            )
        )

    line_review_path = _latest_line_review_snapshot(paths, category)
    rows.append(
        _freshness_record(
            "Existing SKU line review snapshot",
            line_review_path,
            _date_from_json_or_mtime(line_review_path) if line_review_path else None,
            "postgres_mcp_or_redshift_snapshot",
            "Existing Sunco SKU coverage",
            active=bool(line_review_path),
            note="Refreshed by Step 0.",
        )
    )

    stackline_path = latest_category_cache(paths, category)
    rows.append(
        _freshness_record(
            "Redshift Stackline cache",
            stackline_path,
            cache_created_date(stackline_path) if stackline_path else None,
            "redshift_stackline_cache",
            "Amazon/Home Depot market demand",
            active=bool(stackline_path),
            note="Refreshed by Step 0 and checked by Step 3 preflight.",
        )
    )
    return rows


def summarize_freshness_rows(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return "No source freshness records were available."
    parts = []
    for row in rows:
        active = "active" if row.get("active") else "reference"
        age = row.get("age_days")
        age_text = "Unknown age" if age is None else f"{age} days old"
        parts.append(f"{row.get('source')}: {age_text} ({active}; {row.get('role')})")
    return "\n".join(parts)


def active_source_age_days(rows: list[dict[str, Any]]) -> int | None:
    ages = [row.get("age_days") for row in rows if row.get("active") and row.get("age_days") is not None]
    return max(ages) if ages else None


def apply_freshness_rows(data: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    data["freshness_rows"] = rows
    data["freshness_summary"] = summarize_freshness_rows(rows)
    data["active_source_age_days"] = active_source_age_days(rows)
    if data.get("active_source_age_days") is not None:
        data["age_days"] = data["active_source_age_days"]


def include_legacy_gap_evidence() -> bool:
    return os.environ.get(INCLUDE_LEGACY_GAP_EVIDENCE_ENV, "").strip().lower() in {"1", "true", "yes", "y"}


def load_category_data(paths: ProjectPaths, category: Category) -> dict[str, Any]:
    intelligence = load_category_intelligence(paths, category)
    source_path, amazon_rerun_path, amazon_path = _reference_manifest_paths(paths)
    legacy_enabled = include_legacy_gap_evidence()
    source_payload = load_json(source_path) if legacy_enabled and source_path.exists() else {"recommendations": []}
    amazon_rerun_payload = load_json(amazon_rerun_path) if legacy_enabled and amazon_rerun_path.exists() else {"recommendations": []}
    amazon_payload = load_json(amazon_path) if legacy_enabled and amazon_path.exists() else {"recommendations": []}

    all_source_rows = list(source_payload.get("recommendations", []))
    all_amazon_rows = list(amazon_rerun_payload.get("recommendations", []))
    fallback_amazon_rows = list(amazon_payload.get("recommendations", []))

    exact_source_rows = [
        row for row in source_payload.get("recommendations", [])
        if _match_category(category, row.get("subcategory"))
    ]
    exact_amazon_rows = [
        row for row in amazon_rerun_payload.get("recommendations", [])
        if _match_category(category, row.get("subcategory"))
    ]
    if not exact_amazon_rows:
        exact_amazon_rows = [
            row for row in amazon_payload.get("recommendations", [])
            if _match_category(category, row.get("subcategory"))
        ]

    if not exact_source_rows or not exact_amazon_rows:
        for evidence in intelligence.gap_evidence:
            row = _gap_evidence_to_step1_row(evidence, category)
            channel = str(evidence.get("source_channel") or "").lower()
            if "amazon" in channel and not exact_amazon_rows:
                exact_amazon_rows.append(row)
            elif "amazon" not in channel and not exact_source_rows:
                exact_source_rows.append(row)

    source_rows, source_warnings = _backfill_rows(
        category=category,
        exact_rows=exact_source_rows,
        all_rows=all_source_rows,
        row_factory=_source_supplement,
    )
    amazon_backfill_pool = all_amazon_rows or fallback_amazon_rows
    amazon_rows, amazon_warnings = _backfill_rows(
        category=category,
        exact_rows=exact_amazon_rows,
        all_rows=amazon_backfill_pool,
        row_factory=_amazon_supplement,
    )
    source_from_amazon_count = 0
    if not source_rows and exact_amazon_rows:
        source_rows = [_amazon_to_main_tab_row(row, category) for row in exact_amazon_rows]
        source_from_amazon_count = len(source_rows)

    generated = parse_generated_date(source_payload) or parse_generated_date(amazon_payload)
    data = {
        "source_path": str(source_path),
        "amazon_rerun_path": str(amazon_rerun_path),
        "amazon_path": str(amazon_path),
        "generated": generated,
        "legacy_seed_age_days": age_days(generated),
        "age_days": age_days(generated),
        "source_rows": source_rows,
        "amazon_rows": amazon_rows,
        "source_exact_count": len(exact_source_rows),
        "source_supplemental_count": len([row for row in source_rows if row.get("_supplemental")]),
        "source_from_amazon_count": source_from_amazon_count,
        "source_supplemental_warnings": source_warnings,
        "amazon_exact_count": len(exact_amazon_rows),
        "amazon_supplemental_count": max(0, len(amazon_rows) - len(exact_amazon_rows)),
        "amazon_supplemental_warnings": amazon_warnings,
        "supplemental_warnings": _unique_warnings(source_warnings + amazon_warnings),
        "category_intelligence": intelligence,
    }
    apply_freshness_rows(
        data,
        build_step1_freshness_rows(
            paths=paths,
            category=category,
            generated=generated,
            source_path=source_path,
            amazon_rerun_path=amazon_rerun_path,
            amazon_path=amazon_path,
            include_legacy_active=legacy_enabled,
        ),
    )
    return data


def _gap_evidence_to_step1_row(evidence: dict[str, Any], category: Category) -> dict[str, Any]:
    recommendation = evidence.get("recommendation")
    return {
        "subcategory": category.run_name,
        "recommendation": recommendation,
        "classification": evidence.get("classification"),
        "priority": evidence.get("priority") or "Medium",
        "confidence": evidence.get("confidence") or "Medium",
        "example": evidence.get("competitor_example"),
        "source_url": evidence.get("review_url"),
        "review_url": evidence.get("review_url"),
        "sunco_check": evidence.get("sunco_coverage_check"),
        "why_gap": evidence.get("gap_rationale"),
        "evidence": evidence.get("gap_rationale"),
        "pm_action": evidence.get("pm_action"),
        "action": evidence.get("pm_action"),
        "source_systems": evidence.get("source_systems") or ["category_intelligence_db"],
        "local_image": evidence.get("local_image"),
    }


def _template_path(paths: ProjectPaths) -> Path:
    template = paths.templates / "True_Gap_Workbook_Template.xlsx"
    if template.exists():
        return template
    fallback = paths.source_data / "schema_references" / "Sunco_Indoor_Residential_True_Gaps_2026-05-13.xlsx"
    if fallback.exists():
        return fallback
    template.parent.mkdir(parents=True, exist_ok=True)
    create_minimal_workbook(template, TRUE_GAP_HEADERS)
    return template


def _clear_true_gap_workbook(workbook) -> None:
    for sheet_name, headers in TRUE_GAP_HEADERS.items():
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        ws = workbook[sheet_name]
        if sheet_name == "Summary":
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell, MergedCell):
                        cell.value = None
            continue
        if hasattr(ws, "_images"):
            ws._images = []
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col).value = header


def _write_summary(workbook, category: Category, data: dict[str, Any]) -> None:
    ws = workbook["Summary"]
    ws["A1"] = f"{category.run_name} True Gap Recommendations"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:H1")
    rows = [
        ("Category", category.run_name),
        ("Owner", category.owner),
        ("Generated", date.today().isoformat()),
        ("Primary competitor channel", "Amazon"),
        ("Secondary channel", "Home Depot Marketplace"),
        ("Main Recommendations tab rows", len(data["source_rows"])),
        ("Amazon recommendations", len(data["amazon_rows"])),
        ("Exact source recommendations", data["source_exact_count"]),
        ("Amazon-derived main-tab display rows", data.get("source_from_amazon_count", 0)),
        ("Supplemental source candidates", data["source_supplemental_count"]),
        ("Exact Amazon recommendations", data["amazon_exact_count"]),
        ("Supplemental Amazon candidates", data["amazon_supplemental_count"]),
        ("Active source age days", "Unknown" if data["age_days"] is None else str(data["age_days"])),
        ("Legacy seed age days", "Unknown" if data.get("legacy_seed_age_days") is None else str(data["legacy_seed_age_days"])),
        ("Source freshness detail", data.get("freshness_summary") or "No source freshness records were available."),
        ("Why these ideations were selected", SUCCESS_PROXY_TEXT),
        ("Decision tree", _decision_tree_text()),
        ("Supplemental candidate rule", f"Use the natural exact-category result count plus up to {SUPPLEMENTAL_CANDIDATES_PER_TAB} warning-labeled adjacent candidate per recommendation tab."),
        ("Category intelligence database", format_intelligence_audit(data["category_intelligence"])),
    ]
    line_review_context = data.get("line_review_context")
    if line_review_context:
        rows.append(("Existing SKU line review", line_review_context.summary()))
    for index, (label, value) in enumerate(rows, start=3):
        ws.cell(index, 1).value = label
        ws.cell(index, 2).value = value
        ws.cell(index, 1).font = Font(bold=True)
        ws.cell(index, 2).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 80


def _write_source_rows(paths: ProjectPaths, workbook, category: Category, rows: list[dict[str, Any]]) -> list[str]:
    ws = workbook["Recommendations"]
    image_status: list[str] = []
    for index, row in enumerate(rows, start=2):
        image_path = find_cached_image(paths, row.get("local_image"))
        status = add_image(ws, image_path, f"A{index}")
        image_status.append(f"{row.get('recommendation', 'row ' + str(index))}: {status}")
        ws.row_dimensions[index].height = 78
        values = [
            None,
            row.get("subcategory"),
            _pm_recommendation_name(category, row),
            row.get("recommendation"),
            row.get("priority"),
            row.get("confidence"),
            row.get("example"),
            row.get("source_url"),
            row.get("sunco_check"),
            row.get("why_gap"),
            row.get("pm_action"),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(index, col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row.get("source_url"):
            ws.cell(index, 8).hyperlink = row["source_url"]
            ws.cell(index, 8).style = "Hyperlink"
    return image_status


def _write_amazon_rows(paths: ProjectPaths, workbook, category: Category, rows: list[dict[str, Any]]) -> list[str]:
    ws = workbook["Amazon Recommendations"]
    image_status: list[str] = []
    for index, row in enumerate(rows, start=2):
        image_path = find_cached_image(paths, row.get("local_image"))
        status = add_image(ws, image_path, f"A{index}")
        image_status.append(f"{row.get('recommendation', 'row ' + str(index))}: {status}")
        ws.row_dimensions[index].height = 78
        values = [
            None,
            row.get("subcategory"),
            _pm_recommendation_name(category, row),
            row.get("recommendation"),
            row.get("priority"),
            row.get("confidence"),
            row.get("classification"),
            row.get("evidence"),
            row.get("sunco_check"),
            row.get("example"),
            row.get("review_url"),
            row.get("action"),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(index, col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row.get("review_url"):
            ws.cell(index, 11).hyperlink = row["review_url"]
            ws.cell(index, 11).style = "Hyperlink"
    return image_status


def _write_source_audit(workbook, category: Category, data: dict[str, Any]) -> None:
    intelligence = data["category_intelligence"]
    ws = workbook["Sources and Audit"]
    rows = [
        ("Decision tree", category.run_name, "Selection logic", _decision_tree_text(), ""),
        ("Success proxy", category.run_name, "Why these ideations over others", SUCCESS_PROXY_TEXT, ""),
        ("Cache", "All", "Source manifest", data["source_path"], data["source_path"]),
        ("Cache", "All", "Corrected Amazon rerun evidence", data["amazon_rerun_path"], data["amazon_rerun_path"]),
        ("Cache", "All", "Amazon manifest", data["amazon_path"], data["amazon_path"]),
        ("Freshness", "All", "Legacy seed generated date", str(data.get("generated") or "Unknown"), ""),
        ("Freshness", "All", "Active source age days", "Unknown" if data.get("age_days") is None else str(data.get("age_days")), ""),
        ("Freshness", "All", "Source freshness detail", data.get("freshness_summary") or "No source freshness records were available.", ""),
        ("Category intelligence", category.run_name, "Backend SQLite database", format_intelligence_audit(intelligence), str(intelligence.database_path)),
    ]
    for freshness in data.get("freshness_rows", []):
        rows.append((
            "Freshness",
            category.run_name,
            freshness.get("source"),
            (
                f"{freshness.get('generated')} | age_days={freshness.get('age_days')} | "
                f"{'active' if freshness.get('active') else 'reference'} | {freshness.get('note')}"
            ),
            freshness.get("path"),
        ))
    if data.get("source_from_amazon_count"):
        rows.append((
            "Display note",
            category.run_name,
            "Amazon-derived main-tab rows",
            f"{data['source_from_amazon_count']} exact Amazon/Stackline row(s) were mirrored into the main Recommendations tab because no separate Sunco.com/ecommerce rows were available.",
            data["amazon_path"],
        ))
    line_review_context = data.get("line_review_context")
    if line_review_context:
        rows.extend(line_review_source_audit_rows(line_review_context))
    for row in rows:
        ws.append(row)
    for row in data.get("source_rows", []):
        ws.append([
            "Row decision",
            row.get("subcategory"),
            row.get("recommendation"),
            _row_decision_rationale(row, "Sunco.com/ecommerce"),
            row.get("source_url"),
        ])
    for warning in data.get("source_supplemental_warnings", []):
        ws.append(("Supplemental warning", "All", "Natural count +1 candidate", warning, ""))

    aws = workbook["Amazon Source Audit"]
    aws.append([
        "Amazon",
        "Amazon remains the primary competitor channel.",
        "Competitive recommendation ranking and image selection.",
        "Rows are only present where cached or refreshed evidence exists.",
        data["amazon_rerun_path"],
    ])
    aws.append([
        "Category intelligence",
        format_intelligence_audit(intelligence),
        "Backend demand, coverage, and feature-signal context used before Step 1 selection.",
        "Private source rows are only represented by counts and audit notes in user-facing workbooks.",
        str(intelligence.database_path),
    ])
    aws.append([
        "Decision tree",
        _decision_tree_text(),
        "Explains why Step 1 selected these ideations over other possible ideas.",
        "This is a selection framework, not a forecast guarantee.",
        "",
    ])
    aws.append([
        "Success proxy",
        SUCCESS_PROXY_TEXT,
        "Use this answer when asked why these ideas are worth researching first.",
        "Step 2 and Step 3 refine attributes, links, pricing, and launch fit.",
        "",
    ])
    for row in data.get("amazon_rows", []):
        aws.append([
            "Row decision",
            _row_decision_rationale(row, "Amazon"),
            row.get("recommendation"),
            "Use this row-level logic to explain the selection; supplemental rows remain exploratory.",
            row.get("review_url"),
        ])
    for warning in data.get("amazon_supplemental_warnings", []):
        aws.append([
            "Supplemental warning",
            warning,
            "Natural count +1 candidate",
            "Included as one adjacent idea beyond the exact category evidence.",
            "",
        ])

    for audit_ws in [ws, aws]:
        for row in audit_ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row_index in range(2, audit_ws.max_row + 1):
            audit_ws.row_dimensions[row_index].height = 60


def generate_gap_workbook(paths: ProjectPaths, category: Category, force_refresh: bool = False) -> tuple[Path, list[str]]:
    paths.ensure()
    with _category_run_lock(paths, category):
        data = load_category_data(paths, category)
        line_review_context = prepare_line_review_context(paths, category)
        data["line_review_context"] = line_review_context
        template = _template_path(paths)
        output_folder = paths.gap_category_outputs(category.slug)
        output_folder.mkdir(parents=True, exist_ok=True)
        output = output_folder / f"{category.slug}_true_gaps_{timestamp()}.xlsx"
        ensure_template_copy(template, output)

        workbook = load_workbook(output)
        _clear_true_gap_workbook(workbook)
        image_status = []
        _write_summary(workbook, category, data)
        image_status.extend(_write_source_rows(paths, workbook, category, data["source_rows"]))
        image_status.extend(_write_amazon_rows(paths, workbook, category, data["amazon_rows"]))
        _write_source_audit(workbook, category, data)
        write_line_review_sheet(workbook, line_review_context)

        freshness = "Unknown"
        if data["age_days"] is not None:
            freshness = f"{data['age_days']} days old"
        refresh_note = "Force refresh requested." if force_refresh else "Force refresh not requested."
        if data["age_days"] is None or data["age_days"] > 30 or force_refresh:
            refresh_note += " Active source freshness is outside the 30-day target or a force refresh was requested; run Step 0 before final decisions."
        if not data["source_rows"] and not data["amazon_rows"]:
            refresh_note += " No cached rows were available for this category, so this workbook is an audit-ready empty run shell."

        audit_rows = [
            ("Project", "sunco-product-opportunity-engine"),
            ("Category", category.run_name),
            ("Owner", category.owner),
            ("Generated", timestamp()),
            ("Active source freshness", freshness),
            ("Legacy seed age days", "Unknown" if data.get("legacy_seed_age_days") is None else str(data["legacy_seed_age_days"])),
            ("Source freshness detail", data.get("freshness_summary") or "No source freshness records were available."),
            ("Refresh note", refresh_note),
            ("Primary competitor channel", "Amazon"),
            ("Secondary competitor channel", "Home Depot Marketplace"),
            ("Ideation decision tree", _decision_tree_text()),
            ("Why selected over other ideas", SUCCESS_PROXY_TEXT),
            ("Ranking rule", "Exact selected-category rows are selected first. Rows with High priority and High confidence outrank weaker rows. If exact-category coverage is sparse, the tool may include one clearly warning-labeled adjacent candidate per tab for exploration."),
            ("Promotion rule", "A supplemental row is not considered a true gap until live evidence confirms category fit, demand, Sunco coverage gap, and PM actionability."),
            ("Image status", "\n".join(image_status) if image_status else "No images embedded for this category run."),
            ("Supplemental candidate rule", f"Use natural exact-category results plus up to {SUPPLEMENTAL_CANDIDATES_PER_TAB} adjacent supplemental candidate per tab. Supplemental rows are adjacent seeds, not exact category proof."),
            ("Amazon-derived main-tab display rows", str(data.get("source_from_amazon_count", 0))),
            ("Supplemental warnings", "\n".join(data.get("supplemental_warnings", [])) or "No supplemental rows were needed."),
            ("Category intelligence audit", format_intelligence_audit(data["category_intelligence"])),
        ]
        audit_rows.extend(line_review_run_audit_rows(line_review_context))
        add_or_replace_audit_sheet(
            workbook,
            audit_rows,
            collect_sql_text(paths, category.slug),
        )
        workbook.save(output)
        workbook.close()

        issues = validate_workbook(output, ["Summary", "Recommendations", "Sources and Audit", "Amazon Recommendations", "Amazon Source Audit", "Run Audit"])
        ok, message = try_excel_com_open_save(output)
        if message:
            issues.append(message if not ok else message)
        return output, issues
