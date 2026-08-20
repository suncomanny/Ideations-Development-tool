from __future__ import annotations

import json
import re
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.views import Pane, Selection

from .categories import Category
from .category_intelligence import format_intelligence_audit, load_category_intelligence
from .paths import ProjectPaths
from .sql_audit import collect_sql_text
from .url_validation import extract_urls, format_validation_notes, validate_urls
from .utils import newest_file, slugify, timestamp
from .validation import try_excel_com_open_save, validate_workbook
from .workbook_common import add_or_replace_audit_sheet, clear_row_values, ensure_template_copy


IDEATION_MAX_COL = 54
SHIPPING_INCLUSIVE_PRE_ADS_GM_TARGET = 0.525
SHIPPING_INCLUSIVE_PRE_ADS_GM_RANGE = (0.50, 0.55)


COLUMN_MAP = {
    "category": 1,
    "subcategory": 2,
    "ideation_name": 3,
    "sunco_reference_sku": 4,
    "reference_sku_source": 5,
    "strategy": 6,
    "voltage": 7,
    "wattage_primary": 8,
    "wattage_max": 9,
    "selectable_wattage": 10,
    "cct_primary": 11,
    "cct_max": 12,
    "selectable_cct": 13,
    "cri": 14,
    "lumens_target": 15,
    "efficiency": 16,
    "power_factor": 17,
    "dimmable": 18,
    "dimming_type": 19,
    "frequency": 20,
    "driver_type": 21,
    "size_form_factor": 22,
    "mounting_type": 23,
    "material": 24,
    "finish_color": 25,
    "ip_rating": 26,
    "moisture_rating": 27,
    "indoor_outdoor_use": 28,
    "operating_temperature": 29,
    "wiring_type": 30,
    "emergency_battery": 31,
    "run_time": 32,
    "charge_time": 33,
    "switching_time": 34,
    "motion_sensor": 35,
    "motion_duration": 36,
    "daylight_sensor_auto_dimming": 37,
    "smart_connected": 38,
    "linkable": 39,
    "bulb_base_type": 40,
    "bulb_shape": 41,
    "beam_angle": 42,
    "target_msrp": 43,
    "target_margin_shopify": 44,
    "target_margin_amazon": 45,
    "cost_type": 46,
    "target_vendor_cost": 47,
    "certifications": 48,
    "lifetime_hours": 49,
    "warranty": 50,
    "known_competitors": 51,
    "priority_channels": 52,
    "stackline_data": 53,
    "research_notes": 54,
}


PRD_SPEC_COLUMNS = set(COLUMN_MAP) - {"research_notes"}

PM_SKIP_MARKERS = (
    "[skip]",
    "[do not convert]",
    "[do not move forward]",
    "[hold]",
)
PM_SKIP_VALUES = {
    "skip",
    "do not convert",
    "do not move forward",
    "hold",
}

SOURCE_MAPPING_HEADERS = [
    "Ideation Name",
    "Why selected",
    "Key evidence used",
    "Reference SKU rationale",
    "Source URL",
    "URL status",
]

ACTION_NPD = "NPD"
ACTION_REVISION = "Revision"
ACTION_CONCEPT_REVIEW = "Concept Review"
ACTION_HOLD = "Hold"
ACTION_SORT_ORDER = {
    ACTION_NPD: 0,
    ACTION_REVISION: 1,
    ACTION_CONCEPT_REVIEW: 2,
    ACTION_HOLD: 3,
}

REFERENCE_SKU_STOPWORDS = {
    "active",
    "amazon",
    "case",
    "channel",
    "check",
    "coverage",
    "example",
    "feature",
    "listing",
    "product",
    "reference",
    "shopify",
    "source",
    "sku",
    "sunco",
    "title",
}


def _clean_prd_requirement_text(key: str, value: Any) -> Any:
    if key not in PRD_SPEC_COLUMNS or not isinstance(value, str):
        return value
    replacements = {
        "; validate against Step 1 listing and supplier file": "",
        "; validate listing claim": "",
        "; validate exact listing claim": "",
        "; validate against supplier warranty and Sunco category standard": "",
        "; validate battery warranty separately": "; separate battery coverage recommended",
        "TBD from Step 1 evidence and supplier options": "Target from Step 1 evidence",
        "TBD from Step 1 evidence and supplier file": "Target from Step 1 evidence",
        "TBD from supplier file": "target",
        "TBD by supplier": "Target",
        "supplier options": "design options",
        "supplier research": "product definition",
        "supplier and live market refresh": "market evidence",
        "supplier and certification file": "certification requirement",
        "supplier quotes": "quote refresh",
        "Validate against live pricing before launch.": "Use live pricing before launch.",
        "validate demand before adding complexity": "use demand evidence before adding complexity",
        "validate total fixture output": "define total fixture output",
        "validate if integrated LED": "use integrated LED requirement when applicable",
        "validate by design size and design options": "select by design size",
        "validate by design size and supplier options": "select by design size",
    }
    text = value
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+target from Step 1 evidence\b", "", text, flags=re.IGNORECASE)
    if text.strip().lower() == "target from step 1 evidence":
        return None
    if key == "voltage":
        text = re.sub(r"\b(\d{2,3}\s*-\s*\d{2,3})(?!\s*[Vv])\b", r"\1V", text)
    return re.sub(r"\s{2,}", " ", text).strip()


def _clean_research_note_text(value: str) -> str:
    replacements = {
        "validate final": "confirm final",
        "validate category fit": "confirm category fit",
        "validate battery capacity": "confirm battery capacity",
        "validate against supplier files": "confirm against source documents",
        "supplier files": "source documents",
        "supplier file": "source document",
        "before PRD lock": "before RFQ release",
        "Filled fields to validate": "Assumption-backed fields",
        "must be validated before Step 2": "needs confirmation before Step 2",
    }
    text = value
    for old, new in replacements.items():
        text = re.sub(re.escape(old), new, text, flags=re.IGNORECASE)
    text = re.sub(r"\bvalidated\b", "confirmed", text, flags=re.IGNORECASE)
    text = re.sub(r"\bvalidate\b", "confirm", text, flags=re.IGNORECASE)
    return text


def _template_path(paths: ProjectPaths) -> Path:
    template = paths.templates / "PRD_Research_Ideation_Template.xlsx"
    if template.exists():
        return template
    raise FileNotFoundError(f"Missing PRD ideation template: {template}")


def _attribute_profile_path() -> Path:
    return Path(__file__).with_name("category_attribute_profiles.json")


def _load_attribute_profiles() -> dict[str, Any]:
    profile_path = _attribute_profile_path()
    if not profile_path.exists():
        return {}
    return json.loads(profile_path.read_text(encoding="utf-8"))


def _merged_attribute_profile(category: Category, intelligence_defaults: dict[str, Any] | None = None) -> dict[str, Any]:
    profiles = _load_attribute_profiles()
    selected = profiles.get(category.slug) or {}
    inherited_key = selected.get("inherits")
    merged: dict[str, Any] = {}
    if inherited_key and inherited_key in profiles:
        merged.update(profiles[inherited_key])
        merged["defaults"] = dict(profiles[inherited_key].get("defaults") or {})
    merged.update({key: value for key, value in selected.items() if key != "defaults"})
    defaults = dict(merged.get("defaults") or {})
    defaults.update(selected.get("defaults") or {})
    defaults.update({key: value for key, value in (intelligence_defaults or {}).items() if value not in (None, "", [], {})})
    merged["defaults"] = defaults
    return merged


def _load_market_price_samples(paths: ProjectPaths, category: Category) -> tuple[list[dict[str, Any]], Path | None]:
    folder = paths.source_data / "market_price_samples"
    sample_file = newest_file(folder, [f"{category.slug}_market_price_samples_*.json"])
    if sample_file is None:
        return [], None
    payload = json.loads(sample_file.read_text(encoding="utf-8"))
    return list(payload.get("samples") or []), sample_file


def latest_gap_workbook(paths: ProjectPaths, category: Category) -> Path | None:
    return newest_file(
        paths.gap_category_outputs(category.slug),
        [f"{category.slug}_true_gaps_*.xlsx"],
    ) or newest_file(paths.gap_outputs, [f"{category.slug}_true_gaps_*.xlsx"])


def _clear_ideation_template(workbook) -> None:
    ws = workbook["Ideations"]
    ws["B1"] = None
    ws["E1"] = None
    for row in range(4, 104):
        clear_row_values(ws, row, IDEATION_MAX_COL)
    if "Source Mapping" in workbook.sheetnames:
        mapping = workbook["Source Mapping"]
        if mapping.max_row > 1:
            mapping.delete_rows(2, mapping.max_row - 1)


def _reset_sheet_views(workbook) -> None:
    """Keep Excel pane metadata valid after editing a copied template."""
    ws = workbook["Ideations"]
    ws.freeze_panes = None
    ws.sheet_view.pane = Pane(
        xSplit=3,
        ySplit=2,
        topLeftCell="D6",
        activePane="bottomRight",
        state="frozen",
    )
    ws.sheet_view.selection = [
        Selection(pane="topRight", activeCell="D1", sqref="D1"),
        Selection(pane="bottomLeft", activeCell="A6", sqref="A6"),
        Selection(pane="bottomRight", activeCell="D6", sqref="D6"),
    ]

    if "Source Mapping" in workbook.sheetnames:
        mapping = workbook["Source Mapping"]
        mapping.freeze_panes = None
        mapping.sheet_view.pane = Pane(
            ySplit=1,
            topLeftCell="A2",
            activePane="bottomLeft",
            state="frozen",
        )
        mapping.sheet_view.selection = [Selection(pane="bottomLeft", activeCell="A2", sqref="A2")]


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _profile_attribute_mode(profile: dict[str, Any]) -> str:
    mode = str(profile.get("attribute_mode") or "").strip().lower()
    if mode:
        return mode
    if profile.get("inherits") == "decorative_fixture":
        return "decorative_fixture"
    return "integrated_led_fixture"


def _profile_tokens(profile: dict[str, Any], key: str) -> set[str]:
    values = profile.get(key) or []
    if isinstance(values, str):
        values = [values]
    return {slugify(str(value)) for value in values if str(value or "").strip()}


def _profile_allows(profile: dict[str, Any], key: str, *terms: str) -> bool:
    tokens = _profile_tokens(profile, key)
    if not tokens:
        return True
    return any(slugify(term) in tokens for term in terms)


def _profile_sku_split_attributes(profile: dict[str, Any]) -> list[str]:
    values = profile.get("sku_split_attributes") or profile.get("sku_defining_attributes") or []
    if isinstance(values, str):
        values = [values]
    return [slugify(str(value)) for value in values if str(value or "").strip()]


def _record_has_values(record: dict[str, Any]) -> bool:
    return any(_cell_text(value) for value in record.values())


def _classification_from_recommendation(value: Any, fallback: str) -> str:
    text = _cell_text(value)
    lowered = text.lower()
    for label in [
        ACTION_NPD,
        ACTION_REVISION,
        ACTION_CONCEPT_REVIEW,
        ACTION_HOLD,
    ]:
        if lowered.startswith(label.lower()) or f"{label.lower()}:" in lowered:
            return label
    for label in [
        "Existing Sunco coverage, but missing feature",
        "Product Revision or merchandising review",
        "Strategic outlier / High-output watchlist",
        "Strategic outlier watchlist",
        "Possible feature gap",
        "New variant opportunity",
    ]:
        if lowered.startswith(label.lower()) or label.lower() in lowered:
            return "Strategic outlier / High-output watchlist" if label == "Strategic outlier watchlist" else label
    return fallback


def _pm_skip_reason(record: dict[str, Any], recommendation_key: str) -> str | None:
    recommendation = _cell_text(record.get(recommendation_key))
    lowered_recommendation = recommendation.lower()
    for marker in PM_SKIP_MARKERS:
        if lowered_recommendation.startswith(marker):
            return f"PM skip marker in {recommendation_key}: {marker}"
    for key in ["Priority", "Confidence", "PM Action", "PM action"]:
        value = _cell_text(record.get(key)).lower()
        if value in PM_SKIP_VALUES:
            return f"PM skip value in {key}: {record.get(key)}"
    return None


def _gap_row_skip_reason(record: dict[str, Any], recommendation_key: str, evidence_keys: list[str]) -> str | None:
    if not _cell_text(record.get(recommendation_key)):
        return "missing recommendation"
    skip_reason = _pm_skip_reason(record, recommendation_key)
    if skip_reason:
        return skip_reason
    if not any(_cell_text(record.get(key)) for key in evidence_keys):
        return "missing usable evidence, source link, or PM action"
    return None


def _format_skipped_rows(skipped_rows: list[str]) -> str:
    if not skipped_rows:
        return "None"
    shown = skipped_rows[:20]
    suffix = f"\n... {len(skipped_rows) - len(shown)} more skipped row(s)" if len(skipped_rows) > len(shown) else ""
    return "\n".join(shown) + suffix


def _clip_text(value: Any, max_chars: int = 1400) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 4].rstrip() + " ..."


def _source_url_status(link_validation: Any) -> str:
    text = str(link_validation or "").strip()
    if not text:
        return "No URL status available"
    if " -> " in text:
        return text.split(" -> ", 1)[1].strip()
    return text


def _is_pack_count_phrase(value: Any) -> bool:
    text = _cell_text(value).lower()
    if not text:
        return False
    return bool(re.fullmatch(r"(?:multi[-\s]?pack|\d{1,3}\s*[-\s]?(?:pack|pk|pc|piece|pieces))", text))


def _is_pack_count_only_recommendation(value: Any) -> bool:
    text = _cell_text(value)
    if not re.search(r"\bmulti[-\s]?pack\b", text, flags=re.IGNORECASE):
        return False
    tail = re.split(r"\s+-\s+", text)[-1]
    cues = [part.strip() for part in tail.split(",") if part.strip()]
    return bool(cues) and all(_is_pack_count_phrase(cue) for cue in cues)


def _clean_pack_count_recommendation(value: Any) -> Any:
    if value is None:
        return value
    text = str(value)
    text = re.sub(r"\s*,\s*multi[-\s]?pack\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bmulti[-\s]?pack\b\s*,?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+-\s*$", "", text).strip()
    text = re.sub(r"\s{2,}", " ", text)
    return text


def _clean_pack_count_cue_sentences(value: Any) -> Any:
    if value is None:
        return value
    text = str(value)

    def replace_cues(match: re.Match[str]) -> str:
        cues = [cue.strip() for cue in match.group(1).split(",") if cue.strip()]
        cues = [cue for cue in cues if not _is_pack_count_phrase(cue)]
        if not cues:
            return ""
        return f"Industrial-design cues: {', '.join(cues)}."

    text = re.sub(r"Industrial-design cues:\s*([^.\n]+)\.", replace_cues, text, flags=re.IGNORECASE)
    replacements = {
        "design/pack position": "design/spec position",
        "exact-design and pack-size coverage": "exact-design and spec coverage",
        "pack strategy, ": "",
        "pack/price architecture": "spec/price architecture",
        "pack visibility": "spec visibility",
    }
    for old, new in replacements.items():
        text = re.sub(re.escape(old), new, text, flags=re.IGNORECASE)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def _read_gap_rows_with_audit(gap_workbook: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(gap_workbook, data_only=False)
    rows: list[dict[str, Any]] = []
    exact_source_names: set[str] = set()
    line_review_candidates = _line_review_reference_candidates(workbook)
    line_review_reference = max(line_review_candidates, key=lambda item: (item["active_score"], item["total_revenue"])) if line_review_candidates else None
    line_review_sku_lookup = _line_review_sku_lookup(line_review_candidates)
    audit: dict[str, Any] = {
        "recommendations_rows_read": 0,
        "recommendations_rows_selected": 0,
        "recommendations_rows_skipped": 0,
        "amazon_rows_read": 0,
        "amazon_rows_selected": 0,
        "amazon_rows_skipped": 0,
        "skipped_rows": [],
    }

    if "Recommendations" in workbook.sheetnames:
        ws = workbook["Recommendations"]
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        for row_index in range(2, ws.max_row + 1):
            record = {str(headers[col - 1]): ws.cell(row_index, col).value for col in range(1, len(headers) + 1)}
            if not _record_has_values(record):
                continue
            audit["recommendations_rows_read"] += 1
            recommendation = record.get("Recommendation")
            skip_reason = _gap_row_skip_reason(
                record,
                "Recommendation",
                ["Why This Is A True Gap", "Competitor / Ecommerce Example", "Review Link", "Sunco Active-Catalog Check", "PM Action"],
            )
            if skip_reason:
                audit["recommendations_rows_skipped"] += 1
                audit["skipped_rows"].append(f"Recommendations row {row_index}: {skip_reason}")
                continue
            recommendation = _clean_pack_count_recommendation(recommendation)
            why_text = str(record.get("Why This Is A True Gap") or "").lower()
            source_label = "Amazon" if "amazon/stackline-derived display row" in why_text else "Sunco.com/ecommerce"
            exact_source_names.add(str(recommendation).strip().lower())
            audit["recommendations_rows_selected"] += 1
            pm_recommendation_name = record.get("PM Recommendation Name")
            rows.append({
                "source": source_label,
                "classification": _classification_from_recommendation(pm_recommendation_name or recommendation, "Step 1 opportunity"),
                "subcategory": record.get("Subcategory"),
                "pm_recommendation_name": pm_recommendation_name,
                "name": recommendation,
                "priority": record.get("Priority"),
                "confidence": record.get("Confidence"),
                "example": record.get("Competitor / Ecommerce Example"),
                "url": record.get("Review Link"),
                "sunco_check": _clean_pack_count_cue_sentences(record.get("Sunco Active-Catalog Check")),
                "why": _clean_pack_count_cue_sentences(record.get("Why This Is A True Gap")),
                "action": _clean_pack_count_cue_sentences(record.get("PM Action")),
                "_line_review_reference": line_review_reference,
                "_line_review_sku_lookup": line_review_sku_lookup,
            })

    if "Amazon Recommendations" in workbook.sheetnames:
        ws = workbook["Amazon Recommendations"]
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        for row_index in range(2, ws.max_row + 1):
            record = {str(headers[col - 1]): ws.cell(row_index, col).value for col in range(1, len(headers) + 1)}
            if not _record_has_values(record):
                continue
            audit["amazon_rows_read"] += 1
            recommendation = record.get("Amazon-channel recommendation")
            skip_reason = _gap_row_skip_reason(
                record,
                "Amazon-channel recommendation",
                ["Stackline / Amazon evidence", "Example listing", "Review link", "Sunco Amazon coverage check", "PM action"],
            )
            if skip_reason:
                audit["amazon_rows_skipped"] += 1
                audit["skipped_rows"].append(f"Amazon Recommendations row {row_index}: {skip_reason}")
                continue
            if _is_pack_count_only_recommendation(recommendation):
                audit["amazon_rows_skipped"] += 1
                audit["skipped_rows"].append(f"Amazon Recommendations row {row_index}: pack-count-only Amazon cue belongs in the separate pack-size workflow")
                continue
            recommendation = _clean_pack_count_recommendation(recommendation)
            is_supporting_evidence = str(recommendation or "").strip().lower() in exact_source_names
            pm_recommendation_name = record.get("PM Recommendation Name")
            classification = (
                _classification_from_recommendation(pm_recommendation_name, "Amazon recommendation")
                if pm_recommendation_name
                else record.get("Amazon classification") or _classification_from_recommendation(recommendation, "Amazon recommendation")
            )
            if is_supporting_evidence and "supporting" not in str(classification).lower():
                classification = f"{classification} + Step 1 supporting evidence"
            audit["amazon_rows_selected"] += 1
            rows.append({
                "source": "Amazon",
                "classification": classification,
                "subcategory": record.get("Subcategory"),
                "pm_recommendation_name": pm_recommendation_name,
                "name": recommendation,
                "priority": record.get("Priority"),
                "confidence": record.get("Confidence"),
                "example": record.get("Example listing"),
                "url": record.get("Review link"),
                "sunco_check": _clean_pack_count_cue_sentences(record.get("Sunco Amazon coverage check")),
                "why": _clean_pack_count_cue_sentences(record.get("Stackline / Amazon evidence")),
                "action": _clean_pack_count_cue_sentences(record.get("PM action")),
                "_line_review_reference": line_review_reference,
                "_line_review_sku_lookup": line_review_sku_lookup,
            })

    workbook.close()
    deduped_rows = _dedupe_gap_rows(rows)
    audit["rows_after_dedupe"] = len(deduped_rows)
    return deduped_rows, audit


def _read_gap_rows(gap_workbook: Path) -> list[dict[str, Any]]:
    rows, _audit = _read_gap_rows_with_audit(gap_workbook)
    return rows


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]+", "", str(value))
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _line_review_reference_candidates(workbook) -> list[dict[str, Any]]:
    if "Existing SKU Line Review" not in workbook.sheetnames:
        return []
    ws = workbook["Existing SKU Line Review"]
    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    candidates: list[dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        record = {
            headers[col - 1]: ws.cell(row_index, col).value
            for col in range(1, len(headers) + 1)
            if headers[col - 1]
        }
        sku = str(record.get("Family Part Number") or "").strip()
        if not sku or sku.startswith("("):
            continue
        total_revenue = _to_float(record.get("Total Revenue")) or 0.0
        status = str(record.get("Active Status") or "").lower()
        active_score = 1 if "active" in status else 0
        candidates.append({
            "sku": sku,
            "title": record.get("Product Title"),
            "vendor": record.get("Vendor"),
            "vendor_cost": _to_float(record.get("Vendor Cost")),
            "vendor_cost_source": record.get("Vendor Cost Source"),
            "amazon_revenue": _to_float(record.get("Amazon Revenue")) or 0.0,
            "shopify_revenue": _to_float(record.get("Shopify Revenue")) or 0.0,
            "total_revenue": total_revenue,
            "pack_sizes": record.get("Pack Sizes Available"),
            "product_url": record.get("Product URL"),
            "active_score": active_score,
        })
    return candidates


def _best_line_review_reference(workbook) -> dict[str, Any] | None:
    candidates = _line_review_reference_candidates(workbook)
    if not candidates:
        return None
    return max(candidates, key=lambda item: (item["active_score"], item["total_revenue"]))


def _normalize_sku_match_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def _strip_sku_pack_suffix(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"-(?:\d+)?(?:PK|PACK|PC|PCS)(?:[-.].*)?$", "", text, flags=re.IGNORECASE)


def _line_review_sku_lookup(candidates: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for candidate in candidates:
        sku = str(candidate.get("sku") or "").strip()
        if not sku:
            continue
        for key_source in {sku, _strip_sku_pack_suffix(sku)}:
            key = _normalize_sku_match_key(key_source)
            if key and key not in lookup:
                lookup[key] = candidate
    return lookup


def _canonical_line_review_reference(sku: str, lookup: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    if not lookup:
        return None
    for key_source in (sku, _strip_sku_pack_suffix(sku)):
        key = _normalize_sku_match_key(key_source)
        if key in lookup:
            return lookup[key]
    return None


def _looks_like_sunco_reference_sku(value: Any) -> bool:
    text = _cell_text(value)
    if not text:
        return False
    lowered = re.sub(r"[^a-z0-9]+", "", text.lower())
    if lowered in REFERENCE_SKU_STOPWORDS:
        return False
    if not re.fullmatch(r"[A-Z0-9][A-Z0-9_./-]{2,}", text, flags=re.IGNORECASE):
        return False
    compact = re.sub(r"[^A-Za-z0-9]", "", text)
    if len(compact) < 5 or not re.search(r"\d", compact):
        return False
    if re.fullmatch(r"B0[A-Z0-9]{8}", compact, flags=re.IGNORECASE):
        return False
    has_sku_shape = (
        any(delimiter in text for delimiter in ("-", "_", "."))
        or bool(re.search(r"[A-Z]{2,}\d|\d[A-Z]{2,}", compact, flags=re.IGNORECASE))
    )
    return has_sku_shape


def _append_unique(existing: str | None, new: str | None) -> str | None:
    if not new:
        return existing
    if not existing:
        return new
    existing_parts = [part.strip() for part in str(existing).split("\n") if part.strip()]
    if str(new).strip() in existing_parts:
        return existing
    return f"{existing}\n{new}"


def _append_unique_source(existing: str | None, new: str | None) -> str | None:
    if not new:
        return existing
    if not existing:
        return new
    parts = [part.strip() for part in str(existing).split(" + ") if part.strip()]
    if str(new).strip() not in parts:
        parts.append(str(new).strip())
    return " + ".join(parts)


def _dedupe_gap_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: dict[str, dict[str, Any]] = {}
    for item in rows:
        key = str(item.get("name") or "").strip().lower()
        if not key:
            continue
        if key not in deduped:
            deduped[key] = item.copy()
            continue
        existing = deduped[key]
        existing["source"] = _append_unique_source(existing.get("source"), item.get("source"))
        for field in ["classification", "priority", "confidence", "example", "url", "sunco_check", "why", "action"]:
            existing[field] = _append_unique(existing.get(field), item.get(field))
    return list(deduped.values())


def _first_match(text: str, pattern: str, flags: int = re.IGNORECASE) -> str | None:
    match = re.search(pattern, text, flags=flags)
    return match.group(1).strip() if match else None


def _extract_money_values(text: str | None) -> list[float]:
    values: list[float] = []
    for match in re.finditer(r"\$([0-9][0-9,]*(?:\.\d{1,2})?)", str(text or "")):
        value = _to_float(match.group(1))
        if value is not None:
            values.append(value)
    return values


def _competitor_label(item: dict[str, Any]) -> str | None:
    example = str(item.get("example") or "").strip()
    if not example:
        return None
    label = example.split(":", 1)[0].strip()
    return label[:90] if label else None


def _extract_demand_metrics(item: dict[str, Any]) -> dict[str, Any]:
    why = str(item.get("why") or "")
    metrics: dict[str, Any] = {}
    metrics["overlay_score"] = _first_match(why, r"(?:Demand confidence score|Product Demand overlay score):\s*([0-9.]+/100)")
    metrics["score_profile"] = _first_match(why, r"Score profile:\s*([^\n.]+)")
    metrics["observed_stock_decrease"] = _first_match(why, r"Observed stock decrease totals\s*([0-9,.]+)\s*units")
    metrics["decrease_events"] = _first_match(why, r"across\s*([0-9,]+)\s*decrease event")
    metrics["velocity_units_per_week"] = _first_match(why, r"averaging about\s*([0-9,.]+)\s*units/week")
    metrics["days_since_last_decrease"] = _first_match(why, r"Last decrease was\s*([^\n.]+?)\s*from the latest scrape")
    metrics["match_quality"] = _first_match(why, r"match quality\s*([0-9]+(?:\.[0-9]+)?)")
    metrics["stackline_sales"] = _first_match(why, r"Live Redshift Stackline Atlas found\s*(\$[0-9,]+)\s*retail sales")
    metrics["stackline_units"] = _first_match(why, r"retail sales and\s*([0-9,]+)\s*units")
    metrics["stackline_weeks"] = _first_match(why, r"across\s*([0-9,]+)\s*week")
    metrics["stackline_asp"] = _first_match(why, r"Average observed retail price\s*(\$[0-9][0-9,]*(?:\.\d{1,2})?)")
    metrics["stackline_segment"] = _first_match(why, r"Segment:\s*([^\n.]+)")
    metrics["searched_terms"] = _first_match(why, r"Searched/customer terms:\s*([^\n.]+)")
    return {key: value for key, value in metrics.items() if value not in (None, "")}


def _priority_channels_from_evidence(item: dict[str, Any], metrics: dict[str, Any]) -> str | None:
    profile = str(metrics.get("score_profile") or "").lower()
    source = str(item.get("source") or "").lower()
    if "amazon" in profile or "stackline" in profile or source == "amazon":
        return "Amazon first; Shopify parity review if Sunco.com assortment coverage is incomplete."
    if "shopify" in profile or "ecommerce" in profile or "ecommerce" in source:
        return "Shopify/front-end launch review first; Amazon follow-up if Stackline and marketplace economics support."
    return None


def _stackline_or_demand_summary(item: dict[str, Any], metrics: dict[str, Any]) -> str | None:
    if metrics.get("stackline_sales"):
        pieces = [
            f"Stackline/Amazon demand: {metrics.get('stackline_sales')} retail sales",
            f"{metrics.get('stackline_units')} units" if metrics.get("stackline_units") else None,
            f"{metrics.get('stackline_weeks')} weeks" if metrics.get("stackline_weeks") else None,
            f"ASP {metrics.get('stackline_asp')}" if metrics.get("stackline_asp") else None,
            f"segment {metrics.get('stackline_segment')}" if metrics.get("stackline_segment") else None,
            f"overlay score {metrics.get('overlay_score')}" if metrics.get("overlay_score") else None,
        ]
        return "; ".join(part for part in pieces if part) + "."
    if metrics.get("observed_stock_decrease") or metrics.get("velocity_units_per_week"):
        pieces = [
            f"Ecommerce movement: {metrics.get('observed_stock_decrease')} units observed stock decrease" if metrics.get("observed_stock_decrease") else "Ecommerce movement signal",
            f"{metrics.get('decrease_events')} decrease events" if metrics.get("decrease_events") else None,
            f"~{metrics.get('velocity_units_per_week')} units/week" if metrics.get("velocity_units_per_week") else None,
            f"last decrease {metrics.get('days_since_last_decrease')} from latest scrape" if metrics.get("days_since_last_decrease") else None,
            f"match quality {metrics.get('match_quality')}" if metrics.get("match_quality") else None,
            f"overlay score {metrics.get('overlay_score')}" if metrics.get("overlay_score") else None,
        ]
        return "; ".join(part for part in pieces if part) + "."
    return None


def _line_review_reference_text(reference: dict[str, Any]) -> str:
    parts = [
        "Best-selling adjacent active Sunco family by category revenue from Step 1 Existing SKU Line Review",
        str(reference.get("sku") or "").strip(),
    ]
    title = str(reference.get("title") or "").strip()
    if title:
        parts.append(title)
    revenue_bits = [
        f"total {_format_money(float(reference.get('total_revenue') or 0))}",
        f"Shopify {_format_money(float(reference.get('shopify_revenue') or 0))}",
        f"Amazon {_format_money(float(reference.get('amazon_revenue') or 0))}",
    ]
    parts.append(" / ".join(revenue_bits))
    if reference.get("vendor_cost") is not None:
        parts.append(f"vendor cost {_format_money(float(reference.get('vendor_cost')))}")
    return "; ".join(part for part in parts if part)


def _sunco_reference_from_coverage(item: dict[str, Any]) -> dict[str, str] | None:
    check = str(item.get("sunco_check") or "")
    if not check or "0 strong" in check.lower():
        return None
    match = re.search(
        r"(?:Partial Sunco active coverage found[^:]*:\s*|Sunco likely already has comparable active coverage[^:]*:\s*)?([A-Z][A-Z0-9_./-]{2,}):\s*([^\n.]+)",
        check,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    sku = match.group(1).strip()
    title = match.group(2).strip()
    if not _looks_like_sunco_reference_sku(sku):
        return None
    verified_reference = _canonical_line_review_reference(sku, item.get("_line_review_sku_lookup") or {})
    if item.get("_line_review_sku_lookup") and not verified_reference:
        return None
    if verified_reference:
        verified_sku = str(verified_reference.get("sku") or "").strip()
        verified_title = str(verified_reference.get("title") or title).strip()
        return {
            "sku": verified_sku,
            "source": f"Closest Sunco active-catalog coverage named in Step 1: {sku}; mapped to verified line-review reference {verified_sku}; {verified_title}",
        }
    return {
        "sku": sku,
        "source": f"Closest Sunco active-catalog coverage named in Step 1: {sku}; {title}",
    }


def _strategy_from_classification(value: str | None) -> str:
    text = (value or "").lower()
    if "hold" in text:
        return ACTION_CONCEPT_REVIEW
    if "concept review" in text or "strategic outlier" in text or "watchlist" in text or "possible feature gap" in text:
        return ACTION_CONCEPT_REVIEW
    if "revision" in text or "merchandising" in text or "existing sunco coverage" in text or "coverage exists" in text or "defend/optimize" in text or "sunco amazon incumbent" in text:
        return ACTION_REVISION
    if "npd" in text or "new product" in text or "new variant" in text or "true gap" in text or "amazon stackline opportunity" in text or "stackline/amazon leader" in text or "stackline amazon leader" in text:
        return ACTION_NPD
    return ACTION_CONCEPT_REVIEW


def _action_rank(value: Any) -> int:
    text = _cell_text(value).lower()
    if "npd" in text or "new product development" in text:
        return ACTION_SORT_ORDER[ACTION_NPD]
    if "revision" in text or "merchandising" in text or "existing sunco coverage" in text or "coverage exists" in text or "defend/optimize" in text or "sunco amazon incumbent" in text:
        return ACTION_SORT_ORDER[ACTION_REVISION]
    if "concept review" in text or "strategic outlier" in text or "watchlist" in text or "possible feature gap" in text or "partial coverage" in text:
        return ACTION_SORT_ORDER[ACTION_CONCEPT_REVIEW]
    if "hold" in text:
        return ACTION_SORT_ORDER[ACTION_HOLD]
    return 99


def _candidate_action_sort_key(item: dict[str, Any]) -> tuple[int, str]:
    strategy = _strategy_from_classification(item.get("classification") or item.get("name"))
    return (_action_rank(strategy), _cell_text(item.get("name")).lower())


def _strip_action_prefix(value: Any) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    prefixes = [
        ACTION_NPD,
        ACTION_REVISION,
        ACTION_CONCEPT_REVIEW,
        ACTION_HOLD,
        "Existing Sunco coverage, but missing feature",
        "Product Revision or merchandising review",
        "Strategic outlier / High-output watchlist",
        "Strategic outlier watchlist",
        "Possible feature gap",
        "New variant opportunity",
        "Amazon Stackline opportunity",
        "Stackline/Amazon leader",
        "Defend/optimize Sunco Amazon winner",
    ]
    for prefix in prefixes:
        pattern = rf"^\s*{re.escape(prefix)}(?:\s*\([^)]*\))?\s*:\s*"
        cleaned = re.sub(pattern, "", text, count=1, flags=re.IGNORECASE).strip()
        if cleaned != text:
            return cleaned
    return text


def _ideation_name_with_action(item: dict[str, Any], strategy: str) -> str:
    body = _strip_action_prefix(item.get("name"))
    if body:
        return f"{strategy}: {body}"
    return strategy


SOURCE_NAME_MARKERS = (
    "amazon stackline opportunity",
    "amazon-channel recommendation",
    "stackline/amazon leader",
    "stackline amazon leader",
    "amazon recommendation",
    "source audit",
    "step 1 evidence",
    "step 1 ecommerce",
    "competitor evidence",
    "home depot marketplace",
)

GENERIC_NAME_MARKERS = (
    "target unless step 1",
    "only if step 1",
    "named in step 1 evidence",
    "by selected product form factor",
    "by design size",
    "by kit type",
    "lumen output target",
    "cct target",
    "wattage target",
    "integrated led wattage target",
    "total fixture output defined",
    "defined by included bulb",
    "defined by product design",
)


def _title_name_component(value: str) -> str:
    acronyms = {
        "led": "LED",
        "cct": "CCT",
        "cri": "CRI",
        "ip": "IP",
        "ul": "UL",
        "etl": "ETL",
        "fcc": "FCC",
        "0-10v": "0-10V",
        "triac": "TRIAC",
        "ft": "ft",
        "in": "in",
        "lm": "lm",
    }
    words = re.split(r"(\s+|/|-)", value.strip())
    output: list[str] = []
    for word in words:
        key = word.lower()
        if not word or word.isspace() or word in {"/", "-"}:
            output.append(word)
        elif key in acronyms:
            output.append(acronyms[key])
        elif re.fullmatch(r"\d+(?:\.\d+)?(?:x\d+(?:\.\d+)?|ft|in|w|k|lm)?", key):
            output.append(word)
        elif re.fullmatch(r"\d+cct", key):
            output.append(word.upper())
        else:
            output.append(word[:1].upper() + word[1:].lower())
    text = "".join(output)
    text = re.sub(r"\bLed\b", "LED", text)
    text = re.sub(r"\bCct\b", "CCT", text)
    text = re.sub(r"\bIp(\d+)\b", r"IP\1", text)
    text = re.sub(r"\b0-10v\b", "0-10V", text, flags=re.IGNORECASE)
    return text


def _clean_name_component(value: Any, *, allow_generic: bool = False) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    lowered = text.lower()
    if any(marker in lowered for marker in SOURCE_NAME_MARKERS):
        return ""
    if not allow_generic and any(marker in lowered for marker in GENERIC_NAME_MARKERS):
        return ""
    if _is_pack_count_phrase(text) or _is_pack_count_only_recommendation(text):
        return ""
    text = _strip_action_prefix(text)
    text = _clean_pack_count_recommendation(text)
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\s*;\s*", ", ", text)
    text = re.sub(r"\btarget\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" ,;:-")
    if not text or text.lower() in {"yes", "no", "n/a", "none", "optional"}:
        return ""
    return text


def _step1_pm_recommendation_name(item: dict[str, Any], strategy: str) -> str:
    if item.get("_sku_candidate_target"):
        return ""
    text = _cell_text(item.get("pm_recommendation_name"))
    if not text:
        return ""
    lowered = text.lower()
    if any(marker in lowered for marker in SOURCE_NAME_MARKERS):
        return ""
    body = _strip_action_prefix(text)
    body = re.sub(r"\s+", " ", body).strip(" ,;:-")
    if not body:
        return ""
    return f"{strategy}: {body}"


def _dedupe_name_parts(parts: list[str]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for part in parts:
        normalized = re.sub(r"[^a-z0-9]+", "", part.lower())
        if not normalized or normalized in seen:
            continue
        if any(normalized in re.sub(r"[^a-z0-9]+", "", existing.lower()) for existing in output):
            continue
        seen.add(normalized)
        output.append(part)
    return output


def _sku_defining_wattage(value: Any) -> str:
    return _extract_wattage_text(_cell_text(value)) or ""


def _sku_defining_lumens(value: Any) -> str:
    return _extract_lumens_text(_cell_text(value)) or ""


def _sku_defining_cct(enriched: dict[str, Any]) -> str:
    primary = _clean_name_component(enriched.get("cct_primary"))
    if primary:
        return primary
    selectable = _cell_text(enriched.get("selectable_cct")).lower()
    if selectable.startswith("yes") or "selectable cct" in selectable or "switchable cct" in selectable:
        return "Selectable CCT"
    return ""


def _sku_defining_yes_feature(value: Any, label: str) -> str:
    text = _cell_text(value)
    lowered = text.lower()
    if not text or lowered in {"no", "n/a", "none"}:
        return ""
    if "only if" in lowered or "unless" in lowered or "optional" in lowered:
        return ""
    if lowered.startswith("yes") or label.lower() in lowered:
        return label
    return ""


def _sku_defining_dimming(value: Any) -> str:
    text = _cell_text(value)
    lowered = text.lower()
    if not text or "unless" in lowered:
        return ""
    if "0-10v" in lowered or "0-10 v" in lowered:
        return "0-10V Dimming"
    if "triac" in lowered:
        return "TRIAC Dimming"
    if "dimming" in lowered and not any(marker in lowered for marker in GENERIC_NAME_MARKERS):
        return _title_name_component(_clean_name_component(text))
    return ""


def _sku_defining_ip_rating(value: Any) -> str:
    text = _cell_text(value)
    match = re.search(r"\bIP\s*([0-9]{2})\b", text, flags=re.IGNORECASE)
    if match:
        return f"IP{match.group(1)}"
    return ""


def _sku_defining_lens_housing_features(text: str) -> list[str]:
    lowered = text.lower()
    features: list[str] = []
    for needle, label in [
        ("frosted lens", "Frosted Lens"),
        ("prismatic lens", "Prismatic Lens"),
        ("clear lens", "Clear Lens"),
        ("polycarbonate lens", "Polycarbonate Lens"),
        ("acrylic lens", "Acrylic Lens"),
        ("steel housing", "Steel Housing"),
        ("aluminum housing", "Aluminum Housing"),
        ("die-cast housing", "Die-Cast Housing"),
        ("vapor tight housing", "Vapor Tight Housing"),
        ("vaportight housing", "Vapor Tight Housing"),
        ("daisy chain", "Daisy-Chain Ready"),
        ("daisy-chain", "Daisy-Chain Ready"),
    ]:
        if needle in lowered and label not in features:
            features.append(label)
    return features


def _category_name_fallback(category: Category, profile: dict[str, Any]) -> str:
    mode = _profile_attribute_mode(profile)
    category_name = _title_name_component(category.run_name.replace("_", " "))
    if mode == "ceiling_fan":
        if "ceiling fan" in category_name.lower():
            return category_name
        return "Ceiling Fan"
    if mode == "bathroom_fan":
        return "Bathroom Exhaust Fan"
    if mode == "decorative_fixture":
        return f"{category_name} Fixture"
    if "fixture" in category_name.lower() or "light" in category_name.lower():
        return category_name
    return f"LED {category_name} Fixture"


def _sku_defining_ideation_name(
    category: Category,
    item: dict[str, Any],
    enriched: dict[str, Any],
    profile: dict[str, Any],
) -> str:
    strategy = enriched.get("strategy") if enriched.get("strategy") in {ACTION_NPD, ACTION_REVISION, ACTION_CONCEPT_REVIEW} else ACTION_CONCEPT_REVIEW
    step1_pm_name = _step1_pm_recommendation_name(item, strategy)
    if step1_pm_name:
        return step1_pm_name
    form_factor = _clean_name_component(enriched.get("size_form_factor"), allow_generic=True)
    if not form_factor or any(marker in form_factor.lower() for marker in SOURCE_NAME_MARKERS):
        form_factor = _category_name_fallback(category, profile)
    form_factor = _title_name_component(form_factor.replace(", LED", " LED").replace(", led", " LED"))

    parts = [
        form_factor,
        _sku_defining_wattage(enriched.get("wattage_primary")),
        _sku_defining_lumens(enriched.get("lumens_target")),
        _sku_defining_cct(enriched),
    ]
    features = [
        _sku_defining_yes_feature(enriched.get("selectable_wattage"), "Selectable Wattage"),
        _sku_defining_yes_feature(enriched.get("emergency_battery"), "Emergency Backup"),
        _sku_defining_yes_feature(enriched.get("motion_sensor"), "Motion Sensor"),
        _sku_defining_yes_feature(enriched.get("daylight_sensor_auto_dimming"), "Daylight Sensor"),
        _sku_defining_yes_feature(enriched.get("smart_connected"), "Smart Connected"),
        _sku_defining_yes_feature(enriched.get("linkable"), "Linkable"),
        _sku_defining_dimming(enriched.get("dimming_type")),
        _sku_defining_ip_rating(enriched.get("ip_rating")),
    ]
    features.extend(_sku_defining_lens_housing_features(_evidence_text(item)))
    if _profile_attribute_mode(profile) in {"ceiling_fan", "bathroom_fan"}:
        features.extend(_sku_defining_fan_features(_evidence_text(item), enriched, profile))
    clean_parts = _dedupe_name_parts([part for part in parts + features if part])
    body = ", ".join(clean_parts) or _category_name_fallback(category, profile)
    return f"{strategy}: {body}"


def _revision_change_summary(item: dict[str, Any], enriched: dict[str, Any]) -> str | None:
    if enriched.get("strategy") != ACTION_REVISION:
        return None
    target_sku = _cell_text(enriched.get("sunco_reference_sku"))
    target = target_sku if target_sku and not target_sku.lower().startswith("tbd") else "matched active Sunco SKU/family"
    action = _cell_text(item.get("action"))
    if action:
        if action.lower().startswith("revision target"):
            return action
        return f"Revise {target}: {action}"
    feature_text = _cell_text(item.get("name"))
    body = _strip_action_prefix(feature_text)
    if body:
        return f"Revise {target}: compare the active SKU/family against {body} and add the competitor-supported missing feature or merchandising update."
    return f"Revise {target}: use the Step 1 Sunco coverage note and source links to confirm the specific rolling-change or merchandising update."


def _evidence_text(item: dict[str, Any]) -> str:
    return "\n".join(
        str(item.get(key) or "")
        for key in ["name", "classification", "example", "why", "action", "sunco_check", "url"]
    )


def _unique_csv(parts: list[str]) -> str:
    output: list[str] = []
    for part in parts:
        text = part.strip()
        if text and text not in output:
            output.append(text)
    return ", ".join(output)


def _extract_bulb_bases(text: str) -> str | None:
    bases = re.findall(r"\bE(?:12|17|26|39)\b", text, flags=re.IGNORECASE)
    return _unique_csv([base.upper() for base in bases]) or None


def _extract_light_count(text: str) -> str | None:
    match = re.search(r"\b(\d+)\s*(?:-|to)\s*(\d+)\s*light\b", text, flags=re.IGNORECASE)
    if match:
        return f"{match.group(1)}-{match.group(2)} light"
    matches = re.findall(r"\b(\d+)\s*-?\s*light\b", text, flags=re.IGNORECASE)
    if matches:
        first = f"{matches[0]}-light"
        if len(matches) > 1:
            return f"{first}; optional {matches[1]}-light"
        return first
    return None


def _extract_light_count_values(text: str) -> list[str]:
    values: list[str] = []
    for match in re.finditer(r"\b(\d+)\s*(?:-|to)\s*(\d+)\s*light\b", text, flags=re.IGNORECASE):
        for value in [match.group(1), match.group(2)]:
            if value not in values:
                values.append(value)
    for value in re.findall(r"\b(\d+)\s*-?\s*light\b", text, flags=re.IGNORECASE):
        if value not in values:
            values.append(value)
    return values


def _extract_sizes(text: str) -> list[str]:
    return [
        re.sub(r"\s+", " ", match.group(0)).replace("inch", "in").strip()
        for match in re.finditer(r"\b\d+(?:\.\d+)?\s*(?:-|to)\s*\d+(?:\.\d+)?\s*(?:in|inch)\b", text, flags=re.IGNORECASE)
    ]


def _extract_panel_size_tokens(text: str) -> list[str]:
    values: list[str] = []
    for match in re.finditer(r"\b(?:1|2|4)\s*x\s*(?:2|4)\b", text, flags=re.IGNORECASE):
        value = re.sub(r"\s+", "", match.group(0)).lower()
        if value not in values:
            values.append(value)
    return values


def _extract_linear_size_tokens(text: str) -> list[str]:
    values: list[str] = []
    for match in re.finditer(r"\b(?:2|3|4|5|6|8)\s*(?:ft|foot|feet)\b", text, flags=re.IGNORECASE):
        number = re.search(r"\d+", match.group(0))
        if not number:
            continue
        value = f"{number.group(0)} ft"
        if value not in values:
            values.append(value)
    return values


def _extract_wattage_values(text: str) -> list[str]:
    values: list[str] = []
    for match in re.finditer(r"\b\d+(?:\.\d+)?\s*W\b", text, flags=re.IGNORECASE):
        value = match.group(0).replace(" ", "").upper()
        try:
            numeric = float(value.rstrip("W"))
        except ValueError:
            continue
        if numeric <= 0 or numeric > 2000:
            continue
        if value not in values:
            values.append(value)
    return values


def _extract_wattage_text(text: str) -> str | None:
    values = _extract_wattage_values(text)
    return " / ".join(values[:5]) if values else None


def _collect_lumen_values(text: str, pattern: str) -> list[str]:
    values: list[str] = []
    for match in re.finditer(pattern, text, flags=re.IGNORECASE):
        trailing = text[match.end(): match.end() + 8]
        if re.match(r"\s*/\s*w", trailing, flags=re.IGNORECASE):
            continue
        numeric = _to_float(match.group(1))
        if numeric is None or numeric < 300 or numeric > 200000:
            continue
        value = f"{numeric:,.0f}lm" if numeric.is_integer() else f"{numeric:,.1f}lm"
        if value not in values:
            values.append(value)
    return values


def _extract_lumen_values(text: str) -> list[str]:
    target_values = _collect_lumen_values(text, r"brightness target\s+([0-9][0-9,]*(?:\.\d+)?)\s*(?:lm|lumens)\b")
    if target_values:
        return target_values[:4]

    values: list[str] = []
    for value in _collect_lumen_values(text, r"\b([0-9][0-9,]*(?:\.\d+)?)\s*(?:lm|lumens)\b"):
        if value not in values:
            values.append(value)
    return values[:4]


def _extract_lumens_text(text: str) -> str | None:
    values = _extract_lumen_values(text)
    return " / ".join(values[:4]) if values else None


def _extract_cct_text(text: str) -> tuple[str | None, str | None]:
    lowered = text.lower()
    values: list[str] = []
    for match in re.finditer(r"\b(?:27|30|35|40|50|65)00K\b|\b\d{4}\s*K\b", text, flags=re.IGNORECASE):
        value = match.group(0).replace(" ", "").upper()
        if value not in values:
            values.append(value)
    if "5cct" in lowered:
        values = values or ["3000K", "3500K", "4000K", "5000K", "6500K"]
        return "5CCT selectable", "/".join(values[:5]) if values else "5CCT selectable"
    if "3cct" in lowered:
        values = values or ["3000K", "4000K", "5000K"]
        return "3CCT selectable", "/".join(values[:3]) if values else "3CCT selectable"
    if values:
        return "/".join(values[:4]), values[-1]
    if "selectable cct" in lowered or "switchable cct" in lowered:
        return "Selectable CCT", None
    return None, None


def _infer_integrated_mounting(text: str, default: str | None = None, profile: dict[str, Any] | None = None) -> str:
    profile = profile or {}
    lowered = text.lower()
    if "surface mount" in lowered or "surface-mount" in lowered:
        return "Surface mount"
    if "direct mount" in lowered or "direct-mount" in lowered:
        return "Direct ceiling mount"
    if "drop ceiling" in lowered or "lay in" in lowered or "lay-in" in lowered:
        return "Drop ceiling / lay-in grid"
    if ("troffer" in lowered or "center basket" in lowered) and _profile_allows(profile, "form_factor_terms", "troffer", "center basket"):
        return "Recessed grid / troffer"
    if ("flat panel" in lowered or "panel" in lowered) and _profile_allows(profile, "form_factor_terms", "flat panel", "panel"):
        return "Commercial ceiling panel mount"
    return default or "Mounting type by selected product form factor"


def _infer_integrated_form_factor(item: dict[str, Any], text: str, profile: dict[str, Any] | None = None) -> str:
    profile = profile or {}
    lowered = text.lower()
    pieces: list[str] = []
    sizes = _extract_panel_size_tokens(text)
    linear_sizes = _extract_linear_size_tokens(text)
    if sizes and _profile_allows(profile, "form_factor_terms", "panel", "flat panel"):
        pieces.append("/".join(sizes))
    if ("vapor tight" in lowered or "vaportight" in lowered or "vapor proof" in lowered) and _profile_allows(profile, "form_factor_terms", "vapor tight"):
        pieces.extend(linear_sizes[:2])
        pieces.append("LED vapor tight fixture")
    elif ("wraparound" in lowered or "wrap around" in lowered) and _profile_allows(profile, "form_factor_terms", "wraparound"):
        pieces.extend(linear_sizes[:2])
        pieces.append("LED wraparound fixture")
    elif ("strip light" in lowered or "striplight" in lowered) and _profile_allows(profile, "form_factor_terms", "strip light"):
        pieces.extend(linear_sizes[:2])
        pieces.append("LED strip light")
    elif "linear high bay" in lowered and _profile_allows(profile, "form_factor_terms", "linear high bay"):
        pieces.extend(linear_sizes[:2])
        pieces.append("linear high bay")
    elif "center basket" in lowered and _profile_allows(profile, "form_factor_terms", "center basket"):
        pieces.append("center basket troffer")
    elif "troffer" in lowered and _profile_allows(profile, "form_factor_terms", "troffer"):
        pieces.append("troffer")
    elif ("surface mount" in lowered or "surface-mount" in lowered) and _profile_allows(profile, "form_factor_terms", "surface mount panel"):
        pieces.append("surface-mount panel")
    elif "grid frame" in lowered and _profile_allows(profile, "form_factor_terms", "grid frame panel"):
        pieces.append("grid frame panel")
    elif "flat panel" in lowered and _profile_allows(profile, "form_factor_terms", "flat panel"):
        pieces.append("flat panel")
    elif "panel" in lowered and _profile_allows(profile, "form_factor_terms", "panel"):
        pieces.append("panel")
    if not pieces:
        pieces.append(str(item.get("name") or "").strip() or "Integrated LED fixture")
    return ", ".join(piece for piece in pieces if piece)


def _is_decorative_profile(profile: dict[str, Any]) -> bool:
    return _profile_attribute_mode(profile) == "decorative_fixture"


def _apply_integrated_led_enrichment(enriched: dict[str, Any], category: Category, item: dict[str, Any], text: str, profile: dict[str, Any]) -> None:
    cct_primary, cct_max = _extract_cct_text(text)
    lowered = text.lower()
    enriched["size_form_factor"] = _infer_integrated_form_factor(item, text, profile)
    enriched["mounting_type"] = _infer_integrated_mounting(text, enriched.get("mounting_type"), profile)
    enriched["material"] = enriched.get("material")
    enriched["finish_color"] = _infer_finish(text) or enriched.get("finish_color")
    enriched["bulb_base_type"] = None
    enriched["bulb_shape"] = None
    enriched["wattage_primary"] = _extract_wattage_text(text) or enriched.get("wattage_primary") or "Integrated LED wattage target"
    enriched["wattage_max"] = enriched.get("wattage_max") or enriched["wattage_primary"]
    enriched["cct_primary"] = cct_primary or enriched.get("cct_primary") or "CCT target"
    enriched["cct_max"] = cct_max or enriched.get("cct_max")
    enriched["lumens_target"] = _extract_lumens_text(text) or enriched.get("lumens_target") or "Lumen output target"
    enriched["efficiency"] = enriched.get("efficiency") or "Target lm/W derived from final wattage and lumen target"
    enriched["power_factor"] = enriched.get("power_factor") or ">=0.9 target for commercial integrated LED driver"
    enriched["operating_temperature"] = enriched.get("operating_temperature") or "Commercial indoor ambient operating range"
    if "triac" in lowered:
        enriched["dimming_type"] = "TRIAC dimming"
    elif "0-10v" in lowered or "0-10 v" in lowered:
        enriched["dimming_type"] = "0-10V dimming"
    if "selectable wattage" in lowered or "wattage selectable" in lowered or "multi-wattage" in lowered or "5-wattage" in lowered:
        enriched["selectable_wattage"] = "Yes"
    if "selectable cct" in lowered or "switchable cct" in lowered or "3cct" in lowered or "5cct" in lowered:
        enriched["selectable_cct"] = "Yes"
    if "emergency backup" in lowered or "emergency battery" in lowered:
        enriched["emergency_battery"] = "Yes - emergency backup named in Step 1 evidence"
    if "motion" in lowered or "sensor" in lowered:
        enriched["motion_sensor"] = "Yes - sensor/control feature named in Step 1 evidence"


def _extract_blade_span_values(text: str) -> list[str]:
    values: list[str] = []
    for match in re.finditer(
        r"\b([2-8][0-9](?:\.\d+)?)\s*(?:in|inch|inches|\")?\s*(?:ceiling\s*)?(?:fan|blade|span)\b|"
        r"\b(?:fan|blade|span)\s*(?:diameter|size|span)?\s*[:=]?\s*([2-8][0-9](?:\.\d+)?)\s*(?:in|inch|inches|\")\b",
        text,
        flags=re.IGNORECASE,
    ):
        raw = match.group(1) or match.group(2)
        if not raw:
            continue
        number = float(raw)
        value = f"{number:g} in"
        if value not in values:
            values.append(value)
    return values


def _extract_blade_count_values(text: str) -> list[str]:
    values: list[str] = []
    for match in re.finditer(r"\b([2-9])\s*(?:-| )?blade\b|\b([2-9])\s*blades\b", text, flags=re.IGNORECASE):
        raw = match.group(1) or match.group(2)
        if raw and raw not in values:
            values.append(raw)
    return values


def _extract_cfm_text(text: str) -> str | None:
    values: list[str] = []
    for match in re.finditer(r"\b([0-9][0-9,]*(?:\.\d+)?)\s*cfm\b", text, flags=re.IGNORECASE):
        numeric = _to_float(match.group(1))
        if numeric is None or numeric < 20 or numeric > 20000:
            continue
        value = f"{numeric:,.0f} CFM"
        if value not in values:
            values.append(value)
    return " / ".join(values[:4]) if values else None


def _fan_has_light_kit(text: str) -> bool:
    lowered = text.lower()
    if re.search(r"\b(?:no|without|not\s+including|does\s+not\s+include|do\s+not\s+include)\s+(?:a\s+)?(?:light|light\s+kit)\b", lowered):
        return False
    return any(
        phrase in lowered
        for phrase in [
            "light kit",
            "fan light",
            "with light",
            "with led",
            "integrated led light",
            "led light kit",
        ]
    )


def _infer_ceiling_fan_form_factor(item: dict[str, Any], text: str, category: Category) -> str:
    spans = _extract_blade_span_values(text)
    blade_counts = _extract_blade_count_values(text)
    parts = []
    if spans:
        parts.append("/".join(spans))
    if blade_counts:
        parts.append("/".join(f"{value}-blade" for value in blade_counts[:2]))
    category_label = "commercial ceiling fan" if "commercial" in category.slug else "ceiling fan"
    parts.append(category_label)
    if _fan_has_light_kit(text):
        parts.append("with light kit")
    elif "light kit adaptable" in text.lower() or "light-kit adaptable" in text.lower():
        parts.append("light-kit adaptable")
    return " ".join(part for part in parts if part)


def _infer_fan_mounting(text: str, default: str | None = None) -> str:
    lowered = text.lower()
    if "flush mount" in lowered or "low profile" in lowered or "hugger" in lowered:
        return "Flush / low-profile ceiling mount"
    if "downrod" in lowered or "down rod" in lowered:
        return "Downrod ceiling mount"
    if "dual mount" in lowered:
        return "Dual-mount ceiling fan"
    return default or "Ceiling fan-rated junction box mount"


def _fan_controls(text: str) -> tuple[str | None, str | None, str | None]:
    lowered = text.lower()
    remote = "Remote control" if "remote" in lowered else None
    smart = "Yes - smart fan controls named in Step 1 evidence" if "smart" in lowered or "wifi" in lowered or "wi-fi" in lowered else None
    reversible = "Reversible motor" if "reversible" in lowered else None
    return remote, smart, reversible


def _fan_motor_type(text: str, default: str | None = None) -> str:
    lowered = text.lower()
    if "dc motor" in lowered or re.search(r"\bdc\b", lowered):
        return "DC motor target"
    if "ac motor" in lowered or re.search(r"\bac\b", lowered):
        return "AC motor target"
    return default or "Fan motor target by blade span and price tier"


def _apply_ceiling_fan_enrichment(enriched: dict[str, Any], category: Category, item: dict[str, Any], text: str, profile: dict[str, Any]) -> None:
    lowered = text.lower()
    light_kit = _fan_has_light_kit(text)
    cct_primary, cct_max = _extract_cct_text(text)
    remote, smart, reversible = _fan_controls(text)
    enriched["size_form_factor"] = _infer_ceiling_fan_form_factor(item, text, category)
    enriched["mounting_type"] = _infer_fan_mounting(text, enriched.get("mounting_type"))
    enriched["finish_color"] = _infer_finish(text) or enriched.get("finish_color")
    enriched["material"] = _infer_material(text) or enriched.get("material")
    enriched["wattage_primary"] = _extract_wattage_text(text) or enriched.get("wattage_primary") or _fan_motor_type(text)
    enriched["wattage_max"] = enriched.get("wattage_max") or enriched["wattage_primary"]
    enriched["driver_type"] = _fan_motor_type(text, enriched.get("driver_type"))
    enriched["wiring_type"] = "Hardwired ceiling fan electrical box"
    enriched["power_factor"] = None
    enriched["efficiency"] = _extract_cfm_text(text) or "Airflow target by blade span and motor type"
    enriched["beam_angle"] = None
    enriched["linkable"] = "No"
    enriched["bulb_base_type"] = None
    enriched["bulb_shape"] = None
    if light_kit:
        enriched["cct_primary"] = cct_primary or enriched.get("cct_primary") or "Light-kit CCT target"
        enriched["cct_max"] = cct_max or enriched.get("cct_max")
        enriched["lumens_target"] = _extract_lumens_text(text) or enriched.get("lumens_target") or "Light-kit lumen target"
        enriched["dimmable"] = enriched.get("dimmable") or "Yes for light-kit variants"
        if cct_primary and ("selectable" in cct_primary.lower() or "cct" in cct_primary.lower()):
            enriched["selectable_cct"] = "Yes"
    else:
        enriched["cct_primary"] = None
        enriched["cct_max"] = None
        enriched["selectable_cct"] = "Light-kit only"
        enriched["lumens_target"] = None
        enriched["dimmable"] = "Light-kit only"
    if remote:
        enriched["dimming_type"] = remote
    if smart:
        enriched["smart_connected"] = smart
    if reversible:
        enriched["motion_sensor"] = reversible
    if "damp" in lowered:
        enriched["moisture_rating"] = "Damp rated"
    elif "wet" in lowered or "outdoor" in lowered:
        enriched["moisture_rating"] = "Wet rated"
        enriched["indoor_outdoor_use"] = "Indoor/outdoor where rated"
    else:
        enriched["moisture_rating"] = enriched.get("moisture_rating") or "Dry rated indoor use"
    enriched["operating_temperature"] = enriched.get("operating_temperature") or "Residential indoor ambient operating range"


def _infer_bathroom_fan_form_factor(text: str) -> str:
    cfm = _extract_cfm_text(text)
    pieces = [cfm, "bathroom exhaust fan"]
    if "humidity" in text.lower():
        pieces.append("humidity sensor")
    if _fan_has_light_kit(text):
        pieces.append("with light")
    return " ".join(part for part in pieces if part)


def _extract_sone_text(text: str) -> str | None:
    match = re.search(r"\b([0-9](?:\.\d+)?)\s*sone?s?\b", text, flags=re.IGNORECASE)
    return f"{match.group(1)} sones" if match else None


def _apply_bathroom_fan_enrichment(enriched: dict[str, Any], category: Category, item: dict[str, Any], text: str, profile: dict[str, Any]) -> None:
    light_kit = _fan_has_light_kit(text)
    cct_primary, cct_max = _extract_cct_text(text)
    enriched["size_form_factor"] = _infer_bathroom_fan_form_factor(text)
    enriched["mounting_type"] = "Ceiling exhaust fan mount"
    enriched["wiring_type"] = "Hardwired"
    enriched["driver_type"] = "Exhaust fan motor"
    enriched["wattage_primary"] = _extract_wattage_text(text) or enriched.get("wattage_primary") or "Fan motor wattage by airflow target"
    enriched["wattage_max"] = enriched.get("wattage_max") or enriched["wattage_primary"]
    enriched["efficiency"] = _extract_sone_text(text) or "Airflow and sound target"
    enriched["moisture_rating"] = "Damp rated bathroom use"
    enriched["indoor_outdoor_use"] = "Indoor bathroom"
    enriched["operating_temperature"] = "Residential indoor bathroom ambient range"
    enriched["linkable"] = "No"
    enriched["beam_angle"] = None
    enriched["bulb_base_type"] = None
    enriched["bulb_shape"] = None
    if light_kit:
        enriched["cct_primary"] = cct_primary or enriched.get("cct_primary") or "Light-kit CCT target"
        enriched["cct_max"] = cct_max or enriched.get("cct_max")
        enriched["lumens_target"] = _extract_lumens_text(text) or enriched.get("lumens_target") or "Light-kit lumen target"
        enriched["dimmable"] = enriched.get("dimmable") or "Yes for light-kit variants"
    else:
        enriched["cct_primary"] = None
        enriched["cct_max"] = None
        enriched["selectable_cct"] = "Light-kit only"
        enriched["lumens_target"] = None
        enriched["dimmable"] = "Light-kit only"
    if "humidity" in text.lower():
        enriched["motion_sensor"] = "Humidity sensor target"


def _sku_defining_fan_features(text: str, enriched: dict[str, Any], profile: dict[str, Any]) -> list[str]:
    lowered = text.lower()
    features: list[str] = []
    has_light_kit = _fan_has_light_kit(text)
    for needle, label in [
        ("dc motor", "DC Motor"),
        ("remote", "Remote Control"),
        ("reversible", "Reversible Motor"),
        ("light kit", "Light Kit"),
        ("with light", "Light Kit"),
        ("low profile", "Low Profile"),
        ("flush mount", "Flush Mount"),
        ("downrod", "Downrod Mount"),
        ("damp", "Damp Rated"),
        ("wet", "Wet Rated"),
        ("humidity", "Humidity Sensor"),
    ]:
        if label == "Light Kit" and not has_light_kit:
            continue
        if needle in lowered and label not in features:
            features.append(label)
    if _profile_attribute_mode(profile) == "bathroom_fan":
        cfm = _extract_cfm_text(text)
        if cfm:
            features.insert(0, cfm)
    return features


def _infer_finish(text: str) -> str | None:
    lowered = text.lower()
    if "black/gold" in lowered or ("black" in lowered and "gold" in lowered):
        return "Matte black / gold or brass accent target"
    if "matte black" in lowered and "brass" in lowered:
        return "Matte black / brass"
    if "black/wood" in lowered or ("black" in lowered and "wood" in lowered):
        return "Black with wood-look accent target"
    if "rattan" in lowered or "woven" in lowered:
        return "Natural rattan / woven shade target"
    if "black" in lowered:
        return "Black finish target"
    return None


def _infer_material(text: str) -> str | None:
    lowered = text.lower()
    if "rattan" in lowered or "woven" in lowered:
        return "Metal frame with natural rattan/woven shade target"
    if "wood" in lowered:
        return "Metal frame with wood-look accents target"
    if any(term in lowered for term in ["chandelier", "linear", "wagon wheel", "black", "brass", "gold"]):
        return "Metal fixture body target"
    return None


def _infer_mounting(text: str) -> str:
    lowered = text.lower()
    if re.search(r"\brods?\b|\brod[- ]hung\b", lowered) or "linear" in lowered or "island" in lowered:
        return "Ceiling / rod-hung"
    if "wagon wheel" in lowered:
        return "Ceiling / chain-hanging"
    return "Ceiling / chain-hanging or rod-hung"


def _infer_form_factor_label(text: str) -> str:
    lowered = text.lower()
    if "wagon wheel" in lowered:
        return "wagon wheel chandelier"
    if "linear" in lowered or "island" in lowered:
        return "linear island chandelier"
    if "rattan" in lowered or "woven" in lowered:
        return "rattan/woven pendant"
    if "chandelier" in lowered:
        return "chandelier"
    return "fixture"


def _infer_size_form_factor(item: dict[str, Any], text: str) -> str:
    name = str(item.get("name") or "").strip()
    lowered = text.lower()
    pieces: list[str] = []
    light_count = _extract_light_count(text)
    sizes = _extract_sizes(text)
    if light_count:
        pieces.append(light_count)
    if "wagon wheel" in lowered:
        pieces.append("wagon wheel chandelier")
    elif "linear" in lowered or "island" in lowered:
        pieces.append("linear island chandelier")
    else:
        pieces.append(name)
    if sizes:
        pieces.append("; ".join(sizes) + " target")
    return ", ".join(piece for piece in pieces if piece)


def _sample_mentions_light_count(sample: dict[str, Any], light_count: str | None) -> bool:
    if not light_count:
        return False
    number = re.sub(r"\D+", "", str(light_count))
    if not number:
        return False
    haystack = " ".join(str(sample.get(key) or "") for key in ["product", "note", "url"]).lower()
    return bool(re.search(rf"\b{re.escape(number)}\s*[- ]?light\b", haystack, flags=re.IGNORECASE))


def _build_known_competitors(item: dict[str, Any]) -> str:
    parts = []
    for key in ["example", "url"]:
        value = item.get(key)
        if value:
            parts.append(str(value))
    return "\n".join(parts) or "Step 1 competitor evidence or maintained category competitor references"


def _normalize_match_text(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _market_samples_for_item(item: dict[str, Any], samples: list[dict[str, Any]]) -> list[dict[str, Any]]:
    name = _normalize_match_text(item.get("name"))
    matched: list[dict[str, Any]] = []
    for sample in samples:
        match = _normalize_match_text(sample.get("ideation_match"))
        if not match:
            continue
        if match in name or name in match:
            matched.append(sample)
    return matched


def _percentile(values: list[float], percentile: float) -> float:
    if not values:
        raise ValueError("No values provided.")
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * percentile
    lower = int(position)
    upper = min(lower + 1, len(ordered) - 1)
    weight = position - lower
    return ordered[lower] + (ordered[upper] - ordered[lower]) * weight


def _retail_price_point(value: float) -> float:
    rounded = int((value + 9.9999) // 10 * 10) - 0.01
    return max(0.99, round(rounded, 2))


def _format_money(value: float) -> str:
    return f"${value:,.2f}"


def _target_vendor_cost_from_margin(enriched: dict[str, Any], target_msrp: float) -> dict[str, Any] | None:
    margin_pct = SHIPPING_INCLUSIVE_PRE_ADS_GM_TARGET * 100
    floor_cost = target_msrp * (1 - SHIPPING_INCLUSIVE_PRE_ADS_GM_RANGE[1])
    ceiling_cost = target_msrp * (1 - SHIPPING_INCLUSIVE_PRE_ADS_GM_RANGE[0])
    target_cost = round(target_msrp * (1 - SHIPPING_INCLUSIVE_PRE_ADS_GM_TARGET), 2)
    return {
        "target": target_cost,
        "notes": (
            f"Target vendor cost basis: shipping-inclusive landed cost ceiling backsolved from target MSRP "
            f"{_format_money(target_msrp)} using the midpoint of the 50-55% pre-ads gross margin policy "
            f"({margin_pct:.1f}%). Acceptable landed-cost band for that policy is "
            f"{_format_money(floor_cost)}-{_format_money(ceiling_cost)}."
        ),
    }


def _numeric_pricing_cell(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return round(value, 2)
    prices = _extract_money_values(str(value))
    if prices:
        return round(prices[0], 2)
    stripped = str(value).strip().replace(",", "")
    try:
        return round(float(stripped), 2)
    except ValueError:
        return None


def _market_msrp_target(item: dict[str, Any], market_samples: list[dict[str, Any]]) -> dict[str, Any] | None:
    matched_samples = [
        sample for sample in _market_samples_for_item(item, market_samples)
        if isinstance(sample.get("price"), (int, float))
    ]
    target_light_count = str(item.get("_target_light_count") or "").strip()
    pricing_basis = "comparable listings"
    if target_light_count:
        light_count_samples = [sample for sample in matched_samples if _sample_mentions_light_count(sample, target_light_count)]
        pricing_basis = (
            f"exact {target_light_count}-light comparable listings"
            if light_count_samples
            else f"nearest parent-family comparable listings; no exact {target_light_count}-light sample was found in the local cache"
        )
        samples = light_count_samples or matched_samples
    else:
        samples = [sample for sample in matched_samples if sample.get("use_for_base_msrp", True) is not False]
    prices = [float(sample["price"]) for sample in samples]
    if not prices:
        return None

    p50 = _percentile(prices, 0.50)
    p55 = _percentile(prices, 0.55)
    target = _retail_price_point((p50 + p55) / 2)
    sample_labels = [
        f"{sample.get('retailer')} {sample.get('brand')}: {_format_money(float(sample['price']))}"
        for sample in samples
    ]
    sample_links = [
        f"{sample.get('retailer')} {sample.get('brand')}: {sample.get('url')}"
        for sample in samples
        if sample.get("url")
    ]
    return {
        "target": round(target, 2),
        "text": (
            f"{_format_money(target)} market MSRP target; based on p50-p55 of {pricing_basis} "
            f"({_format_money(p50)}-{_format_money(p55)}). Use live pricing before launch."
        ),
        "vendor_cost_text": (
            f"Backsolve from AQ MSRP target {_format_money(target)} after quote refresh; "
            "do not use margin targets as the initial MSRP anchor."
        ),
        "notes": f"Market MSRP pricing basis: {pricing_basis}. Samples: " + "; ".join(sample_labels),
        "links": "\n".join(sample_links),
    }


def _market_msrp_target_from_step1(item: dict[str, Any]) -> dict[str, Any] | None:
    prices = _extract_money_values(str(item.get("example") or ""))
    if not prices:
        return None
    observed = prices[-1]
    label = _competitor_label(item) or "Step 1 competitor listing"
    url = str(item.get("url") or "").strip()
    return {
        "target": round(observed, 2),
        "text": (
            f"{_format_money(observed)} provisional market MSRP anchor from Step 1 competitor listing; "
            "use broader p50-p55 market pricing when refreshed comps are available."
        ),
        "vendor_cost_text": (
            f"Backsolve from provisional market MSRP anchor {_format_money(observed)} after quote refresh; "
            "do not use margin targets as the initial MSRP anchor."
        ),
        "notes": f"Market MSRP fallback basis: {label} observed at {_format_money(observed)} in Step 1 evidence.",
        "links": f"{label}: {url}" if url else "",
    }


def _variant_size_form_factor(text: str, light_count: str | None, size: str | None) -> str:
    form_factor = _infer_form_factor_label(text)
    lowered = text.lower()
    if light_count == "1" and ("rattan" in lowered or "woven" in lowered):
        form_factor = "dome rattan/woven pendant"
    elif light_count == "3" and ("rattan" in lowered or "woven" in lowered):
        form_factor = "linear island rattan/woven pendant"
    pieces = []
    if light_count:
        pieces.append(f"{light_count}-light")
    pieces.append(form_factor)
    if size:
        pieces.append(size)
    return ", ".join(pieces)


def _variant_label(text: str, light_count: str | None, size: str | None) -> str:
    form_factor = _infer_form_factor_label(text)
    lowered = text.lower()
    if light_count == "1" and ("rattan" in lowered or "woven" in lowered):
        form_factor = "dome rattan/woven pendant"
    elif light_count == "3" and ("rattan" in lowered or "woven" in lowered):
        form_factor = "linear rattan/woven island pendant"
    finish = _infer_finish(text)
    parts: list[str] = []
    if light_count:
        parts.append(f"{light_count}-light")
    if size:
        parts.append(size)
    parts.append(form_factor)
    if finish:
        parts.append(finish.replace(" target", ""))
    return " ".join(parts)


def _candidate_name(base_name: str, label: str) -> str:
    supplemental_prefix = "[Supplemental candidate] "
    prefix = supplemental_prefix if base_name.startswith(supplemental_prefix) else ""
    clean_base = base_name.removeprefix(supplemental_prefix).strip()
    return f"{prefix}{clean_base} - {label}"


def _extract_mounting_values(text: str) -> list[str]:
    lowered = text.lower()
    values: list[str] = []
    for needle, label in [
        ("surface mount", "Surface mount"),
        ("surface-mount", "Surface mount"),
        ("direct mount", "Direct ceiling mount"),
        ("direct-mount", "Direct ceiling mount"),
        ("drop ceiling", "Drop ceiling / lay-in grid"),
        ("lay-in", "Drop ceiling / lay-in grid"),
        ("magnetic", "Magnetic mount"),
        ("adhesive", "Adhesive mount"),
        ("plug-in", "Plug-in"),
        ("hardwired", "Hardwired"),
        ("chain", "Chain-hung"),
        ("rod", "Rod-hung"),
    ]:
        if needle in lowered and label not in values:
            values.append(label)
    return values


def _extract_control_type_values(text: str) -> list[str]:
    lowered = text.lower()
    values: list[str] = []
    for needle, label in [
        ("0-10v", "0-10V dimming"),
        ("0-10 v", "0-10V dimming"),
        ("triac", "TRIAC dimming"),
        ("motion", "Motion sensor"),
        ("sensor", "Sensor control"),
        ("remote", "Remote control"),
        ("smart", "Smart connected"),
    ]:
        if needle in lowered and label not in values:
            values.append(label)
    return values


def _extract_finish_values(text: str) -> list[str]:
    lowered = text.lower()
    values: list[str] = []
    for needle, label in [
        ("matte black", "Matte black"),
        ("black", "Black"),
        ("white", "White"),
        ("bronze", "Bronze"),
        ("brass", "Brass"),
        ("gold", "Gold"),
        ("nickel", "Nickel"),
        ("rattan", "Natural rattan"),
        ("woven", "Woven shade"),
        ("wood", "Wood-look accent"),
    ]:
        if needle in lowered and label not in values:
            values.append(label)
    if "Matte black" in values and "Black" in values:
        values.remove("Black")
    return values


def _sku_attribute_values(attribute: str, text: str) -> list[str]:
    if attribute == "light_count":
        return _extract_light_count_values(text)
    if attribute == "blade_span":
        return _extract_blade_span_values(text)
    if attribute == "blade_count":
        return _extract_blade_count_values(text)
    if attribute in {"fixture_size", "size"}:
        return _extract_sizes(text)
    if attribute == "panel_size":
        return _extract_panel_size_tokens(text)
    if attribute == "linear_size":
        return _extract_linear_size_tokens(text)
    if attribute in {"output_tier", "lumens", "lumen_output"}:
        return _extract_lumen_values(text)
    if attribute in {"wattage", "wattage_tier"}:
        return _extract_wattage_values(text)
    if attribute in {"mounting", "mounting_type"}:
        return _extract_mounting_values(text)
    if attribute in {"control", "control_type"}:
        return _extract_control_type_values(text)
    if attribute in {"finish", "finish_color"}:
        return _extract_finish_values(text)
    return []


def _sku_attribute_label(attribute: str, value: str) -> str:
    if attribute == "light_count":
        return f"{value}-light" if value.isdigit() else value
    if attribute == "blade_count":
        return f"{value}-blade" if value.isdigit() else value
    return value


def _candidate_dimension_values(profile: dict[str, Any], text: str) -> list[dict[str, Any]]:
    dimensions = []
    for attribute in _profile_sku_split_attributes(profile):
        values = _sku_attribute_values(attribute, text)
        if values:
            dimensions.append({"attribute": attribute, "values": values})
    return dimensions


def _candidate_values_for_index(dimensions: list[dict[str, Any]], index: int) -> dict[str, str]:
    output: dict[str, str] = {}
    for dimension in dimensions:
        values = list(dimension.get("values") or [])
        if not values:
            continue
        value = values[index] if index < len(values) else values[0]
        output[str(dimension.get("attribute"))] = value
    return output


def _candidate_target_label(attribute_values: dict[str, str]) -> str:
    parts = [
        _sku_attribute_label(attribute, value)
        for attribute, value in attribute_values.items()
        if value
    ]
    return ", ".join(dict.fromkeys(parts))


def _candidate_field_overrides(
    category: Category,
    profile: dict[str, Any],
    text: str,
    attribute_values: dict[str, str],
) -> dict[str, Any]:
    overrides: dict[str, Any] = {}
    light_count = attribute_values.get("light_count")
    size = attribute_values.get("fixture_size") or attribute_values.get("size")
    if light_count or size:
        overrides["size_form_factor"] = _variant_size_form_factor(text, light_count, size)
    if light_count:
        overrides["wattage_primary"] = f"Replaceable LED bulb design; target {light_count}-light fixture"
    if attribute_values.get("panel_size"):
        overrides["size_form_factor"] = f"{attribute_values['panel_size']} integrated LED fixture"
    if attribute_values.get("linear_size"):
        form_factor = _infer_integrated_form_factor({"name": category.run_name}, text, profile)
        overrides["size_form_factor"] = f"{attribute_values['linear_size']} {form_factor}"
    if attribute_values.get("blade_span") or attribute_values.get("blade_count"):
        parts = [
            attribute_values.get("blade_span"),
            _sku_attribute_label("blade_count", attribute_values.get("blade_count") or ""),
            "ceiling fan" if _profile_attribute_mode(profile) == "ceiling_fan" else None,
        ]
        overrides["size_form_factor"] = " ".join(part for part in parts if part)
    if attribute_values.get("output_tier"):
        overrides["lumens_target"] = attribute_values["output_tier"]
    if attribute_values.get("lumens"):
        overrides["lumens_target"] = attribute_values["lumens"]
    if attribute_values.get("wattage") or attribute_values.get("wattage_tier"):
        overrides["wattage_primary"] = attribute_values.get("wattage") or attribute_values.get("wattage_tier")
    if attribute_values.get("mounting_type") or attribute_values.get("mounting"):
        overrides["mounting_type"] = attribute_values.get("mounting_type") or attribute_values.get("mounting")
    control_type = attribute_values.get("control_type") or attribute_values.get("control")
    if control_type:
        if "dimming" in control_type.lower():
            overrides["dimming_type"] = control_type
        elif "remote" in control_type.lower():
            overrides["dimming_type"] = control_type
        elif "sensor" in control_type.lower():
            overrides["motion_sensor"] = f"Yes - {control_type} named in Step 1 evidence"
        elif "smart" in control_type.lower():
            overrides["smart_connected"] = "Yes - smart control named in Step 1 evidence"
    if attribute_values.get("finish_color") or attribute_values.get("finish"):
        overrides["finish_color"] = attribute_values.get("finish_color") or attribute_values.get("finish")
    return overrides


def _sku_candidate_items(category: Category, profile: dict[str, Any], item: dict[str, Any]) -> list[dict[str, Any]]:
    text = _evidence_text(item)
    base_name = str(item.get("name") or "").strip()
    dimensions = _candidate_dimension_values(profile, text)
    if not dimensions:
        return [item]
    split_count = max((len(dimension["values"]) for dimension in dimensions), default=1)
    if split_count <= 1:
        return [item]

    candidates: list[dict[str, Any]] = []
    max_candidates = int(profile.get("max_sku_candidates_per_row") or 4)
    for index in range(min(split_count, max_candidates)):
        attribute_values = _candidate_values_for_index(dimensions, index)
        label = _candidate_target_label(attribute_values)
        if not label:
            continue
        candidate = deepcopy(item)
        candidate["name"] = _candidate_name(base_name, label)
        candidate["_parent_opportunity"] = base_name
        candidate["_target_light_count"] = attribute_values.get("light_count")
        candidate["_target_size"] = (
            attribute_values.get("fixture_size")
            or attribute_values.get("size")
            or attribute_values.get("panel_size")
            or attribute_values.get("linear_size")
        )
        candidate["_sku_candidate_target"] = label
        candidate["_sku_candidate_rationale"] = (
            "Split from the parent opportunity because the selected category profile allows this SKU-defining "
            "attribute and Step 1 evidence or PM action explicitly named multiple options. Step 3 should "
            "research this row as one candidate SKU, not as a broad family."
        )
        candidate["_field_overrides"] = _candidate_field_overrides(category, profile, text, attribute_values)
        candidates.append(candidate)

    return candidates or [item]


def _step2_candidate_rows(category: Category, profile: dict[str, Any], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidate_rows: list[dict[str, Any]] = []
    for item in rows:
        candidate_rows.extend(_sku_candidate_items(category, profile, item))
    candidate_rows.sort(key=_candidate_action_sort_key)
    return candidate_rows


def _url_validation_summary(cache: dict[str, dict[str, Any]]) -> str:
    if not cache:
        return "No review URLs found for review."
    counts: dict[str, int] = {}
    flagged: list[str] = []
    for result in cache.values():
        state = str(result.get("state") or "unverified")
        counts[state] = counts.get(state, 0) + 1
        if state != "verified":
            flagged.append(f"{result.get('url')} -> {result.get('note')}")
    count_text = ", ".join(f"{state}: {count}" for state, count in sorted(counts.items()))
    if flagged:
        return count_text + "\nFlagged links:\n" + "\n".join(flagged)
    return count_text


def _build_research_notes(
    item: dict[str, Any],
    enriched: dict[str, Any],
    profile: dict[str, Any],
    pricing: dict[str, Any] | None,
    link_validation: str | None,
) -> str:
    notes = [
        f"Source: {item.get('source')}",
        f"Recommended Product Action: {enriched.get('strategy') or item.get('classification')}",
        f"Revision target SKU: {enriched.get('sunco_reference_sku')}" if enriched.get("strategy") == ACTION_REVISION else None,
        f"Recommended revision changes: {_revision_change_summary(item, enriched)}" if enriched.get("strategy") == ACTION_REVISION else None,
        f"Demand summary: {enriched.get('stackline_data')}" if enriched.get("stackline_data") else None,
        f"Priority channel rationale: {enriched.get('priority_channels')}" if enriched.get("priority_channels") else None,
        f"Parent opportunity: {item.get('_parent_opportunity')}" if item.get("_parent_opportunity") else None,
        f"SKU candidate target: {item.get('_sku_candidate_target')}" if item.get("_sku_candidate_target") else None,
        f"SKU candidate rationale: {item.get('_sku_candidate_rationale')}" if item.get("_sku_candidate_rationale") else None,
        f"Evidence: {item.get('why')}",
        f"Sunco check: {item.get('sunco_check')}",
        f"Recommended action: {item.get('action')}",
        f"Review link: {item.get('url')}",
        f"Review link check: {link_validation}" if link_validation else None,
        "Attribute enrichment: fields were filled from Step 1 evidence first, then category-intelligence/cache defaults, then selected category-profile defaults when safe.",
        "Confidence note: final electrical specs, certifications, packaging, and claims should be source-confirmed before RFQ release.",
    ]
    if profile.get("notes"):
        notes.append(f"Profile note: {profile.get('notes')}")
    if pricing:
        notes.append(f"Target MSRP basis: {pricing['text']}")
        if pricing.get("vendor_cost_notes"):
            notes.append(pricing["vendor_cost_notes"])
        notes.append(pricing["notes"])
        if pricing.get("links"):
            notes.append("Market pricing links:\n" + pricing["links"])
    assumed = [
        key for key in [
            "voltage",
            "frequency",
            "dimmable",
            "dimming_type",
            "mounting_type",
            "wiring_type",
            "certifications",
        ]
        if enriched.get(key)
    ]
    if assumed:
        notes.append("Assumption-backed fields: " + ", ".join(assumed))
    return _clean_research_note_text("\n".join(part for part in notes if part and not str(part).endswith("None")))


def _enrich_item(
    category: Category,
    item: dict[str, Any],
    profile: dict[str, Any],
    market_samples: list[dict[str, Any]],
    url_validation_cache: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    defaults = dict(profile.get("defaults") or {})
    text = _evidence_text(item)
    bulb_base = _extract_bulb_bases(text)
    light_count = _extract_light_count(text)
    is_decorative = _is_decorative_profile(profile)
    attribute_mode = _profile_attribute_mode(profile)
    metrics = _extract_demand_metrics(item)
    pricing = _market_msrp_target(item, market_samples) or _market_msrp_target_from_step1(item)
    link_validation = format_validation_notes(validate_urls(extract_urls(item.get("url")), url_validation_cache))
    strategy = _strategy_from_classification(item.get("classification") or item.get("name"))
    enriched = {
        "category": category.run_name,
        "subcategory": defaults.get("subcategory") or category.run_name,
        "ideation_name": _ideation_name_with_action(item, strategy),
        "sunco_reference_sku": "TBD - use best-selling adjacent Sunco family by category revenue",
        "reference_sku_source": f"No exact Sunco {category.run_name} SKU identified from Step 1 evidence; use revenue proxy during refresh",
        "strategy": strategy,
        "known_competitors": _build_known_competitors(item),
    }
    enriched.update(defaults)
    coverage_reference = _sunco_reference_from_coverage(item)
    line_review_reference = item.get("_line_review_reference") or {}
    if coverage_reference:
        enriched["sunco_reference_sku"] = coverage_reference["sku"]
        enriched["reference_sku_source"] = coverage_reference["source"]
    elif line_review_reference.get("sku"):
        enriched["sunco_reference_sku"] = line_review_reference["sku"]
        enriched["reference_sku_source"] = _line_review_reference_text(line_review_reference)
    priority_channels = _priority_channels_from_evidence(item, metrics)
    if priority_channels:
        enriched["priority_channels"] = priority_channels
    demand_summary = _stackline_or_demand_summary(item, metrics)
    if demand_summary:
        enriched["stackline_data"] = demand_summary

    if is_decorative:
        enriched["size_form_factor"] = _infer_size_form_factor(item, text)
        enriched["mounting_type"] = _infer_mounting(text)
        enriched["material"] = _infer_material(text) or enriched.get("material")
        enriched["finish_color"] = _infer_finish(text) or enriched.get("finish_color")
        enriched["bulb_base_type"] = bulb_base or "E12/E26 target by design size"
        if light_count and bulb_base:
            normalized_count = light_count.replace("; optional", " / optional")
            enriched["wattage_primary"] = f"Replaceable LED bulb design; target {normalized_count} using {bulb_base} LED bulbs"
        elif light_count:
            enriched["wattage_primary"] = f"Replaceable LED bulb design; target {light_count}"
        else:
            enriched["wattage_primary"] = "Replaceable LED bulb design; bulb count defined by product design"
        enriched["wattage_max"] = "Socket max by certification requirement"
        enriched["cct_primary"] = "2700K-3000K warm white target if bulbs are included"
        enriched["cct_max"] = "3000K target; avoid selectable CCT unless integrated LED"
        enriched["lumens_target"] = "Total fixture output defined by included bulb bundle"
        enriched["efficiency"] = "Lamp efficiency based on selected LED bulb bundle"
        enriched["power_factor"] = "N/A for replaceable bulbs; integrated LED requirement when applicable"
        enriched["operating_temperature"] = "Residential indoor ambient operating range"
    elif attribute_mode == "under_cabinet_task_lighting":
        enriched["size_form_factor"] = _infer_integrated_form_factor(item, text, profile)
        enriched["mounting_type"] = _infer_integrated_mounting(text, enriched.get("mounting_type"), profile)
        enriched["material"] = enriched.get("material")
        enriched["finish_color"] = _infer_finish(text) or enriched.get("finish_color")
        enriched["bulb_base_type"] = None
        enriched["bulb_shape"] = None
    elif attribute_mode == "ceiling_fan":
        _apply_ceiling_fan_enrichment(enriched, category, item, text, profile)
    elif attribute_mode == "bathroom_fan":
        _apply_bathroom_fan_enrichment(enriched, category, item, text, profile)
    else:
        _apply_integrated_led_enrichment(enriched, category, item, text, profile)

    if attribute_mode == "under_cabinet_task_lighting":
        lowered = text.lower()
        if "magnetic" in lowered or "rechargeable" in lowered or "puck" in lowered:
            enriched["mounting_type"] = "Magnetic and adhesive under-cabinet mount target"
            enriched["wiring_type"] = "Rechargeable battery / no-wire installation"
            enriched["motion_sensor"] = "Yes for rechargeable puck/bar variants"
        elif "tape" in lowered or "strip" in lowered:
            enriched["mounting_type"] = "Adhesive tape/channel under-cabinet mount target"
            enriched["wiring_type"] = "Plug-in or hardwired low-voltage driver options"
            enriched["motion_sensor"] = "No unless selected as a motion-kit variant"
            enriched["linkable"] = "Yes; extension/corner accessories expected"
        enriched["size_form_factor"] = str(enriched.get("size_form_factor") or "").replace("Ceiling / chain-hanging or rod-hung", "under-cabinet kit")
        enriched["wattage_primary"] = "Integrated LED; wattage by puck/bar/tape length and kit format"
        enriched["lumens_target"] = "Task-light output by kit type and per-light output target"
        enriched["bulb_base_type"] = None
        enriched["bulb_shape"] = None
    if pricing:
        target_msrp = round(float(pricing["target"]), 2)
        enriched["target_msrp"] = target_msrp
        enriched["target_margin_shopify"] = SHIPPING_INCLUSIVE_PRE_ADS_GM_TARGET * 100
        enriched["target_margin_amazon"] = SHIPPING_INCLUSIVE_PRE_ADS_GM_TARGET * 100
        enriched["cost_type"] = "Landed"
        vendor_cost = _target_vendor_cost_from_margin(enriched, target_msrp)
        if vendor_cost:
            enriched["target_vendor_cost"] = vendor_cost["target"]
            pricing["vendor_cost_notes"] = vendor_cost["notes"]
    for key, value in dict(item.get("_field_overrides") or {}).items():
        if key in COLUMN_MAP and value:
            enriched[key] = value
    enriched["target_msrp"] = _numeric_pricing_cell(enriched.get("target_msrp"))
    enriched["target_vendor_cost"] = _numeric_pricing_cell(enriched.get("target_vendor_cost"))
    enriched["ideation_name"] = _sku_defining_ideation_name(category, item, enriched, profile)
    enriched["research_notes"] = _build_research_notes(item, enriched, profile, pricing, link_validation)
    enriched["_link_validation"] = link_validation
    return enriched


def _fill_ideation_row(
    ws,
    row_index: int,
    category: Category,
    item: dict[str, Any],
    market_samples: list[dict[str, Any]],
    url_validation_cache: dict[str, dict[str, Any]],
    profile: dict[str, Any] | None = None,
    intelligence_defaults: dict[str, Any] | None = None,
) -> dict[str, Any]:
    profile = profile or _merged_attribute_profile(category, intelligence_defaults)
    enriched = _enrich_item(category, item, profile, market_samples, url_validation_cache)
    values = {
        COLUMN_MAP[key]: _clean_prd_requirement_text(key, value)
        for key, value in enriched.items()
        if key in COLUMN_MAP
    }
    for col, value in values.items():
        cell = ws.cell(row_index, col)
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return enriched


def generate_prd_ideation_workbook(paths: ProjectPaths, category: Category, gap_workbook: Path | None = None) -> tuple[Path, list[str]]:
    paths.ensure()
    intelligence = load_category_intelligence(paths, category)
    selected_gap = gap_workbook or latest_gap_workbook(paths, category)
    if selected_gap is None:
        raise FileNotFoundError(f"No Step 1 gap workbook found for {category.run_name}. Run Step 1 first.")

    output_folder = paths.prd_ideation_category_outputs(category.slug)
    output_folder.mkdir(parents=True, exist_ok=True)
    output = output_folder / f"{category.slug}_prd_ideations_{timestamp()}.xlsx"
    ensure_template_copy(_template_path(paths), output)
    workbook = load_workbook(output)
    _clear_ideation_template(workbook)
    profile = _merged_attribute_profile(category, intelligence.attribute_defaults)

    ws = workbook["Ideations"]
    ws["B1"] = category.owner
    ws["E1"] = f"Prepared from {selected_gap.name} on {timestamp()}"
    ws.cell(3, COLUMN_MAP["strategy"]).value = "Recommended Product Action *"
    rows, intake_audit = _read_gap_rows_with_audit(selected_gap)
    candidate_rows = _step2_candidate_rows(category, profile, rows)
    market_samples, market_sample_path = _load_market_price_samples(paths, category)
    url_validation_cache: dict[str, dict[str, Any]] = {}
    enriched_rows: list[dict[str, Any]] = []
    for offset, item in enumerate(candidate_rows[:100], start=4):
        enriched_rows.append(_fill_ideation_row(ws, offset, category, item, market_samples, url_validation_cache, profile=profile))

    if "Source Mapping" in workbook.sheetnames:
        mapping = workbook["Source Mapping"]
        for col_index, header in enumerate(SOURCE_MAPPING_HEADERS, start=1):
            cell = mapping.cell(1, col_index)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for item, enriched in zip(candidate_rows[:100], enriched_rows):
            evidence_parts = [
                item.get("why"),
                f"Review link check: {enriched.get('_link_validation')}" if enriched.get("_link_validation") else None,
                f"SKU candidate: {item.get('_sku_candidate_rationale')}" if item.get("_sku_candidate_rationale") else None,
            ]
            source_url = str(item.get("url") or "").strip()
            url_status = _source_url_status(enriched.get("_link_validation")) if source_url else "No source URL provided"
            mapping.append([
                enriched.get("ideation_name") or item.get("name"),
                (
                    f"Recommended Product Action: {enriched.get('strategy') or item.get('classification')}"
                    f"; supporting gap reason {item.get('classification') or 'not specified'}"
                    f"; priority {item.get('priority') or 'not specified'}"
                    f"; confidence {item.get('confidence') or 'not specified'}"
                    + (
                        f"; parent opportunity: {item.get('_parent_opportunity')}"
                        if item.get("_parent_opportunity")
                        else ""
                    )
                ),
                _clip_text(" | ".join(str(part) for part in evidence_parts if part)),
                enriched.get("reference_sku_source")
                or "Use Sunco.com/Shopify SKU first when available; otherwise use best-selling item by category revenue.",
                source_url,
                url_status,
            ])
            for cell in mapping[mapping.max_row]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        mapping.column_dimensions["A"].width = 42
        mapping.column_dimensions["B"].width = 48
        mapping.column_dimensions["C"].width = 72
        mapping.column_dimensions["D"].width = 72
        mapping.column_dimensions["E"].width = 64
        mapping.column_dimensions["F"].width = 52

    add_or_replace_audit_sheet(
        workbook,
        [
            ("Project", "sunco-product-opportunity-engine"),
            ("Category", category.run_name),
            ("Owner", category.owner),
            ("Generated", timestamp()),
            ("Source Step 1 workbook", str(selected_gap)),
            ("Step 1 recommendation rows read", str(intake_audit.get("recommendations_rows_read", 0))),
            ("Step 1 recommendation rows converted", str(intake_audit.get("recommendations_rows_selected", 0))),
            ("Step 1 Amazon rows read", str(intake_audit.get("amazon_rows_read", 0))),
            ("Step 1 Amazon rows converted", str(intake_audit.get("amazon_rows_selected", 0))),
            ("Step 1 rows skipped", str(intake_audit.get("recommendations_rows_skipped", 0) + intake_audit.get("amazon_rows_skipped", 0))),
            ("Skipped row detail", _format_skipped_rows(list(intake_audit.get("skipped_rows") or []))),
            ("Step 1 opportunities after dedupe", str(intake_audit.get("rows_after_dedupe", len(rows)))),
            ("Step 2 SKU candidate rows", str(len(candidate_rows[:100]))),
            ("SKU expansion rule", "Default is one PM-kept Step 1 row to one Step 2 SKU concept row. Step 2 splits only when the selected category profile allows the SKU-defining attribute and Step 1 evidence or PM action explicitly names multiple options. It does not force a minimum row count."),
            ("Category profile rule", "Only the selected category profile and its declared inherited profile, if any, guide Step 2 extraction/defaults. Category profiles do not create product ideas without Step 1 evidence or PM action text."),
            ("Category profile mode", _profile_attribute_mode(profile)),
            ("Category SKU split attributes", ", ".join(_profile_sku_split_attributes(profile)) or "None"),
            ("Pack-size rule", "Pack-size and pack-count recommendations are intentionally excluded from Step 2. Use the separate pack-size recommendation workflow for pack-count decisions."),
            ("Selection rule", "PM row deletion is the gate. Step 2 converts remaining usable Step 1 rows, preserving Priority and Confidence as context instead of using them as hard filters."),
            ("Recommended Product Action rule", "Step 2 uses the controlled PM-facing action vocabulary: NPD, Revision, Concept Review, Hold."),
            ("Reference SKU rule", "Sunco.com/Shopify coverage first, canonicalized against the Existing SKU Line Review; if absent, use best-selling item by category revenue."),
            ("MSRP target rule", "When market price samples exist, Target MSRP is based on the 50th-55th percentile of comparable listings and is independent from margin targets."),
            ("URL status rule", "Step 2 checks Step 1 review URLs live when the workbook is generated. 2xx/3xx means verified; 403/429 means blocked by retailer/CDN and needs manual browser review; 404/410 means invalid and should be replaced."),
            ("URL status summary", _url_validation_summary(url_validation_cache)),
            ("Market price sample file", str(market_sample_path) if market_sample_path else "No category market price sample file found."),
            ("Category intelligence audit", format_intelligence_audit(intelligence)),
        ],
        collect_sql_text(paths, category.slug),
    )

    _reset_sheet_views(workbook)
    for col in range(1, IDEATION_MAX_COL + 1):
        ws.cell(3, col).font = Font(bold=True)

    workbook.save(output)
    workbook.close()

    issues = validate_workbook(output, ["Instructions", "Ideations", "Source Mapping", "Run Audit"])
    ok, message = try_excel_com_open_save(output)
    if message:
        issues.append(message if not ok else message)
    return output, issues
