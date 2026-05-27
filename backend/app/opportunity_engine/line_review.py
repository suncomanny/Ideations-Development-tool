from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .categories import Category
from .paths import ProjectPaths
from .source_policy import classify_source, path_has_forbidden_reference, source_policy_text


LINE_REVIEW_SHEET_NAME = "Existing SKU Line Review"
LINE_REVIEW_SQL_FILENAME = "line_review_postgres.sql"

LINE_REVIEW_HEADERS = [
    "Family Part Number",
    "Product Title",
    "Vendor",
    "Vendor Cost",
    "Vendor Cost Source",
    "Amazon Revenue",
    "Shopify Revenue",
    "Total Revenue",
    "Pack Sizes Available",
    "First PO Placed Date",
    "Oldest Legacy Date",
    "Active Status",
    "Product URL",
    "Image URL",
    "Source Notes",
]

LINE_REVIEW_CATEGORY_ALIASES = {
    "bulbs_plus_tubes": ["bulb", "bulbs", "led bulb", "tube", "tubes", "t8", "lamp"],
    "ceiling_fixtures": ["ceiling fixture", "ceiling light", "flush mount", "fixture"],
    "chandeliers": ["chandelier", "chandeliers"],
    "under_cabinet": ["under cabinet", "under-cabinet", "under counter"],
    "tape_rope_light": ["tape light", "rope light", "led neon rope"],
    "panels": ["panel", "panel light", "ceiling panel", "flat panel"],
    "emergency": ["emergency", "exit sign", "exit signs"],
    "striplights": ["striplight", "strip light", "strip lights"],
    "vaportights": ["vapor tight", "vaportight", "vapor tights"],
    "wraparounds": ["wraparound", "wrap around", "utility ceiling"],
    "wall_packs": ["wall pack", "wall packs"],
    "flood_lights": ["flood light", "flood lights"],
    "area_lights": ["area light", "area lights", "shoebox"],
    "canopy": ["canopy light", "canopy"],
    "string_lights": ["string light", "string lights"],
    "commercial_grow_lights": ["commercial grow light", "grow light", "grow lights"],
    "residential_grow_lights": ["residential grow light", "grow light", "grow lights"],
}


@dataclass(frozen=True)
class LineReviewContext:
    category: Category
    rows: list[dict[str, Any]]
    sql_path: Path
    sql_text: str
    source_references: list[str]
    source_notes: list[str]
    rejected_sources: list[str]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def summary(self) -> str:
        if self.rows:
            refs = "; ".join(self.source_references) if self.source_references else "approved DB snapshot"
            return f"{self.row_count} existing SKU family row(s) loaded from approved Postgres/Redshift source(s): {refs}"
        if self.source_references:
            refs = "; ".join(self.source_references)
            return f"Approved Postgres/Redshift line-review snapshot was found, but it returned 0 existing SKU family row(s): {refs}"
        return (
            "No approved Postgres/Redshift line-review snapshot was found for this category. "
            "The sheet is present as a source-policy placeholder and no legacy local Line Review files were used."
        )


def _sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _category_terms(category: Category) -> list[str]:
    values = {
        category.name,
        category.run_name,
        category.slug.replace("_", " "),
        *LINE_REVIEW_CATEGORY_ALIASES.get(category.slug, []),
    }
    terms = []
    for value in values:
        clean = re.sub(r"\s+", " ", str(value or "").strip().lower())
        if clean and clean not in terms:
            terms.append(clean)
    return sorted(terms, key=lambda item: (len(item), item))


def _decoder_prefixes(paths: ProjectPaths, category: Category) -> list[str]:
    path = paths.source_data / "sku_decoder" / "sku_decoder_clean.csv"
    if not path.exists():
        return []
    prefixes: list[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if (row.get("mapped_category_slug") or "").strip() != category.slug:
                continue
            if str(row.get("line_review_match") or "").strip().lower() not in {"1", "true", "yes"}:
                continue
            prefix = (row.get("match_prefix") or "").strip().upper()
            if prefix and prefix not in prefixes:
                prefixes.append(prefix)
    return sorted(prefixes, key=lambda item: (len(item), item))


def _values_cte(values: list[str]) -> str:
    if not values:
        return "SELECT NULL::text WHERE FALSE"
    return "VALUES\n        " + ",\n        ".join(f"({_sql_literal(value)})" for value in values)


def _snapshot_candidates(paths: ProjectPaths, category: Category) -> list[Path]:
    roots = [
        paths.source_data / "postgres_exports" / "line_reviews",
        paths.source_data / "redshift_exports" / "line_reviews",
    ]
    patterns = [
        f"{category.slug}_line_review*.json",
        f"{category.slug}*line_review*.json",
        f"*{category.slug}*line_review*.json",
    ]
    candidates: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        for pattern in patterns:
            candidates.extend(root.glob(pattern))
    return sorted({path.resolve(): path for path in candidates}.values(), key=lambda path: path.stat().st_mtime, reverse=True)


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _payload_rows(payload: Any) -> list[dict[str, Any]]:
    rows = payload.get("rows") if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        return []
    return [row for row in rows if isinstance(row, dict)]


def _source_from_payload(payload: Any, path: Path) -> tuple[str, str]:
    if isinstance(payload, dict):
        source = (
            payload.get("source_system")
            or payload.get("source")
            or payload.get("source_label")
            or payload.get("database")
            or ""
        )
        reference = payload.get("source_reference") or payload.get("generated_from") or str(path)
        return str(source), str(reference)
    return "", str(path)


def _category_matches_payload(payload: Any, category: Category, path: Path) -> bool:
    if not isinstance(payload, dict):
        return category.slug in path.stem.lower()
    values = [
        payload.get("category_slug"),
        payload.get("category"),
        payload.get("category_run_name"),
        payload.get("run_name"),
    ]
    slugs = {re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_") for value in values if value}
    if not slugs:
        return category.slug in path.stem.lower()
    return category.slug in slugs


def _first(row: dict[str, Any], *keys: str) -> Any:
    normalized = {str(key).strip().lower().replace(" ", "_"): value for key, value in row.items()}
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
        value = normalized.get(key.strip().lower().replace(" ", "_"))
        if value not in (None, ""):
            return value
    return None


def _money(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value if item not in (None, ""))
    return str(value)


def normalize_line_review_row(row: dict[str, Any], source_note: str) -> dict[str, Any]:
    amazon_revenue = _money(_first(row, "amazon_revenue", "Amazon Revenue"))
    shopify_revenue = _money(_first(row, "shopify_revenue", "Shopify Revenue"))
    total_revenue = _money(_first(row, "total_revenue", "gross_revenue", "Total Revenue"))
    if total_revenue is None and (amazon_revenue is not None or shopify_revenue is not None):
        total_revenue = (amazon_revenue or 0) + (shopify_revenue or 0)

    return {
        "Family Part Number": _text(_first(row, "family_part_number", "family", "Family Part Number", "sku_family")),
        "Product Title": _text(_first(row, "product_title", "title", "Product Title", "name")),
        "Vendor": _text(_first(row, "vendor", "supplier", "Vendor")),
        "Vendor Cost": _money(_first(row, "vendor_cost", "cost_per_item", "Vendor Cost")),
        "Vendor Cost Source": _text(_first(row, "vendor_cost_source", "Vendor Cost Source", "cost_source")),
        "Amazon Revenue": amazon_revenue,
        "Shopify Revenue": shopify_revenue,
        "Total Revenue": total_revenue,
        "Pack Sizes Available": _text(_first(row, "pack_sizes_available", "pack_sizes", "Pack Sizes Available", "shopify_pack_sizes")),
        "First PO Placed Date": _text(_first(row, "first_po_placed_date", "First PO Placed Date", "first_po_date")),
        "Oldest Legacy Date": _text(_first(row, "oldest_legacy_date", "Oldest Legacy Date", "created", "created_at")),
        "Active Status": _text(_first(row, "active_status", "status", "Active Status")),
        "Product URL": _text(_first(row, "product_url", "shopify_url", "Product URL", "url")),
        "Image URL": _text(_first(row, "image_url", "Image URL")),
        "Source Notes": _text(_first(row, "source_notes", "Source Notes")) or source_note,
    }


def load_line_review_rows(paths: ProjectPaths, category: Category) -> tuple[list[dict[str, Any]], list[str], list[str], list[str]]:
    rows: list[dict[str, Any]] = []
    source_references: list[str] = []
    source_notes: list[str] = []
    rejected_sources: list[str] = []

    for path in _snapshot_candidates(paths, category):
        if path_has_forbidden_reference(path):
            rejected_sources.append(f"{path}: blocked path name.")
            continue
        try:
            payload = _read_json(path)
        except (OSError, json.JSONDecodeError) as exc:
            rejected_sources.append(f"{path}: could not read JSON ({exc}).")
            continue
        if not _category_matches_payload(payload, category, path):
            rejected_sources.append(f"{path}: category metadata did not match {category.slug}.")
            continue
        source, reference = _source_from_payload(payload, path)
        ok, note = classify_source(source, reference)
        if not ok:
            rejected_sources.append(f"{path}: {note}")
            continue
        payload_rows = _payload_rows(payload)
        source_references.append(str(reference))
        if not payload_rows:
            source_notes.append(f"{path}: approved DB source but no rows were present.")
            return rows, source_references, source_notes, rejected_sources
        source_note = f"{note} Source={source}; Reference={reference}"
        rows.extend(normalize_line_review_row(row, source_note) for row in payload_rows)
        return rows, source_references, source_notes, rejected_sources

    if not rows and not source_notes:
        source_notes.append(
            f"No approved line-review snapshot found under {paths.source_data / 'postgres_exports' / 'line_reviews'} "
            f"or {paths.source_data / 'redshift_exports' / 'line_reviews'}."
        )
    return rows, source_references, source_notes, rejected_sources


def build_line_review_sql(
    category: Category,
    paths: ProjectPaths | None = None,
    include_purchase_order_facts: bool = True,
) -> str:
    values = ",\n        ".join(f"({_sql_literal(term)})" for term in _category_terms(category))
    decoder_prefixes = _decoder_prefixes(paths, category) if paths else []
    decoder_values = _values_cte(decoder_prefixes)
    decoder_note = ", ".join(decoder_prefixes) if decoder_prefixes else "none available"
    if include_purchase_order_facts:
        po_facts_cte = """
,
po_facts AS (
    SELECT
        pb.product_id AS product_id,
        MIN(po.created::date) AS first_po_placed_date,
        string_agg(DISTINCT COALESCE(NULLIF(ss.nickname, ''), ss.name), ', ') AS po_vendors,
        MIN(NULLIF(poli.unit_price, 0)) AS po_unit_cost
    FROM product_base pb
    JOIN products_product p ON p.id = pb.product_id
    LEFT JOIN suppliers_supplierproduct sup_prod ON sup_prod.product_id = p.id
    LEFT JOIN suppliers_supplierproductsku sup_sku ON sup_sku.product_id = p.id
    LEFT JOIN purchasing_purchaseorderlineitem poli
        ON poli.supplier_product_id = sup_prod.id
        OR poli.supplier_product_sku_id = sup_sku.id
    LEFT JOIN purchasing_purchaseorder po ON po.id = poli.po_id
    LEFT JOIN suppliers_supplier ss ON ss.id = COALESCE(sup_prod.supplier_id, po.supplier_id, p.supplier_id)
    GROUP BY pb.product_id
)"""
        vendor_expression = "string_agg(DISTINCT COALESCE(NULLIF(pf.po_vendors, ''), NULLIF(pb.shopify_vendor, '')), ', ')"
        vendor_cost_expression = "COALESCE(MIN(pb.product_vendor_cost), MIN(pf.po_unit_cost))"
        vendor_cost_source_expression = """CASE
            WHEN MIN(pb.product_vendor_cost) IS NOT NULL THEN 'products_product.vendor_cost'
            WHEN MIN(pf.po_unit_cost) IS NOT NULL THEN 'purchasing_purchaseorderlineitem.unit_price'
            ELSE 'missing in approved Postgres source'
        END"""
        first_po_expression = "MIN(pf.first_po_placed_date)"
        po_join = "    LEFT JOIN po_facts pf ON pf.product_id = pb.product_id"
        source_note = "Postgres query generated by Step 1 line review module"
    else:
        po_facts_cte = ""
        vendor_expression = "string_agg(DISTINCT NULLIF(pb.shopify_vendor, ''), ', ')"
        vendor_cost_expression = "MIN(pb.product_vendor_cost)"
        vendor_cost_source_expression = """CASE
            WHEN MIN(pb.product_vendor_cost) IS NOT NULL THEN 'products_product.vendor_cost'
            ELSE 'missing in approved Postgres source; fallback refresh skipped PO facts'
        END"""
        first_po_expression = "NULL::date"
        po_join = ""
        source_note = "Postgres MCP fallback query generated by Step 1 line review module; PO facts skipped"
    return f"""-- Existing SKU Line Review for {category.run_name}
-- Source policy: Postgres only. Do not supplement this query with legacy local CSV/workbook exports.
-- Window: trailing 365 days from CURRENT_DATE. Channels: Shopify=12585, Amazon US=11929.
-- SKU decoder product-type prefixes used for category matching: {decoder_note}
-- Purchase order facts included: {str(include_purchase_order_facts).lower()}
WITH params AS (
    SELECT
        CURRENT_DATE - INTERVAL '365 days' AS start_date,
        CURRENT_DATE AS end_date
),
category_terms(term) AS (
    VALUES
        {values}
),
decoder_prefixes(match_prefix) AS (
    {decoder_values}
),
variant_by_product AS (
    SELECT DISTINCT ON (product_id) *
    FROM shopify_productvariantatshopify
    WHERE product_id IS NOT NULL
    ORDER BY product_id, id
),
variant_by_sku AS (
    SELECT DISTINCT ON (UPPER(sku)) *
    FROM shopify_productvariantatshopify
    WHERE sku IS NOT NULL AND sku <> ''
    ORDER BY UPPER(sku), id
),
product_base AS (
    SELECT DISTINCT
        p.id AS product_id,
        UPPER(COALESCE(NULLIF(p.master_sku, ''), NULLIF(svp.sku, ''), NULLIF(svs.sku, ''))) AS sku,
        regexp_replace(UPPER(COALESCE(NULLIF(p.master_sku, ''), NULLIF(svp.sku, ''), NULLIF(svs.sku, ''))), '-[0-9]+PK$', '', 'i') AS family_part_number,
        COALESCE(NULLIF(shp.title, ''), NULLIF(p.name, ''), p.master_sku) AS product_title,
        p.status AS product_status,
        NULLIF(p.vendor_cost, 0) AS product_vendor_cost,
        p.vendor_cost_currency,
        COALESCE((substring(UPPER(COALESCE(NULLIF(p.master_sku, ''), NULLIF(svp.sku, ''), NULLIF(svs.sku, ''))) from '-([0-9]+)PK$'))::int, NULLIF(p.pack_size, 0), 1) AS pack_size,
        pc.name AS product_category_name,
        shp.vendor AS shopify_vendor,
        shp.status AS shopify_status,
        shp.product_type AS shopify_product_type,
        COALESCE(svp.sku, svs.sku) AS shopify_variant_sku,
        COALESCE(svp.price, svs.price) AS shopify_price,
        CASE
            WHEN shp.response_json ? 'handle' THEN 'https://www.sunco.com/products/' || (shp.response_json ->> 'handle')
            ELSE 'https://www.sunco.com/search?q=' || regexp_replace(COALESCE(NULLIF(COALESCE(svp.sku, svs.sku), ''), NULLIF(p.master_sku, '')), '[^A-Za-z0-9_-]+', '+', 'g')
        END AS product_url,
        COALESCE(sm.original_source, product_media.original_source) AS image_url,
        p.created::date AS product_created_date
    FROM products_product p
    LEFT JOIN products_category pc ON pc.id = p.category_id
    LEFT JOIN variant_by_product svp ON svp.product_id = p.id
    LEFT JOIN variant_by_sku svs ON svp.id IS NULL AND UPPER(svs.sku) = UPPER(p.master_sku)
    LEFT JOIN shopify_shopifyproduct shp ON shp.id = COALESCE(svp.shopify_product_id, svs.shopify_product_id)
    LEFT JOIN shopify_shopifymedia sm ON sm.id = COALESCE(svp.shopify_media_id, svs.shopify_media_id)
    LEFT JOIN shopify_shopifymedia product_media
        ON product_media.product_id = shp.id
        AND product_media.position = 1
    WHERE COALESCE(p.master_sku, svp.sku, svs.sku) IS NOT NULL
      AND (
          EXISTS (
              SELECT 1
              FROM category_terms t
              WHERE lower(
                  COALESCE(pc.name, '') || ' ' ||
                  COALESCE(pc.sku, '') || ' ' ||
                  COALESCE(shp.product_type, '') || ' ' ||
                  COALESCE(shp.title, '') || ' ' ||
                  COALESCE(p.name, '') || ' ' ||
                  COALESCE(p.master_sku, '')
              ) LIKE '%' || t.term || '%'
          )
          OR EXISTS (
              SELECT 1
              FROM decoder_prefixes d
              WHERE d.match_prefix IS NOT NULL
                AND left(
                    UPPER(COALESCE(NULLIF(p.master_sku, ''), NULLIF(svp.sku, ''), NULLIF(svs.sku, ''))),
                    length(d.match_prefix)
                ) = d.match_prefix
                AND (
                    length(UPPER(COALESCE(NULLIF(p.master_sku, ''), NULLIF(svp.sku, ''), NULLIF(svs.sku, '')))) = length(d.match_prefix)
                    OR substring(
                        UPPER(COALESCE(NULLIF(p.master_sku, ''), NULLIF(svp.sku, ''), NULLIF(svs.sku, '')))
                        from length(d.match_prefix) + 1 for 1
                    ) IN ('-', '_')
                )
          )
      )
),
product_families AS (
    SELECT DISTINCT family_part_number
    FROM product_base
    WHERE family_part_number IS NOT NULL
),
sales AS (
    SELECT
        regexp_replace(UPPER(i.listing_sku), '-[0-9]+PK$', '', 'i') AS family_part_number,
        ROUND(SUM(CASE WHEN o.sales_channel_id = 11929 THEN i.sales_price * i.quantity_ordered ELSE 0 END)::numeric, 2) AS amazon_revenue,
        ROUND(SUM(CASE WHEN o.sales_channel_id = 12585 THEN i.sales_price * i.quantity_ordered ELSE 0 END)::numeric, 2) AS shopify_revenue,
        ROUND(SUM(i.sales_price * i.quantity_ordered)::numeric, 2) AS total_revenue
    FROM skubana_orderitem i
    JOIN skubana_order o ON i.order_id = o.order_id
    JOIN params p ON o.order_date >= p.start_date AND o.order_date < p.end_date
    JOIN product_families f ON f.family_part_number = regexp_replace(UPPER(i.listing_sku), '-[0-9]+PK$', '', 'i')
    WHERE o.sales_channel_id IN (11929, 12585)
      AND COALESCE(o.order_status, '') NOT IN ('CANCELLED', 'Canceled', 'Cancelled')
    GROUP BY 1
){po_facts_cte},
family_rollup AS (
    SELECT
        pb.family_part_number,
        MIN(pb.product_title) AS product_title,
        {vendor_expression} AS vendor,
        {vendor_cost_expression} AS vendor_cost,
        {vendor_cost_source_expression} AS vendor_cost_source,
        string_agg(DISTINCT pb.pack_size::text, ', ' ORDER BY pb.pack_size::text) AS pack_sizes_available,
        {first_po_expression} AS first_po_placed_date,
        MIN(pb.product_created_date) AS oldest_legacy_date,
        string_agg(DISTINCT COALESCE(pb.shopify_status, pb.product_status), ', ') AS active_status,
        MIN(pb.product_url) AS product_url,
        MIN(pb.image_url) AS image_url
    FROM product_base pb
{po_join}
    GROUP BY pb.family_part_number
)
SELECT
    fr.family_part_number,
    fr.product_title,
    fr.vendor,
    fr.vendor_cost,
    fr.vendor_cost_source,
    COALESCE(s.amazon_revenue, 0) AS amazon_revenue,
    COALESCE(s.shopify_revenue, 0) AS shopify_revenue,
    COALESCE(s.total_revenue, 0) AS total_revenue,
    fr.pack_sizes_available,
    fr.first_po_placed_date,
    fr.oldest_legacy_date,
    fr.active_status,
    fr.product_url,
    fr.image_url,
    '{source_note}' AS source_notes
FROM family_rollup fr
LEFT JOIN sales s ON s.family_part_number = fr.family_part_number
ORDER BY COALESCE(s.total_revenue, 0) DESC, fr.family_part_number;"""


def write_line_review_sql(paths: ProjectPaths, category: Category) -> tuple[Path, str]:
    sql_text = build_line_review_sql(category, paths)
    sql_folder = paths.cache / "ideation_data" / category.slug / "sql"
    sql_folder.mkdir(parents=True, exist_ok=True)
    sql_path = sql_folder / LINE_REVIEW_SQL_FILENAME
    sql_path.write_text(sql_text, encoding="utf-8")
    return sql_path, sql_text


def prepare_line_review_context(paths: ProjectPaths, category: Category) -> LineReviewContext:
    sql_path, sql_text = write_line_review_sql(paths, category)
    rows, source_references, source_notes, rejected_sources = load_line_review_rows(paths, category)
    return LineReviewContext(
        category=category,
        rows=rows,
        sql_path=sql_path,
        sql_text=sql_text,
        source_references=source_references,
        source_notes=source_notes,
        rejected_sources=rejected_sources,
    )


def _create_line_review_sheet(workbook):
    if LINE_REVIEW_SHEET_NAME in workbook.sheetnames:
        del workbook[LINE_REVIEW_SHEET_NAME]
    try:
        index = workbook.sheetnames.index("Amazon Source Audit") + 1
    except ValueError:
        index = len(workbook.sheetnames)
    return workbook.create_sheet(LINE_REVIEW_SHEET_NAME, index)


def _write_headers(ws) -> None:
    for col, header in enumerate(LINE_REVIEW_HEADERS, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(LINE_REVIEW_HEADERS))}1"


def write_line_review_sheet(workbook, context: LineReviewContext) -> None:
    ws = _create_line_review_sheet(workbook)
    _write_headers(ws)

    if context.rows:
        for row_index, row in enumerate(context.rows, start=2):
            for col_index, header in enumerate(LINE_REVIEW_HEADERS, start=1):
                cell = ws.cell(row_index, col_index)
                cell.value = row.get(header)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if header in {"Vendor Cost", "Amazon Revenue", "Shopify Revenue", "Total Revenue"}:
                    cell.number_format = '$#,##0.00;[Red]($#,##0.00);-'
            product_url = row.get("Product URL")
            image_url = row.get("Image URL")
            if product_url:
                ws.cell(row_index, 13).hyperlink = product_url
                ws.cell(row_index, 13).style = "Hyperlink"
            if image_url:
                ws.cell(row_index, 14).hyperlink = image_url
                ws.cell(row_index, 14).style = "Hyperlink"
            ws.row_dimensions[row_index].height = 48
    else:
        ws.cell(2, 1).value = (
            "(approved DB line review snapshot returned no rows)"
            if context.source_references
            else "(no approved DB line review snapshot found)"
        )
        ws.cell(2, 15).value = context.summary()
        for cell in ws[2]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 72

    widths = {
        1: 24,
        2: 42,
        3: 28,
        4: 16,
        5: 30,
        6: 18,
        7: 18,
        8: 18,
        9: 22,
        10: 20,
        11: 18,
        12: 18,
        13: 42,
        14: 42,
        15: 62,
    }
    for col_index, width in widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = width


def source_audit_rows(context: LineReviewContext) -> list[tuple[str, str, str, str, str]]:
    rows = [
        (
            "Line review",
            context.category.run_name,
            "Source policy",
            source_policy_text(),
            "",
        ),
        (
            "Line review",
            context.category.run_name,
            "Generated Postgres SQL",
            str(context.sql_path),
            str(context.sql_path),
        ),
        (
            "Line review",
            context.category.run_name,
            "Existing SKU rows loaded",
            context.summary(),
            "; ".join(context.source_references),
        ),
    ]
    for note in context.source_notes:
        rows.append(("Line review", context.category.run_name, "Source note", note, ""))
    for rejection in context.rejected_sources[:20]:
        rows.append(("Line review", context.category.run_name, "Rejected source", rejection, ""))
    return rows


def run_audit_rows(context: LineReviewContext) -> list[tuple[str, str]]:
    rejected = "\n".join(context.rejected_sources) if context.rejected_sources else "No rejected line-review sources."
    return [
        ("Line review source policy", source_policy_text()),
        ("Line review SQL", str(context.sql_path)),
        ("Line review rows loaded", str(context.row_count)),
        ("Line review source references", "\n".join(context.source_references) or "No approved DB snapshot was loaded."),
        ("Line review notes", "\n".join(context.source_notes) or context.summary()),
        ("Line review rejected sources", rejected),
    ]
